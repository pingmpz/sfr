from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.http import HttpResponse
import requests
from dateutil import parser
import json
# System
from django.conf import settings
import pyodbc
# Date time
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import time
# File 
from sfr.settings import MEDIA_ROOT
from openpyxl import load_workbook, Workbook
import csv
from django.core.files.storage import FileSystemStorage
from pathlib import Path
import glob, os, shutil
import io
# Token
import secrets
import string
# Other 

# EMAIL
# from django.core.mail import EmailMessage
# from sfr.settings import EMAIL_HOST_USER
# import smtplib
# import traceback
# import threading
# from django.template.loader import get_template

import webbrowser
import subprocess
import wget

HOST_URL = 'http://129.1.100.190:8080/'
TEMPLATE_OVERTIME = 'email_templates/overtime.html'
OVER_EST_RATING = 1.5
DEFAULT_ACTIVE_DATE = '1999-01-01'
DEFAULT_INACTIVE_DATE = '2999-01-01'
AB_GRAPH_LIMIT_DAY = 3

#------------------------------------------------------------------------ EMAIL

################################################################################
##################################### PAGES ####################################
################################################################################

def blank(request):
    update_employee_master()
    context = {
    }
    return render(request, 'blank.html', context)

def first_page(request):
    return redirect('/transaction/0')

def index(request):
    return redirect('/transaction/0')

#------------------------------------------------------------------- TRANSACTION

def transaction(request, orderoprno):
    start_time = time.time()
    #CONST
    ip_address = getClientIP(request)
    # ip_address = '129.1.114.149'
    empAtComputerList = getEmpAtComputerList(ip_address)
    overtimehour = 0
    canMP = False
    refreshSecond = 300
    orderNo = ""
    operationNo = ""
    order = None
    operation = None
    isFirst = False
    isOperating = False
    isOvertime = False
    lastCanceledOrder = None
    hasReportTime = False
    remainQty = -1
    state = "ERROR" #-- FIRSTPAGE / NODATAFOUND / NOOPERATIONFOUND / DATAFOUND
    cancelInfo = None
    #-- Left Content List
    operationList = []
    operationStatusList = []
    modList = []
    overTimeOperatorList = []
    overTimeWorkCenterList = []
    joinList = []
    bomList = []
    #-- History
    historyOperateList = []
    historyConfirmList = []
    historyJoinList = []
    #-- ETC
    rejectReasonList = [] #-- All
    materialGroupList = [] #-- All
    purchaseGroupList = [] #-- All
    currencyList = [] #-- All
    #-- IDLE
    idleList = []
    total_idle_time = 0
    #-- OVER EST TIME
    actual_setup_time = 0
    actual_oper_time = 0
    actual_labor_time = 0
    actual_idle_time = 0
    ai_time = 0
    est_setup_time_sum = 0
    est_oper_time_sum = 0
    est_labor_time_sum = 0
    setup_time_percent = -1
    oper_time_percent = -1
    labor_time_percent = -1
    ai_time_percent = -1
    is_over_est_time = False

    currentOperation = -1
    operationBefore = -1 #-- For Prev Operation Button
    operationAfter = -1 #-- For Next Operation Button

    #-- TOOL LIST
    historyToolList = []
    historyToolLifeQtyList = []
    historyToolLifeMinList = []
    hasToolConfirmedQtyList = []
    toolHeaderList = []
    toolStepLists = []
    toolItemLists = []
    toolLifeQtyLists = []
    toolLifeMinLists = []

    idleTypeList = []
    # print(f"Init, excute time : {time.time() - start_time}")
    #--
    if orderoprno == "0":
        state = "FIRSTPAGE"
    else:
        orderNo = orderoprno[0:len(orderoprno) - 4]
        operationNo = orderoprno[len(orderoprno) - 4:len(orderoprno)]
        if isExistOrder(orderNo) == False and isExistSAPOrder(orderNo) == False:
            state = "NODATAFOUND"
        else:
            state = "NOOPERATIONFOUND"
            if isExistOrder(orderNo) == False:
                setDataFromSAP(orderNo)
            order = getOrder(orderNo)
            operationList = getOperationList(orderNo)
            #-- GET OPERATION WITH REMAINING QTY > 0
            if len(operationList) > 0:
                currentOperation = operationList[0].OperationNo
                for i in range(len(operationList)):
                    tempRemainQty = operationList[i].ProcessQty - (operationList[i].AcceptedQty + operationList[i].RejectedQty)
                    if tempRemainQty > 0:
                        currentOperation = operationList[i].OperationNo
                        break
                #-- IF NO OPERATION INPUT (0000), WILL AUTO REDIRECT TO CURRENT PAGE
                if operationNo == '0000':
                    return redirect('/transaction/' + orderNo + currentOperation)
            #-- Check Cancel Order
            if isCanceledOrder(orderNo):
                 state = "CANCEL"
                 cancelInfo = getCanceledOrderInfo(orderNo)
            elif isExistOperation(orderNo, operationNo):
                state = "DATAFOUND"
                #-- CONST
                overtimehour = getOvertimeHour()
                canMP = True if getManualReportAllow() or not order.ProcessStop or (order.ProcessStop + timedelta(days=3) > datetime.today()) else False
                refreshSecond = getRefreshSecond()
                #-- OPRATION DETIAL
                operation = getOperation(orderNo, operationNo)
                isFirst = isFirstOperation(orderNo, operationNo)
                isOperating = isOperatingOperation(orderNo, operationNo)
                isOvertime = isOvertimeOperation(orderNo, operationNo)
                lastCanceledOrder = getLastCanceledOrder()
                # hasReportTime = hasReportTimeToSAP(orderNo, operationNo)
                remainQty = operation.ProcessQty - (operation.AcceptedQty + operation.RejectedQty)
                #-- CHECK CLOSED
                hasNoMoreQty = True
                for i in range(len(operationList)):
                    tempRemainQty = operationList[i].ProcessQty - (operationList[i].AcceptedQty + operationList[i].RejectedQty)
                    if tempRemainQty > 0:
                        hasNoMoreQty = False
                        break
                #-- IF CLOSED DONT SHOW CURRENT OPERATION
                if hasNoMoreQty:
                    currentOperation = -1
                #-- SET STATUS OF EACH OPERATION IN LIST
                # isPartial = False
                # for i in range(len(operationList)):
                #     tempRemainQty = operationList[i].ProcessQty - (operationList[i].AcceptedQty + operationList[i].RejectedQty)
                #     if isPartial == False and tempRemainQty > 0:
                #         isPartial = True
                #     temphasReportTime = hasReportTimeToSAP(orderNo, operationList[i].OperationNo)
                #     if operationList[i].ProcessQty == 0 and hasNoMoreQty:
                #         operationStatusList.append("REJECTED")
                #     elif operationList[i].ProcessQty == 0:
                #         operationStatusList.append("WAITING")
                #     elif operationList[i].JoinToOrderNo != None and operationList[i].JoinToOperationNo != None:
                #         operationStatusList.append("JOINING")
                #     elif isOvertimeOperation(orderNo, operationList[i].OperationNo):
                #         operationStatusList.append("OVERTIME")
                #     elif tempRemainQty > 0 and operationList[i].ProcessStart != None:
                #         operationStatusList.append("WORKING")
                #     elif tempRemainQty > 0 and operationList[i].ProcessStart == None:
                #         operationStatusList.append("READY")
                #     elif tempRemainQty == 0 and isPartial == False and temphasReportTime:
                #         operationStatusList.append("COMPLETE")
                #     elif tempRemainQty == 0 and isPartial == False:
                #         operationStatusList.append("COMPLETE, NO WORKING TIME REPORT")
                #     elif tempRemainQty == 0 and isPartial == True:
                #         operationStatusList.append("PARTIALCOMPLETE")
                #     else:
                #         operationStatusList.append("ERROR")
                #     #-- GET PREV & NEXT OPERATION
                #     if operationNo == operationList[i].OperationNo.strip():
                #         if i != 0:
                #             operationBefore = operationList[i-1].OperationNo
                #         if i != len(operationList) - 1:
                #             operationAfter = operationList[i+1].OperationNo
                #-- BOM TAB
                bomList = getBomList(orderNo)
                # print(f"Neccessory Data, excute time : {time.time() - start_time}")
                #-- HISTORY TAB
                historyOperateList = getHistoryOperateList(orderNo, operationNo)
                historyConfirmList = getHistoryConfirmList(orderNo, operationNo)
                historyJoinList = getHistoryJoinList(orderNo, operationNo)
                modList = getModList(orderNo)
                # print(f"History, excute time : {time.time() - start_time}")
                #-- OVERTIME TAB
                overTimeOperatorList = getOverTimeOperatorList(orderNo, operationNo)
                overTimeWorkCenterList = getOverTimeWorkCenterList(orderNo, operationNo)
                # print(f"Overtime, excute time : {time.time() - start_time}")
                #-- JOIN LIST
                if operation.JoinToOrderNo == None and operation.JoinToOperationNo == None:
                    joinList = getJoinList(orderNo, operationNo)
                # print(f"Join, excute time : {time.time() - start_time}")
                #-- ETC LIST
                rejectReasonList = getRejectReasonList()
                materialGroupList = getMaterialGroupList()
                purchaseGroupList = getPurchaseGroupList()
                idleTypeList = getIdleTypeList()
                currencyList = getCurrencyList()
                # print(f"ETC, excute time : {time.time() - start_time}")
                #-- IDLE LIST
                idleList = getIdleList(orderNo, operationNo) 
                xIdle = getTotalIdleTime(orderNo, operationNo)  
                total_idle_time = int(xIdle.Idle) if xIdle and xIdle.Idle else 0
                # print(f"Idle, excute time : {time.time() - start_time}")
                #-- OVER EST TIME
                est_setup_time_sum = float(operation.EstSetupTime) * int(operation.ProcessQty)
                est_oper_time_sum = float(operation.EstOperationTime) * int(operation.ProcessQty)
                est_labor_time_sum = float(operation.EstLaborTime) * int(operation.ProcessQty)
                actual_time = getActualTime(orderNo, operationNo)
                if actual_time:
                    actual_setup_time = int(actual_time.Setup)
                    actual_oper_time = int(actual_time.Oper)
                    actual_labor_time = int(actual_time.Labor)
                    actual_idle_time = int(actual_time.Idle)
                    ai_time = int(total_idle_time + actual_oper_time)
                    setup_time_percent = -1 if est_setup_time_sum == 0 else int(actual_setup_time/est_setup_time_sum * 100)
                    oper_time_percent = -1 if est_oper_time_sum == 0 else int(actual_oper_time/est_oper_time_sum * 100)
                    labor_time_percent = -1 if est_labor_time_sum == 0 else int(actual_labor_time/est_labor_time_sum * 100)
                    ai_time_percent = -1 if est_oper_time_sum == 0 else int((actual_oper_time + total_idle_time)/est_oper_time_sum * 100)
                if (actual_setup_time > est_setup_time_sum and est_setup_time_sum == 0) or (actual_oper_time > est_oper_time_sum and est_oper_time_sum == 0) or (actual_labor_time > est_labor_time_sum and est_labor_time_sum == 0):
                    is_over_est_time = True
                elif (actual_setup_time > est_setup_time_sum * OVER_EST_RATING) or (actual_oper_time > est_oper_time_sum * OVER_EST_RATING) or (actual_labor_time > est_labor_time_sum * OVER_EST_RATING) or (ai_time > est_oper_time_sum * OVER_EST_RATING):
                    is_over_est_time = True
                # print(f"Over Estimate, excute time : {time.time() - start_time}")
                #-- TOOL LIST
                historyToolList = getHistoryToolList(orderNo, operationNo)
                for tooli in historyToolList:
                    historyToolLifeQty = round((tooli.ConfirmedQty / tooli.ToolLifeQty) * 100, 2) if tooli.ToolLifeQty > 0 else 'NA'
                    historyToolLifeMin = round(tooli.ConfirmedQty * (tooli.ToolLifeMin / tooli.ToolLifeQty), 2) if tooli.ToolLifeQty > 0 else 'NA'
                    historyToolLifeQtyList.append(historyToolLifeQty)
                    historyToolLifeMinList.append(historyToolLifeMin)
                toolHeaderList = getToolHeaderListByOrder(orderNo, operationNo)
                for toolh in toolHeaderList:
                    hasToolConfirmedQty = False
                    toolStepLists.append(getToolStepList(toolh.ID))
                    toolItemList = getToolItemList(toolh.ID)
                    toolItemLists.append(toolItemList)
                    toolLifeQtyList = []
                    toolLifeMinList = []
                    for tooli in toolItemList:
                        if tooli.ConfirmedQty > 0:
                            hasToolConfirmedQty = True
                        toolLifeQty = round((tooli.ConfirmedQty / tooli.ToolLifeQty) * 100, 2) if tooli.ToolLifeQty > 0 else 'NA'
                        toolLifeMin = round(tooli.ConfirmedQty * (tooli.ToolLifeMin / tooli.ToolLifeQty), 2) if tooli.ToolLifeQty > 0 else 'NA'
                        toolLifeQtyList.append(toolLifeQty)
                        toolLifeMinList.append(toolLifeMin)
                    toolLifeQtyLists.append(toolLifeQtyList)
                    toolLifeMinLists.append(toolLifeMinList)
                    hasToolConfirmedQtyList.append(hasToolConfirmedQty)
                # print(f"Tool, excute time : {time.time() - start_time}")
    printString(f"{orderNo}-{operationNo} ({state}) excute time : {time.time() - start_time}")
    context = {
        'empAtComputerList' : empAtComputerList,
        'ip_address' : ip_address,
        'overtimehour' : overtimehour,
        'canMP' : canMP,
        'refreshSecond' : refreshSecond,
        'orderNo' : orderNo,
        'operationNo' : operationNo,
        'state' : state,
        'cancelInfo': cancelInfo,
        'order' : order,
        'operation' : operation,
        'isFirst' : isFirst,
        'isOperating' : isOperating,
        'isOvertime' : isOvertime,
        'lastCanceledOrder' : lastCanceledOrder,
        'hasReportTime' : hasReportTime,
        'remainQty' : remainQty,
        'operationList' : operationList,
        'operationStatusList' : operationStatusList,
        'modList' : modList,
        'overTimeOperatorList' : overTimeOperatorList,
        'overTimeWorkCenterList' : overTimeWorkCenterList,
        'joinList' : joinList,
        'bomList' : bomList,
        'historyOperateList' : historyOperateList,
        'historyConfirmList' : historyConfirmList,
        'historyJoinList' : historyJoinList,
        'rejectReasonList' : rejectReasonList,
        'materialGroupList' : materialGroupList,
        'purchaseGroupList' : purchaseGroupList,
        'idleTypeList' : idleTypeList,
        'currencyList' : currencyList,
        'actual_setup_time': actual_setup_time,
        'actual_oper_time': actual_oper_time,
        'actual_labor_time': actual_labor_time,
        'actual_idle_time': actual_idle_time,
        'ai_time': ai_time,
        'est_setup_time_sum': est_setup_time_sum,
        'est_oper_time_sum': est_oper_time_sum,
        'est_labor_time_sum': est_labor_time_sum,
        'setup_time_percent': setup_time_percent,
        'oper_time_percent': oper_time_percent,
        'labor_time_percent': labor_time_percent,
        'ai_time_percent': ai_time_percent,
        'is_over_est_time': is_over_est_time,
        'idleList': idleList,
        'total_idle_time': total_idle_time,
        'currentOperation' : currentOperation,
        'operationBefore' : operationBefore,
        'operationAfter' : operationAfter,
        'historyToolList': historyToolList,
        'historyToolLifeQtyList': historyToolLifeQtyList,
        'historyToolLifeMinList': historyToolLifeMinList,
        'hasToolConfirmedQtyList': hasToolConfirmedQtyList,
        'toolHeaderList': toolHeaderList,
        'toolStepLists': toolStepLists,
        'toolItemLists': toolItemLists,
        'toolLifeQtyLists': toolLifeQtyLists,
        'toolLifeMinLists': toolLifeMinLists,
    }
    return render(request, 'transaction.html', context)

def join_activity(request, orderoprno):
    orderNo = orderoprno[0:len(orderoprno) - 4]
    operationNo = orderoprno[len(orderoprno) - 4:len(orderoprno)]
    order = getOrder(orderNo)
    operation = getOperation(orderNo, operationNo)
    joinableList = getJoinableList(orderNo, operation.WorkCenterGroup)
    context = {
        'orderNo' : orderNo,
        'operationNo' : operationNo,
        'operation' : operation,
        'joinableList' : joinableList,
    }
    return render(request, 'join_activity.html', context)

def lot_traveller(request, orderno, lotno):
    orderNo = ""
    LotNo = ""
    order = None
    state = ""
    pltList = []
    plt = None
    operationList = []
    operationList1 = []
    operationList2 = []
    operationList3 = []
    planDayCountList1 = []
    planDayCountList2 = []
    planDayCountList3 = []
    pageCount = 1
    maxRows = 13
    counter = 0
    #--
    if orderno == "0":
        state = "FIRSTPAGE"
    else:
        orderNo = orderno
        LotNo = lotno
        if isExistOrder(orderNo) == False:
            state = "NODATAFOUND"
        else:
            state = "DATAFOUND"
            order = getOrder(orderNo)
            operationList = getOperationList(orderNo)
            pltList = getPTLList(orderNo)
            #-- IF COME FROM FIRST PAGE
            if LotNo == "-1":
                #-- IF NEVER DO PLT GO TO 0
                if pltList == []:
                    return redirect('/lot_traveller/' + orderNo + "&0")
                #-- IF EVER DO PLT GO TO LASTEST PLT
                else:
                    return redirect('/lot_traveller/' + orderNo + "&" + str(pltList[0].LotNo))
            #-- IF TRY TO GO TO 0 AFTER EVER DONE PLT, WILL KICK TO LASTEST PLT
            elif LotNo == "0":
                if pltList != []:
                    return redirect('/lot_traveller/' + orderNo + "&" + str(pltList[0].LotNo))
            else:
                #-- IF TRY TO GO TO # AFTER NEVER DONE PLT, WILL KICK TO 0
                if pltList == []:
                    return redirect('/lot_traveller/' + orderNo + "&0")
                plt = getPTL(orderNo, LotNo)
                for opr in operationList:
                    if opr.OperationNo >= plt.StartOperationNo:
                        if counter == maxRows:
                            pageCount += 1
                            counter = 0
                        start_date = datetime.strptime(opr.PlanStartDate, '%Y-%m-%d')
                        stop_date = datetime.strptime(opr.PlanFinishDate, '%Y-%m-%d')
                        days = stop_date - start_date
                        if pageCount == 1:
                            operationList1.append(opr)
                            planDayCountList1.append(days.days)
                        elif pageCount == 2:
                            operationList2.append(opr)
                            planDayCountList2.append(days.days)
                        else:
                            operationList3.append(opr)
                            planDayCountList3.append(days.days)
                        counter += 1
    context = {
        'orderNo' : orderNo,
        'LotNo' : LotNo,
        'order' : order,
        'state' : state,
        'pltList' : pltList,
        'plt' : plt,
        'operationList' : operationList,
        'operationList1' : operationList1,
        'operationList2' : operationList2,
        'operationList3' : operationList3,
        'planDayCountList1': planDayCountList1,
        'planDayCountList2': planDayCountList2,
        'planDayCountList3': planDayCountList3,
        'pageCount': pageCount,
    }
    return render(request, 'lot_traveller.html', context)

def ab_graph_mn(request, frt, fdate):
    onRoutingList = getOnRoutingList()
    if frt == "FIRST":
        frt = onRoutingList[0].WorkCenterNo
    if fdate == "NOW":
        fdate = datetime.today().strftime('%Y-%m-%d')
    act_date = datetime.strptime(fdate, '%Y-%m-%d')
    ndate = (act_date + timedelta(days=(1))).strftime('%Y-%m-%d')
    is_able_to_edit = ((datetime.today() - act_date <= timedelta(days=(AB_GRAPH_LIMIT_DAY))) and datetime.today() > act_date)
    is_today = True if act_date.date() == datetime.today().date() else False
    mcs = getMachineOnRoutingIsActiveList(frt)
    states, hrs, mins, nhrs, nmins, rhours = [], [], [], [], [], []
    states_changables = [False] * 24 if is_today else [True] * 24
    if is_today:
        for idx, i in enumerate(states_changables):
            hr_now = int(datetime.today().strftime('%H'))
            if idx <= hr_now:
                states_changables[idx] = True
    for mc in mcs: 
        state = ['N'] * 24
        hr = ""
        min = ""
        nhr = ""
        nmin = ""
        rhour = ""
        wc = mc.WorkCenterNo
        st = getABGraphManualData(wc, fdate)
        if st:
            state[0] = st.H00
            state[1] = st.H01
            state[2] = st.H02
            state[3] = st.H03
            state[4] = st.H04
            state[5] = st.H05
            state[6] = st.H06
            state[7] = st.H07
            state[8] = st.H08
            state[9] = st.H09
            state[10] = st.H10
            state[11] = st.H11
            state[12] = st.H12
            state[13] = st.H13
            state[14] = st.H14
            state[15] = st.H15
            state[16] = st.H16
            state[17] = st.H17
            state[18] = st.H18
            state[19] = st.H19
            state[20] = st.H20
            state[21] = st.H21
            state[22] = st.H22
            state[23] = st.H23
            hr = st.Hour
            min = st.Min
        nst = getABGraphManualData(wc, ndate)
        if nst:
            nhr = nst.Hour
            nmin = nst.Min
            if hr != "" and min != "" and nhr != "" and nmin != "":
                xhr = nhr
                xmin = nmin
                if int(nmin) < int(min):
                    xhr = int(xhr) - 1
                    xmin = int(xmin) + 60
                rmin = int(xmin) - int(min)
                rhour = int(xhr) - int(hr)
                if rmin > 30:
                    rhour = rhour + 1
        states.append(state)
        hrs.append(hr)
        mins.append(min)
        nhrs.append(nhr)
        nmins.append(nmin)
        rhours.append(rhour)
    context = {
        'onRoutingList': onRoutingList,
        'frt': frt,
        'fdate': fdate,
        'is_able_to_edit': is_able_to_edit,
        'is_today': is_today,
        'states_changables': states_changables,
        'mcs': mcs,
        'states': states,
        'hrs': hrs,
        'mins': mins,
        'ndate': ndate,
        'nhrs': nhrs,
        'nmins': nmins,
        'rhours': rhours,
    }
    return render(request, 'ab_graph_mn.html', context)

def tool_store(request):
    toolHeaderList = getToolHeaderAll()
    context = {
        'toolHeaderList': toolHeaderList,
    }
    return render(request, 'tool_store.html', context)    

#------------------------------------------------------------------------ MASTER

def master_wc(request):
    workCenterList = getWorkCenterList()
    context = {
        'workCenterList' : workCenterList,
    }
    return render(request, 'master/wc.html', context)

def master_emp(request):
    operatorList = getOperatorList()
    context = {
        'operatorList' : operatorList,
    }
    return render(request, 'master/emp.html', context)

def master_rej(request):
    rejectReasonList = getRejectReasonList()
    context = {
        'rejectReasonList' : rejectReasonList,
    }
    return render(request, 'master/rej.html', context)

def master_matg(request):
    materialGroupList = getMaterialGroupList()
    context = {
        'materialGroupList' : materialGroupList,
    }
    return render(request, 'master/matg.html', context)

def master_purg(request):
    purchaseGroupList = getPurchaseGroupList()
    context = {
        'purchaseGroupList' : purchaseGroupList,
    }
    return render(request, 'master/purg.html', context)

def master_curr(request):
    currencyList = getCurrencyList()
    context = {
        'currencyList' : currencyList,
    }
    return render(request, 'master/curr.html', context)

def master_tool(request):
    toolMasterList = getToolMasterList()
    context = {
        'toolMasterList' : toolMasterList,
    }
    return render(request, 'master/tool.html', context)

def master_ctt(request):
    cycleTimeTargetList = getCycleTimeTargetList()
    context = {
        'cycleTimeTargetList' : cycleTimeTargetList,
    }
    return render(request, 'master/ctt.html', context)

#--------------------------------------------------------------------- DATA PAGE

def wc(request, wcno, fmonth):
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    workCenter = getWorkCenter(wcno)
    historyTransList = getWorkCenterHistoryTransactionList(wcno, fmonth)
    historyOvertimeList = getWorkCenterHistoryOvertimeList(wcno, fmonth)
    #STATISTIC
    totalSetupTime = 0
    totalOperTime = 0
    totalWorkingTime = 0
    for trans in historyTransList:
        totalSetupTime = totalSetupTime + trans.Setup
        totalOperTime = totalOperTime + trans.Oper
    totalWorkingTime = totalSetupTime + totalOperTime
    totalSetupTime = str(int(totalSetupTime/60)) + " Hours " + str(int(totalSetupTime%60)) + " Minutes"
    totalOperTime = str(int(totalOperTime/60)) + " Hours " + str(int(totalOperTime%60)) + " Minutes"
    totalWorkingTime = str(int(totalWorkingTime/60)) + " Hours " + str(int(totalWorkingTime%60)) + " Minutes"
    context = {
        'workCenter': workCenter,
        'fmonth': fmonth,
        'historyTransList': historyTransList,
        'historyOvertimeList': historyOvertimeList,
        'totalSetupTime': totalSetupTime,
        'totalOperTime': totalOperTime,
        'totalWorkingTime': totalWorkingTime,
    }
    return render(request, 'wc.html', context)

def emp(request, empid, fmonth):
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    employee = getOperator(empid)
    historyTransList = getEmployeeHistoryTransactionList(empid, fmonth)
    historyConfirmList = getEmployeeHistoryConfirmList(empid, fmonth)
    historyOvertimeList = getEmployeeHistoryOvertimeList(empid, fmonth)
    #STATISTIC
    totalSetupTime = 0
    totalLaborTime = 0
    totalWorkingTime = 0
    for trans in historyTransList:
        totalSetupTime = totalSetupTime + trans.Setup
        totalLaborTime = totalLaborTime + trans.Labor
    totalWorkingTime = totalSetupTime + totalLaborTime
    totalSetupTime = str(int(totalSetupTime/60)) + " Hours " + str(int(totalSetupTime%60)) + " Minutes"
    totalLaborTime = str(int(totalLaborTime/60)) + " Hours " + str(int(totalLaborTime%60)) + " Minutes"
    totalWorkingTime = str(int(totalWorkingTime/60)) + " Hours " + str(int(totalWorkingTime%60)) + " Minutes"
    context = {
        'employee': employee,
        'fmonth': fmonth,
        'historyTransList': historyTransList,
        'historyConfirmList': historyConfirmList,
        'historyOvertimeList': historyOvertimeList,
        'totalSetupTime': totalSetupTime,
        'totalLaborTime': totalLaborTime,
        'totalWorkingTime': totalWorkingTime,
    }
    return render(request, 'emp.html', context)

#-------------------------------------------------------------------- MONITORING

def working_order(request):
    workingOrderList = getWorkingOrderList()
    context = {
        'workingOrderList': workingOrderList,
    }
    return render(request, 'monitoring/working_order.html', context)

def working_wc(request):
    overtimehour = getOvertimeHour()
    warninghour = overtimehour - 2
    workingWorkCenterList = getWorkingWorkCenterList()
    context = {
        'overtimehour': overtimehour,
        'warninghour': warninghour,
        'workingWorkCenterList': workingWorkCenterList,
    }
    return render(request, 'monitoring/working_wc.html', context)

def working_emp(request):
    # send_email_overtime()
    overtimehour = getOvertimeHour()
    warninghour = overtimehour - 2
    workingOperatorList = getWorkingOperatorList()
    context = {
        'overtimehour': overtimehour,
        'warninghour': warninghour,
        'workingOperatorList': workingOperatorList,
    }
    return render(request, 'monitoring/working_emp.html', context)

def delay_operation(request, fwc):
    # profitCenterList = getProfitCenterList()
    # workCenterGroupList = getWorkCenterGroupList()
    workCenterList = getWorkCenterRoutingList()
    if fwc == "FIRST":
        fwc = workCenterList[0].WorkCenterNo
    # if fwcg == "FIRST":
    #     fwcg = profitCenterList[0].WorkCenterNo
    # if fwc == "FIRST":
    #     fpfc = workCenterList[0].WorkCenterNo
    SAPDelayOperationList = getSAPDelayOperationList(fwc)
    SFRDelayOperationList = getSFRDelayOperationList(fwc)
    SFRDelayWorkActualList = []
    for op in SFRDelayOperationList:
        prev_op = getPreviousOperation(op.OrderNo, op.OperationNo)
        if prev_op != None:
            first_con = getFirstConfirmTime(prev_op.OrderNo, prev_op.OperationNo)
            if first_con != None:
                SFRDelayWorkActualList.append(str((datetime.today() - first_con.ConfirmDateTime).days))
            else:
                SFRDelayWorkActualList.append('Error')
        else:
            release_date = datetime.strptime(getOrder(op.OrderNo).ReleaseDate, '%Y-%m-%d')
            SFRDelayWorkActualList.append(str((datetime.today() - release_date).days))
    delay_list_len = len(SAPDelayOperationList) + len(SFRDelayOperationList)
    context = {
        'fwc': fwc,
        # 'profitCenterList': profitCenterList,
        # 'workCenterGroupList': workCenterGroupList,
        'workCenterList': workCenterList,
        'SAPDelayOperationList': SAPDelayOperationList,
        'SFRDelayOperationList': SFRDelayOperationList,
        'SFRDelayWorkActualList': SFRDelayWorkActualList,
        'delay_list_len': delay_list_len,
    }
    return render(request, 'monitoring/delay_operation.html', context)

def none_working_wc(request):
    workCenterList = getNoneWorkingWorkCenterList()
    context = {
        'workCenterList' : workCenterList,
    }
    return render(request, 'monitoring/none_working_wc.html', context)

def none_start_order(request):
    noneStartOrderList = getNoneStartOrderList()
    req_dates = []
    for ord in noneStartOrderList:
        if ord.RequestDate and ord.RequestDate != '00.00.0000' and ord.RequestDate != '' and ord.RequestDate != ' ':
            req_dates.append(datetime.strptime(ord.RequestDate, '%d.%m.%Y'))
        else:
            req_dates.append(None)
    context = {
        'noneStartOrderList': noneStartOrderList,
        'req_dates': req_dates,
    }
    return render(request, 'monitoring/none_start_order.html', context)

def pending_pln_fai(request):
    pendingPLNFAIList = getPendingPLNFAIList()
    context = {
        'pendingPLNFAIList': pendingPLNFAIList,
    }
    return render(request, 'monitoring/pending_pln_fai.html', context)

#------------------------------------------------------------------------ REPORT

def ot_table(request, fmonth):
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    overtimeOperatorList = getOvertimeOperatorList(fmonth)
    overtimeWorkCenterList = getOvertimeWorkCenterList(fmonth)
    overtimeCounterList = getOvertimeCounterList(fmonth)
    context = {
        'fmonth': fmonth,
        'overtimeOperatorList': overtimeOperatorList,
        'overtimeWorkCenterList': overtimeWorkCenterList,
        'overtimeCounterList': overtimeCounterList,
    }
    return render(request, 'report/ot_table.html', context)

def mp_ot_auto_machine(request, fmonth):
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    printString(fmonth)
    ReportList = getAutoMachineManualReportOvertimeList(fmonth)
    context = {
        'fmonth': fmonth,
        'ReportList': ReportList,
    }
    return render(request, 'report/mp_ot_auto_machine.html', context)

def oper_no_time(request, fmonth):
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    operationNoTimeList = getOperationNoTimeList(fmonth)
    context = {
        'fmonth': fmonth,
        'operationNoTimeList': operationNoTimeList,
    }
    return render(request, 'report/oper_no_time.html', context)

def emp_work_time(request, fstartdate, fstopdate, emp_type):
    if fstartdate == "NOW":
        fstartdate = datetime.today().strftime('%Y-%m-%d')
    if fstopdate == "NOW":
        fstopdate = datetime.today().strftime('%Y-%m-%d')
    if datetime.strptime(fstopdate, '%Y-%m-%d') < datetime.strptime(fstartdate, '%Y-%m-%d'):
        fstartdate = fstopdate
    fmonth = fstartdate[0:7]
    empWorkTimeList = getEmpWorkTimeList(fstartdate, fstopdate, emp_type)
    cmsTimeList = []
    hrTimeList = []
    percentList = []
    gradeList = []
    opthrTimeList = []
    optpercentList = []
    optgradeList = []
    for rec in empWorkTimeList:
        cmswt = 0
        # try: 
        #     response = requests.get('http://129.1.100.185:8200/api/get_emp_work_time/%s&%s&%s' % (rec.EmpID.strip(), fstartdate, fstopdate))
        #     data = json.loads(response.text)
        #     cmswt = data['work_time']
        # except Exception as e:
        #     print(f"Can not connect to CMS : {e}")
        cmsTimeList.append(cmswt)

        wt = 0
        optwt = 0
        is_leave = False
        try:
            hr = getTotalEmpWorkTimeFromHRFocus(rec.EmpID, fstartdate, fstopdate)
            wt = hr.TotalTime
            optwt = hr.OptTime
        except Exception as e:
            print(f'Can not connect to HR Focus SQL Server: {e}')
        percentage = '-'
        grade = '-'
        if wt != 0:
            percent = round((rec.Total / wt) * 100, 2)
            percentage = str(percent) + '%'
            if percent >= 81:
                grade = 'A'
            elif percent >= 61:
                grade = 'B'
            elif percent >= 26:
                grade = 'C'
            else:
                grade = 'D'
        hrTimeList.append(wt)
        percentList.append(percentage)
        gradeList.append(grade)
        optpercentage = '-'
        optgrade = '-'
        if optwt != 0:
            percent = round((rec.Total / optwt) * 100, 2)
            optpercentage = str(percent) + '%'
            if percent >= 81:
                optgrade = 'A'
            elif percent >= 61:
                optgrade = 'B'
            elif percent >= 26:
                optgrade = 'C'
            else:
                optgrade = 'D'
        opthrTimeList.append(optwt)
        optpercentList.append(optpercentage)
        optgradeList.append(optgrade)
    context = {
        'fstartdate': fstartdate,
        'fstopdate': fstopdate,
        'emp_type': emp_type,
        'fmonth': fmonth,
        'empWorkTimeList': empWorkTimeList,
        'cmsTimeList' : cmsTimeList,
        'hrTimeList': hrTimeList,
        'percentList': percentList,
        'gradeList': gradeList,
        'opthrTimeList': opthrTimeList,
        'optpercentList': optpercentList,
        'optgradeList': optgradeList,
    }
    return render(request, 'report/emp_work_time.html', context)

def completed_order(request, ftype, fdate, fmonth, fstartdate, fstopdate):
    if fdate == "NOW":
        fdate = datetime.today().strftime('%Y-%m-%d')
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    if fstartdate == "NOW":
        fstartdate = datetime.today().strftime('%Y-%m-%d')
    if fstopdate == "NOW":
        fstopdate = datetime.today().strftime('%Y-%m-%d')
    completedOrderList = getCompletedOrderList(ftype, fdate, fmonth, fstartdate, fstopdate)
    process_qtys = []
    accepted_qtys = []
    for ord in completedOrderList:
        process_qtys.append(getFirstOperation(ord.OrderNo).ProcessQty)
        accepted_qtys.append(getLastOperation(ord.OrderNo).AcceptedQty)
    context = {
        'ftype': ftype,
        'fdate': fdate,
        'fmonth': fmonth,
        'fstartdate': fstartdate,
        'fstopdate': fstopdate,
        'completedOrderList': completedOrderList,
        'process_qtys': process_qtys,
        'accepted_qtys': accepted_qtys,
    }
    return render(request, 'report/closed_prod/completed_order.html', context)

def rejected_order(request, ftype, fdate, fmonth, fstartdate, fstopdate):
    if fdate == "NOW":
        fdate = datetime.today().strftime('%Y-%m-%d')
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    if fstartdate == "NOW":
        fstartdate = datetime.today().strftime('%Y-%m-%d')
    if fstopdate == "NOW":
        fstopdate = datetime.today().strftime('%Y-%m-%d')
    rejectedOrderList = getRejectedOrderList(ftype, fdate, fmonth, fstartdate, fstopdate)
    context = {
        'ftype': ftype,
        'fdate': fdate,
        'fmonth': fmonth,
        'fstartdate': fstartdate,
        'fstopdate': fstopdate,
        'rejectedOrderList': rejectedOrderList,
    }
    return render(request, 'report/closed_prod/rejected_order.html', context)

def canceled_order(request, ftype, fdate, fmonth, fstartdate, fstopdate):
    if fdate == "NOW":
        fdate = datetime.today().strftime('%Y-%m-%d')
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    if fstartdate == "NOW":
        fstartdate = datetime.today().strftime('%Y-%m-%d')
    if fstopdate == "NOW":
        fstopdate = datetime.today().strftime('%Y-%m-%d')
    canceledOrderList = getCanceledOrderList(ftype, fdate, fmonth, fstartdate, fstopdate)
    context = {
        'ftype': ftype,
        'fdate': fdate,
        'fmonth': fmonth,
        'fstartdate': fstartdate,
        'fstopdate': fstopdate,
        'canceledOrderList': canceledOrderList,
    }
    return render(request, 'report/closed_prod/canceled_order.html', context)

def work_records(request, ftype, fdate, fmonth, fstartdate, fstopdate):
    if fdate == "NOW":
        fdate = datetime.today().strftime('%Y-%m-%d')
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    if fstartdate == "NOW":
        fstartdate = datetime.today().strftime('%Y-%m-%d')
    if fstopdate == "NOW":
        fstopdate = datetime.today().strftime('%Y-%m-%d')
    empWorkRecords = getEmpWorkRecordsList(ftype, fdate, fmonth, fstartdate, fstopdate)
    workCenterWorkRecords = getWorkCenterWorkRecordsList(ftype, fdate, fmonth, fstartdate, fstopdate)
    context = {
        'ftype': ftype,
        'fdate': fdate,
        'fmonth': fmonth,
        'fstartdate': fstartdate,
        'fstopdate': fstopdate,
        'empWorkRecords': empWorkRecords,
        'workCenterWorkRecords': workCenterWorkRecords,
    }
    return render(request, 'report/work_records.html', context)

def ab_graph_wcg(request, fwcg, ftype, fmonth, fyear):
    workCenterGroupList = getMachineWorkCenterGroupList()
    if fwcg == "FIRST":
        fwcg = workCenterGroupList[0].WorkCenterGroup
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    if fyear == "NOW":
        fyear = datetime.today().strftime('%Y')
    year = fmonth[0:4]
    month = fmonth[5:7]
    workCenterInGroupList = getWorkCenterInGroupActiveList(fwcg, fmonth)
    #-- INITIALIZE
    x_size = 0
    y_size = 0
    max_hour_day = 0
    max_hour_month = 0
    working_hour_month = 0
    working_hour_month_percent = 0
    max_cap_month = 0
    max_hour_present = 0
    working_hour_present = 0
    working_hour_present_percent = 0
    wcg_oper = []
    wcg_manual = []
    wc_oper_list = []
    wc_setup_list = []
    wc_manual_list = []
    wc_target_list = []
    wc_cap_list = []
    if ftype == "MONTHLY":
        x_size = get_day_count(month, year)
    elif ftype == "YEARLY":
        x_size = 12
    wcg_oper = [0] * x_size
    wcg_setup = [0] * x_size
    wcg_manual = [0] * x_size
    isoweekdays = get_isoweekdays(x_size, month, year)
    is_pass = is_previous_month(month, year)
    if ftype == "MONTHLY" and workCenterInGroupList:
        is_collected = isCollectedABGraphData('WCG', fwcg, month, year)
        if not is_collected:
            for rs in getMonthlyWorkCenterOperForABGraph('WCG', fwcg, fmonth):
                wcg_oper[rs.Fday - 1] = rs.Foper
                if is_pass: 
                    saveABGraphData('WCG',fwcg,'W', rs.Foper, rs.Fday, month, year)
            for rs in getMonthlyWorkCenterSetupForABGraph('WCG', fwcg, fmonth):
                wcg_setup[rs.Fday - 1] = rs.Fsetup
                if is_pass: 
                    saveABGraphData('WCG',fwcg,'S', rs.Fsetup, rs.Fday, month, year)
        else:
            for rs in getABGraphCollectedData('WCG', fwcg, 'W', month, year):
                wcg_oper[rs.Fday - 1] = rs.Hour
            for rs in getABGraphCollectedData('WCG', fwcg, 'S', month, year):
                wcg_setup[rs.Fday - 1] = rs.Hour
        # No Record For Manual Now
        for rs in getMonthlyWorkCenterManualForABGraph('WCG', fwcg, fmonth):
            wcg_manual[rs.Fday - 1] = rs.WorkingHour
        y_size = 1.25 * (max(wcg_oper) + max(wcg_setup)) # higth of graph 25% more than highest bar
        max_hour_day = 24 * len(workCenterInGroupList)
        max_hour_month = max_hour_day * x_size
        working_hour_month = sum(wcg_oper) + sum(wcg_setup)
        working_hour_month_percent = round((working_hour_month / max_hour_month) * 100, 2)
        if(datetime.today().strftime('%Y-%m') == fmonth):
            current_day = int(datetime.today().strftime('%d'))
            max_hour_present = max_hour_day * current_day
            working_hour_present = sum(wcg_oper[0:current_day - 1]) + sum(wcg_setup[0:current_day - 1])
            working_hour_present_percent = round((working_hour_present / max_hour_present) * 100, 2)
        else:
            max_hour_present = max_hour_month
            working_hour_present = working_hour_month
            working_hour_present_percent = working_hour_month_percent
        for wc in workCenterInGroupList:
            temp_oper = [0] * x_size
            temp_setup = [0] * x_size
            temp_manual = [0] * x_size
            is_collected = isCollectedABGraphData('WC', wc.WorkCenterNo, month, year)
            if not is_collected:
                for rs in getMonthlyWorkCenterOperForABGraph('WC', wc.WorkCenterNo, fmonth):
                    temp_oper[rs.Fday - 1] = rs.Foper
                    if is_pass: 
                        saveABGraphData('WC',wc.WorkCenterNo,'W', rs.Foper, rs.Fday, month, year)
                for rs in getMonthlyWorkCenterSetupForABGraph('WC', wc.WorkCenterNo, fmonth):
                    temp_setup[rs.Fday - 1] = rs.Fsetup
                    if is_pass: 
                        saveABGraphData('WC',wc.WorkCenterNo,'S', rs.Fsetup, rs.Fday, month, year)
            else:
                for rs in getABGraphCollectedData('WC', wc.WorkCenterNo, 'W', month, year):
                    temp_oper[rs.Fday - 1] = rs.Hour
                for rs in getABGraphCollectedData('WC', wc.WorkCenterNo, 'S', month, year):
                    temp_setup[rs.Fday - 1] = rs.Hour   
            # No Record For Manual Now
            for rs in getMonthlyWorkCenterManualForABGraph('WC', wc.WorkCenterNo, fmonth):
                    temp_manual[rs.Fday - 1] = rs.WorkingHour 
            wc_oper_list.append(temp_oper)
            wc_setup_list.append(temp_setup)
            wc_manual_list.append(temp_manual)
            wc_target_list.append(wc.Target)
            wc_cap_list.append(wc.Capacity)
            max_cap_month = sum(wc_cap_list * x_size)
    context = {
        'workCenterGroupList': workCenterGroupList,
        'fwcg': fwcg,
        'ftype': ftype,
        'fmonth': fmonth,
        'fyear': fyear,
        'x_size': x_size,
        'y_size': y_size,
        'max_hour_day': max_hour_day,
        'max_hour_month': max_hour_month,
        'working_hour_month': working_hour_month,
        'working_hour_month_percent': working_hour_month_percent,
        'max_cap_month': max_cap_month,
        'max_hour_present': max_hour_present,
        'working_hour_present': working_hour_present,
        'working_hour_present_percent': working_hour_present_percent,
        'wcg_oper': wcg_oper,
        'wcg_setup': wcg_setup,
        'wcg_manual': wcg_manual,
        'workCenterInGroupList': workCenterInGroupList,
        'wc_oper_list': wc_oper_list,
        'wc_setup_list': wc_setup_list,
        'wc_manual_list': wc_manual_list,
        'wc_target_list': wc_target_list,
        'wc_cap_list': wc_cap_list,
        'isoweekdays' : isoweekdays,
    }
    return render(request, 'report/graph/ab_graph/ab_graph_wcg.html', context)

def ab_graph_rt(request, frt, ftype, fmonth, fyear):
    onRoutingList = getOnRoutingList()
    if frt == "FIRST":
        frt = onRoutingList[0].WorkCenterNo
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    if fyear == "NOW":
        fyear = datetime.today().strftime('%Y')
    year = fmonth[0:4]
    month = fmonth[5:7]
    workCenterInGroupList = getWorkCenterOnRoutingActiveList(frt, fmonth)
    #-- INITIALIZE
    x_size = 0
    y_size = 0
    max_hour_day = 0
    max_hour_month = 0
    working_hour_month = 0
    working_hour_month_percent = 0
    max_cap_month = 0
    max_hour_present = 0
    working_hour_present = 0
    working_hour_present_percent = 0
    wcg_oper = []
    wcg_setup = []
    wcg_manual = []
    wc_oper_list = []
    wc_setup_list = []
    wc_manual_list = []
    wc_target_list = []
    wc_cap_list = []
    if ftype == "MONTHLY":
        x_size = get_day_count(month, year)
    elif ftype == "YEARLY":
        x_size = 12
    wcg_oper = [0] * x_size
    wcg_setup = [0] * x_size
    wcg_manual = [0] * x_size
    isoweekdays = get_isoweekdays(x_size, month, year)
    is_pass = is_previous_month(month, year)
    if ftype == "MONTHLY" and workCenterInGroupList:
        is_collected = isCollectedABGraphData('RT', frt, month, year)
        if not is_collected:
            for rs in getMonthlyWorkCenterOperForABGraph('RT', frt, fmonth):
                wcg_oper[rs.Fday - 1] = rs.Foper
                if is_pass: 
                    saveABGraphData('RT',frt,'W', rs.Foper, rs.Fday, month, year)
            for rs in getMonthlyWorkCenterSetupForABGraph('RT', frt, fmonth):
                wcg_setup[rs.Fday - 1] = rs.Fsetup
                if is_pass: 
                 saveABGraphData('RT',frt,'S', rs.Fsetup, rs.Fday, month, year)
        else:
            for rs in getABGraphCollectedData('RT', frt, 'W', month, year):
                wcg_oper[rs.Fday - 1] = rs.Hour
            for rs in getABGraphCollectedData('RT', frt, 'S', month, year):
                wcg_setup[rs.Fday - 1] = rs.Hour
        # No Record For Manual Now
        for rs in getMonthlyWorkCenterManualForABGraph('RT', frt, fmonth):
            wcg_manual[rs.Fday - 1] = rs.WorkingHour
        y_size = 1.25 * (max(wcg_oper) + max(wcg_setup))
        max_hour_day = 24 * len(workCenterInGroupList)
        max_hour_month = max_hour_day * x_size
        working_hour_month = sum(wcg_oper) + sum(wcg_setup)
        working_hour_month_percent = round((working_hour_month / max_hour_month) * 100, 2)
        if(datetime.today().strftime('%Y-%m') == fmonth):
            current_day = int(datetime.today().strftime('%d'))
            max_hour_present = max_hour_day * current_day
            working_hour_present = sum(wcg_oper[0:current_day - 1]) + sum(wcg_setup[0:current_day - 1])
            working_hour_present_percent = round((working_hour_present / max_hour_present) * 100, 2)
        else:
            max_hour_present = max_hour_month
            working_hour_present = working_hour_month
            working_hour_present_percent = working_hour_month_percent
        for wc in workCenterInGroupList:
            temp_oper = [0] * x_size
            temp_setup = [0] * x_size
            temp_manual = [0] * x_size
            is_collected = isCollectedABGraphData('WC', wc.WorkCenterNo, month, year)
            if not is_collected:
                for rs in getMonthlyWorkCenterOperForABGraph('WC', wc.WorkCenterNo, fmonth):
                    temp_oper[rs.Fday - 1] = rs.Foper
                    if is_pass: 
                        saveABGraphData('WC',wc.WorkCenterNo,'W', rs.Foper, rs.Fday, month, year)
                for rs in getMonthlyWorkCenterSetupForABGraph('WC', wc.WorkCenterNo, fmonth):
                    temp_setup[rs.Fday - 1] = rs.Fsetup
                    if is_pass: 
                        saveABGraphData('WC',wc.WorkCenterNo,'S', rs.Fsetup, rs.Fday, month, year)
            else:
                for rs in getABGraphCollectedData('WC', wc.WorkCenterNo, 'W', month, year):
                    temp_oper[rs.Fday - 1] = rs.Hour
                for rs in getABGraphCollectedData('WC', wc.WorkCenterNo, 'S', month, year):
                    temp_setup[rs.Fday - 1] = rs.Hour  
            # No Record For Manual Now
            for rs in getMonthlyWorkCenterManualForABGraph('WC', wc.WorkCenterNo, fmonth):
                    temp_manual[rs.Fday - 1] = rs.WorkingHour
            wc_oper_list.append(temp_oper)
            wc_setup_list.append(temp_setup)
            wc_manual_list.append(temp_manual)
            wc_target_list.append(wc.Target)
            wc_cap_list.append(wc.Capacity)
            max_cap_month = sum(wc_cap_list * x_size)
    context = {
        'onRoutingList': onRoutingList,
        'frt': frt,
        'ftype': ftype,
        'fmonth': fmonth,
        'fyear': fyear,
        'x_size': x_size,
        'y_size': y_size,
        'max_hour_day': max_hour_day,
        'max_hour_month': max_hour_month,
        'working_hour_month': working_hour_month,
        'working_hour_month_percent': working_hour_month_percent,
        'max_cap_month': max_cap_month,
        'max_hour_present': max_hour_present,
        'working_hour_present': working_hour_present,
        'working_hour_present_percent': working_hour_present_percent,
        'wcg_oper': wcg_oper,
        'wcg_setup': wcg_setup,
        'wcg_manual': wcg_manual,
        'workCenterInGroupList': workCenterInGroupList,
        'wc_oper_list': wc_oper_list,
        'wc_setup_list': wc_setup_list,
        'wc_manual_list': wc_manual_list,
        'wc_target_list': wc_target_list,
        'wc_cap_list': wc_cap_list,
        'isoweekdays': isoweekdays,
    }
    return render(request, 'report/graph/ab_graph/ab_graph_rt.html', context)

def reject_per_pfc(request, fpfc, fmonth):
    profitCenterList = getProfitCenterList()
    if fpfc == "FIRST":
        fpfc = profitCenterList[0].ProfitCenter
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    year = fmonth[0:4]
    month = fmonth[5:7]
    x_size = 0
    y_size = 1
    x_size = get_day_count(month, year)
    rejs, pro_qtys, rej_qtys = [0] * x_size, [0] * x_size, [0] * x_size
    mper, mpro, mrej = 0, 0, 0
    isoweekdays = get_isoweekdays(x_size, month, year)
    for x in range(x_size):
        day = x + 1
        item = None
        item = getRejectPerDataPFC(fpfc, day, month, year)
        if item:
            pro_qtys[x] = item.ProcessQty
            rej_qtys[x] = item.RejectedQty
            mpro = mpro + pro_qtys[x] if pro_qtys[x] else mpro
            mrej = mrej + rej_qtys[x] if rej_qtys[x] else mrej
            if pro_qtys[x] and pro_qtys[x] > 0:
                rejs[x] = round(((rej_qtys[x] / pro_qtys[x]) * 100), 2)
        y_size = rejs[x] if rejs[x] > y_size else y_size
    y_size = int(y_size) * 1.25
    week_titles, week_pros, week_rejs, week_pers = get_week_rej_per_info(isoweekdays, pro_qtys, rej_qtys)
    # Whole Month
    if mpro and mpro > 0:
        mper = round(((mrej / mpro) * 100), 2)
    context = {
        'fpfc': fpfc,
        'fmonth': fmonth,
        'profitCenterList': profitCenterList,
        'x_size': x_size,
        'y_size' : y_size,
        'rejs' : rejs,
        'pro_qtys': pro_qtys,
        'rej_qtys': rej_qtys,
        'isoweekdays': isoweekdays,
        'mper': mper,
        'mrej': mrej,
        'mpro': mpro,
        'week_titles': week_titles,
        'week_pros': week_pros,
        'week_rejs': week_rejs,
        'week_pers': week_pers,
    }
    return render(request, 'report/graph/reject_per/reject_per_pfc.html', context)

def reject_rc_pfc(request, fpfc, fmonth):
    profitCenterList = getProfitCenterList()
    if fpfc == "FIRST":
        fpfc = profitCenterList[0].ProfitCenter
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    year = fmonth[0:4]
    month = fmonth[5:7]
    recs = getRecordRejectionPFC(fpfc, month, year)
    context = {
        'fpfc': fpfc,
        'fmonth': fmonth,
        'profitCenterList': profitCenterList,
        'recs': recs,
    }
    return render(request, 'report/graph/reject_per/reject_rc_pfc.html', context)

def reject_per_wcg(request, fwcg, fmonth):
    workCenterGroupList = getMachineWorkCenterGroupList()
    if fwcg == "FIRST":
        fwcg = workCenterGroupList[0].WorkCenterGroup
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    year = fmonth[0:4]
    month = fmonth[5:7]
    x_size = 0
    y_size = 1
    x_size = get_day_count(month, year)
    rejs, pro_qtys, rej_qtys = [0] * x_size, [0] * x_size, [0] * x_size
    mper, mpro, mrej = 0, 0, 0
    isoweekdays = get_isoweekdays(x_size, month, year)
    for x in range(x_size):
        day = x + 1
        item = None
        item = getRejectPerDataWCG(fwcg, day, month, year)
        if item:
            pro_qtys[x] = item.ProcessQty
            rej_qtys[x] = item.RejectedQty
            mpro = mpro + pro_qtys[x] if pro_qtys[x] else mpro
            mrej = mrej + rej_qtys[x] if rej_qtys[x] else mrej
            if pro_qtys[x] and pro_qtys[x] > 0:
                rejs[x] = round(((rej_qtys[x] / pro_qtys[x]) * 100), 2)
        y_size = rejs[x] if rejs[x] > y_size else y_size
    y_size = int(y_size) * 1.25
    week_titles, week_pros, week_rejs, week_pers = get_week_rej_per_info(isoweekdays, pro_qtys, rej_qtys)
    # Whole Month
    if mpro and mpro > 0:
        mper = round(((mrej / mpro) * 100), 2)
    context = {
        'fwcg': fwcg,
        'fmonth': fmonth,
        'workCenterGroupList': workCenterGroupList,
        'x_size': x_size,
        'y_size' : y_size,
        'rejs' : rejs,
        'pro_qtys': pro_qtys,
        'rej_qtys': rej_qtys,
        'isoweekdays': isoweekdays,
        'mper': mper,
        'mrej': mrej,
        'mpro': mpro,
        'week_titles': week_titles,
        'week_pros': week_pros,
        'week_rejs': week_rejs,
        'week_pers': week_pers,
    }
    return render(request, 'report/graph/reject_per/reject_per_wcg.html', context)

def reject_per_rt(request, fwc, fmonth):
    workCenterList = getWorkCenterRoutingList()
    if fwc == "FIRST":
        fwc = workCenterList[0].WorkCenterNo
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    year = fmonth[0:4]
    month = fmonth[5:7]
    x_size = 0
    y_size = 1
    x_size = get_day_count(month, year)
    rejs, pro_qtys, rej_qtys = [0] * x_size, [0] * x_size, [0] * x_size
    mper, mpro, mrej = 0, 0, 0
    isoweekdays = get_isoweekdays(x_size, month, year)
    for x in range(x_size):
        day = x + 1
        item = getRejectPerDataRT(fwc, day, month, year)
        if item:
            pro_qtys[x] = item.ProcessQty
            rej_qtys[x] = item.RejectedQty
            mpro = mpro + pro_qtys[x] if pro_qtys[x] else mpro
            mrej = mrej + rej_qtys[x] if rej_qtys[x] else mrej
            if pro_qtys[x] and pro_qtys[x] > 0:
                rejs[x] = round(((rej_qtys[x] / pro_qtys[x]) * 100), 2)
        y_size = rejs[x] if rejs[x] > y_size else y_size
    y_size = int(y_size) * 1.25
    week_titles, week_pros, week_rejs, week_pers = get_week_rej_per_info(isoweekdays, pro_qtys, rej_qtys)
    # Whole Month
    if mpro and mpro > 0:
        mper = round(((mrej / mpro) * 100), 2)
    context = {
        'fwc': fwc,
        'fmonth': fmonth,
        'workCenterList': workCenterList,
        'x_size': x_size,
        'y_size' : y_size,
        'rejs' : rejs,
        'pro_qtys': pro_qtys,
        'rej_qtys': rej_qtys,
        'isoweekdays': isoweekdays,
        'mper': mper,
        'mrej': mrej,
        'mpro': mpro,
        'week_titles': week_titles,
        'week_pros': week_pros,
        'week_rejs': week_rejs,
        'week_pers': week_pers,
    }
    return render(request, 'report/graph/reject_per/reject_per_rt.html', context)

def efficiency_pfc(request, fpfc, fmonth):
    profitCenterList = getProfitCenterList()
    if fpfc == "FIRST":
        fpfc = profitCenterList[0].ProfitCenter
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    year = fmonth[0:4]
    month = fmonth[5:7]
    x_size = 0
    y_size = 100
    x_size = get_day_count(month, year)
    setups, est_setups, act_setups = [0] * x_size, [0] * x_size, [0] * x_size
    opers, est_opers, act_opers = [0] * x_size, [0] * x_size, [0] * x_size
    labors, est_labors, act_labors = [0] * x_size, [0] * x_size, [0] * x_size
    isoweekdays = get_isoweekdays(x_size, month, year)
    for x in range(x_size):
        day = x + 1
        item = getEfficiencyDataPFC(fpfc, day, month, year)
        if item:
            est_setups[x] = item.EstSetup
            est_opers[x] = item.EstOper
            est_labors[x] = item.EstLabor
            act_setups[x] = item.ActSetup
            act_opers[x] = item.ActOper
            act_labors[x] = item.ActLabor
            if act_setups[x] and act_setups[x] > 0:
                setups[x] = int((est_setups[x] / act_setups[x]) * 100)
            if act_opers[x] and act_opers[x] > 0:
                opers[x] = int((est_opers[x] / act_opers[x]) * 100)
            if act_labors[x] and act_labors[x] > 0:
                labors[x] = int((est_labors[x] / act_labors[x]) * 100)
        y_size = opers[x] if opers[x] > y_size else y_size
        y_size = setups[x] if setups[x] > y_size else y_size
        y_size = labors[x] if labors[x] > y_size else y_size
    y_size = int(y_size) * 1.25
    context = {
        'fpfc': fpfc,
        'fmonth': fmonth,
        'profitCenterList': profitCenterList,
        'x_size': x_size,
        'y_size' : y_size,
        'setups' : setups,
        'opers': opers,
        'labors': labors,
        'isoweekdays': isoweekdays,
    }
    return render(request, 'report/graph/efficiency/efficiency_pfc.html', context)

def efficiency_wcg(request, fwcg, fmonth):
    workCenterGroupList = getMachineWorkCenterGroupList()
    if fwcg == "FIRST":
        fwcg = workCenterGroupList[0].WorkCenterGroup
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    year = fmonth[0:4]
    month = fmonth[5:7]
    x_size = 0
    y_size = 100
    x_size = get_day_count(month, year)
    setups, est_setups, act_setups = [0] * x_size, [0] * x_size, [0] * x_size
    opers, est_opers, act_opers = [0] * x_size, [0] * x_size, [0] * x_size
    labors, est_labors, act_labors = [0] * x_size, [0] * x_size, [0] * x_size
    isoweekdays = get_isoweekdays(x_size, month, year)
    for x in range(x_size):
        day = x + 1
        item = getEfficiencyDataWCG(fwcg, day, month, year)
        if item:
            est_setups[x] = item.EstSetup
            est_opers[x] = item.EstOper
            est_labors[x] = item.EstLabor
            act_setups[x] = item.ActSetup
            act_opers[x] = item.ActOper
            act_labors[x] = item.ActLabor
            if act_setups[x] and act_setups[x] > 0:
                setups[x] = int((est_setups[x] / act_setups[x]) * 100)
            if act_opers[x] and act_opers[x] > 0:
                opers[x] = int((est_opers[x] / act_opers[x]) * 100)
            if act_labors[x] and act_labors[x] > 0:
                labors[x] = int((est_labors[x] / act_labors[x]) * 100)
        y_size = opers[x] if opers[x] > y_size else y_size
        y_size = setups[x] if setups[x] > y_size else y_size
        y_size = labors[x] if labors[x] > y_size else y_size
    y_size = int(y_size) * 1.25
    context = {
        'fwcg': fwcg,
        'fmonth': fmonth,
        'workCenterGroupList': workCenterGroupList,
        'x_size': x_size,
        'y_size' : y_size,
        'setups' : setups,
        'opers': opers,
        'labors': labors,
        'isoweekdays': isoweekdays,
    }
    return render(request, 'report/graph/efficiency/efficiency_wcg.html', context)

def efficiency_wc(request, fwc, fmonth):
    workCenterList = getWorkCenterRoutingList()
    if fwc == "FIRST":
        fwc = workCenterList[0].WorkCenterNo
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    year = fmonth[0:4]
    month = fmonth[5:7]
    x_size = 0
    y_size = 100
    x_size = get_day_count(month, year)
    setups, est_setups, act_setups = [0] * x_size, [0] * x_size, [0] * x_size
    opers, est_opers, act_opers = [0] * x_size, [0] * x_size, [0] * x_size
    labors, est_labors, act_labors = [0] * x_size, [0] * x_size, [0] * x_size
    isoweekdays = get_isoweekdays(x_size, month, year)
    for x in range(x_size):
        day = x + 1
        item = getEfficiencyDataWC(fwc, day, month, year)
        if item:
            est_setups[x] = item.EstSetup
            est_opers[x] = item.EstOper
            est_labors[x] = item.EstLabor
            act_setups[x] = item.ActSetup
            act_opers[x] = item.ActOper
            act_labors[x] = item.ActLabor
            if act_setups[x] and act_setups[x] > 0:
                setups[x] = int((est_setups[x] / act_setups[x]) * 100)
            if act_opers[x] and act_opers[x] > 0:
                opers[x] = int((est_opers[x] / act_opers[x]) * 100)
            if act_labors[x] and act_labors[x] > 0:
                labors[x] = int((est_labors[x] / act_labors[x]) * 100)
        y_size = opers[x] if opers[x] > y_size else y_size
        y_size = setups[x] if setups[x] > y_size else y_size
        y_size = labors[x] if labors[x] > y_size else y_size
    y_size = int(y_size) * 1.25
    context = {
        'fwc': fwc,
        'fmonth': fmonth,
        'workCenterList': workCenterList,
        'x_size': x_size,
        'y_size' : y_size,
        'setups' : setups,
        'opers': opers,
        'labors': labors,
        'isoweekdays': isoweekdays,
    }
    return render(request, 'report/graph/efficiency/efficiency_wc.html', context)

def cumulative_rejection(request):
    pfcs = getProfitCenterList()
    pfc = request.POST.get('pfc', pfcs[0].ProfitCenter)
    month = request.POST.get('month', datetime.today().strftime('%Y-%m')) if request.POST.get('month') != '' else datetime.today().strftime('%Y-%m')
    month_no = month[5:7]
    year = month[0:4]
    day_count = get_day_count(month_no, year)
    days = []
    orders = []
    rejects = []
    percentages = []
    total_order = 0
    total_reject = 0
    total_percentage = 0
    is_sundays = get_isoweekdays(day_count, month_no, year)
    cums = getCumulativeRejectionList(pfc, int(month_no), int(year))
    day = 1
    for cum in cums:
        # handle case no order finish in that day - previous and between
        while int(cum.Day) != day:
            tmp = datetime(int(year), int(month_no), (day)).strftime('%a %d')
            days.append(tmp)
            orders.append(0)
            rejects.append(0)
            percentages.append(0)
            day = day + 1
        tmp = datetime(int(year), int(month_no), (day)).strftime('%a %d')
        days.append(tmp)
        orders.append(cum.OrderQty)
        rejects.append(cum.RejectQty)
        percentages.append(cum.RejectPercentage)
        total_order = total_order + cum.OrderQty
        total_reject = total_reject + cum.RejectQty
        day = day + 1
    # handle case no order finish in that day - after
    while len(orders) < day_count:
        tmp = datetime(int(year), int(month_no), (day)).strftime('%a %d')
        days.append(tmp)
        orders.append(0)
        rejects.append(0)
        percentages.append(0)
        day = day + 1
    if total_order != 0:
        total_percentage = round(((total_reject / total_order) * 100), 3)
    context = {
        'pfcs': pfcs,
        'month': month,
        'pfc': pfc,
        'day_count' : day_count,
        'days' : days,
        'orders' : orders,
        'rejects' : rejects,
        'percentages' : percentages,
        'total_order' : total_order,
        'total_reject' : total_reject,
        'total_percentage' : total_percentage,
        'is_sundays' : is_sundays,
    }
    return render(request, 'report/graph/cumulative_rejection.html', context)

def cumulative_rejection_record(request):
    pfcs = getProfitCenterList()
    pfc = request.POST.get('pfc', pfcs[0].ProfitCenter)
    is_rej_only = request.POST.get('is_rej_only', False)
    is_rej_only = True if is_rej_only == 'Yes' else False
    month = request.POST.get('month', datetime.today().strftime('%Y-%m')) if request.POST.get('month') != '' else datetime.today().strftime('%Y-%m')
    month_no = month[5:7]
    year = month[0:4]
    recs = getCumulativeRejectionRecordList(pfc, int(month_no), int(year), is_rej_only)
    context = {
        'pfcs': pfcs,
        'month': month,
        'pfc': pfc,
        'is_rej_only' : is_rej_only,
        'recs' : recs,
    }
    return render(request, 'report/graph/cumulative_rejection_record.html', context)

def con_operation(request, fwc, fmonth):
    workCenterList = getWorkCenterRoutingList()
    if fwc == "FIRST":
        fwc = workCenterList[0].WorkCenterNo
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    confirmOperationList = getConfirmOperationList(fwc, fmonth)
    context = {
        'fwc': fwc,
        'fmonth': fmonth,
        'workCenterList': workCenterList,
        'confirmOperationList': confirmOperationList,
    }
    return render(request, 'report/con_operation.html', context)

def operating_history(request, fwc, fmc, ftype, fstartdate, fstopdate, fmonth):
    wc_list = getWorkCenterRoutingList()
    mc_list = getWorkCenterMachineList()
    if fwc == "FIRST":
        fwc = wc_list[0].WorkCenterNo
    if fmc == "FIRST":
        fmc = mc_list[0].WorkCenterNo
    if fstartdate == "NOW":
        fstartdate = datetime.today().strftime('%Y-%m-%d')
    if fstopdate == "NOW":
        fstopdate = datetime.today().strftime('%Y-%m-%d')
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    con_list = getOperatingHistoryConfirm(fwc, fmc, ftype, fstartdate, fstopdate, fmonth)
    opr_list = getOperatingHistoryOperate(fwc, fmc, ftype, fstartdate, fstopdate, fmonth)
    context = {
        'wc_list': wc_list,
        'mc_list': mc_list,
        'fwc': fwc,
        'fmc': fmc,
        'ftype': ftype,
        'fstopdate': fstopdate,
        'fstartdate': fstartdate,
        'fmonth': fmonth,
        'con_list': con_list,
        'opr_list': opr_list,
    }
    return render(request, 'report/operating_history.html', context)

def over_est_operation(request, fwc, fweek):
    workCenterList = getWorkCenterRoutingList()
    if fwc == "FIRST":
        fwc = workCenterList[0].WorkCenterNo
    if fweek == "NOW":
        fweek = datetime.today().strftime('%Y-W%W')
    overEstOperationList = getOverEstOperationList(fwc, fweek)
    start_date, end_date = get_date_between_week(fweek)
    mc_list = []
    est_setup_sum = []
    est_oper_sum = []
    target_sum = []
    est_labor_sum = []
    actual_setup = []
    actual_oper = []
    actual_labor = []
    actual_ai = []
    setup_percent = []
    oper_percent = []
    tar_percent = []
    labor_percent = []
    ai_percent = []
    yellow_setup = []
    yellow_oper = []
    yellow_labor = []
    yellow_ai = []
    red_setup = []
    red_oper = []
    red_labor = []
    red_ai = []
    for op in overEstOperationList:
        tmps = []
        mcs = getMachineWorkOnOperation(op.OrderNo, op.OperationNo)
        for mc in mcs:
            if mc.WorkCenterNo not in tmps:
                tmps.append(mc.WorkCenterNo)
        mc_list.append(','.join(str(e) for e in tmps))
        est_setup = int(op.EstSetupTime * op.ProcessQty)
        est_oper = int(op.EstOperationTime * op.ProcessQty)
        est_labor = int(op.EstLaborTime * op.ProcessQty)
        est_setup_sum.append(est_setup)
        est_oper_sum.append(est_oper)
        target = 0 if not op.TargetValue else int(op.TargetValue * op.ProcessQty)
        target_sum.append(target)
        est_labor_sum.append(est_labor)
        act_setup = op.ActualSetup if op.ActualSetup else 0
        act_oper = op.ActualOper if op.ActualOper else 0
        act_labor = op.ActualLabor if op.ActualLabor else 0
        xIdle = getTotalIdleTime(op.OrderNo, op.OperationNo)
        idle = int(xIdle.Idle) if xIdle and xIdle.Idle else 0
        act_ai = int(act_oper) + int(idle)
        actual_setup.append(int(act_setup))
        actual_oper.append(int(act_oper))
        actual_labor.append(int(act_labor))
        actual_ai.append(act_ai)
        if act_setup > 0:
            setup_per = int(((op.EstSetupTime * op.ProcessQty) / act_setup) * 100)
            setup_percent.append(str(setup_per) + "%")
        else:
            setup_percent.append("-")
        if act_oper > 0:
            oper_per = int(((op.EstOperationTime * op.ProcessQty) / act_oper) * 100)
            oper_percent.append(str(oper_per) + "%")
            tar_percent.append(str(int((target / act_oper) * 100)) + "%")
        else:
            oper_percent.append("-")
            tar_percent.append("-")
        if act_labor > 0:
            labor_per = int(((op.EstLaborTime * op.ProcessQty) / act_labor) * 100)
            labor_percent.append(str(labor_per) + "%")
        else:
            labor_percent.append("-")
        if act_ai > 0:
            ai_per = int(((op.EstOperationTime * op.ProcessQty) / act_ai) * 100)
            ai_percent.append(str(ai_per) + "%")
        else:
            ai_percent.append("-")
        red_setup.append(True if (est_setup == 0 and act_setup > 0) or act_setup > est_setup * OVER_EST_RATING else False)
        red_oper.append(True if (est_oper == 0 and act_oper > 0) or act_oper > est_oper * OVER_EST_RATING else False)
        red_labor.append(True if (est_labor == 0 and act_labor > 0) or act_labor > est_labor * OVER_EST_RATING else False)
        red_ai.append(True if (est_oper == 0 and act_ai > 0) or act_ai > est_oper * OVER_EST_RATING else False)
        yellow_setup.append(True if act_setup > est_setup else False)
        yellow_oper.append(True if act_oper > est_oper else False)
        yellow_labor.append(True if act_labor > est_labor else False)
        yellow_ai.append(True if act_ai > est_oper else False)
    context = {
        'fwc': fwc,
        'fweek': fweek,
        'workCenterList': workCenterList,
        'overEstOperationList': overEstOperationList,
        'start_date': start_date,
        'end_date': end_date,
        'mc_list': mc_list,
        'est_setup_sum': est_setup_sum,
        'est_oper_sum': est_oper_sum,
        'target_sum': target_sum,
        'est_labor_sum': est_labor_sum,
        'actual_setup': actual_setup,
        'actual_oper': actual_oper,
        'actual_labor': actual_labor,
        'actual_ai': actual_ai,
        'setup_percent': setup_percent,
        'oper_percent': oper_percent,
        'tar_percent': tar_percent,
        'labor_percent': labor_percent,
        'ai_percent': ai_percent,
        'yellow_setup': yellow_setup,
        'yellow_oper': yellow_oper,
        'yellow_labor': yellow_labor,
        'yellow_ai': yellow_ai,
        'red_setup': red_setup,
        'red_oper': red_oper,
        'red_labor': red_labor,
        'red_ai': red_ai,
    }
    return render(request, 'report/over_est_operation.html', context)

def plan_act_cycle_time(request, ffgcode, fwc, ftype, fstartdate, fstopdate, fmonth):
    wc_list = getWorkCenterRoutingList()
    planActualCycleTimeList = []
    per_setups, per_opers, per_labors, acts, per_tars, per_ais = [], [], [], [], [], []
    yellow_setup, yellow_oper, yellow_labor, yellow_ai, red_setup, red_oper, red_labor, red_ai = [], [], [], [], [], [], [], []
    if fwc == "FIRST":
        fwc = wc_list[0].WorkCenterNo
    if fstartdate == "NOW":
        fstartdate = datetime.today().strftime('%Y-%m-%d')
    if fstopdate == "NOW":
        fstopdate = datetime.today().strftime('%Y-%m-%d')
    if fmonth == "NOW":
        fmonth = datetime.today().strftime('%Y-%m')
    if ffgcode != "NONE":
        planActualCycleTimeList = getPlanActualCycleTimeList(ffgcode, fwc, ftype, fstartdate, fstopdate, fmonth)
        for item in planActualCycleTimeList:
            per_setup = '-'
            per_oper = '-'
            per_labor = '-'
            per_tar = '-'
            act = '-'
            per_ai = '-'
            item.AI = item.AI if item.AI else item.ActOper
            if item.Est2Setup > 0:
                per_setup = str(round((item.ActSetup / item.Est2Setup) * 100, 2)) + '%'
            if item.Est2Oper > 0:
                per_oper = str(round((item.ActOper / item.Est2Oper) * 100, 2)) + '%'
                per_ai = str(round((item.AI / item.Est2Oper) * 100, 2)) + '%'
            if item.Est2Labor > 0:
                per_labor = str(round((item.ActLabor / item.Est2Labor) * 100, 2)) + '%'
            if item.TargetValue:
                act = int(item.TargetValue * item.ProcessQty)
                per_tar = str(round((item.ActOper / act) * 100, 2)) + '%'
            per_setups.append(per_setup)
            per_opers.append(per_oper)
            per_labors.append(per_labor)
            acts.append(act)
            per_tars.append(per_tar)
            per_ais.append(per_ai)
            red_setup.append(True if (item.Est2Setup == 0 and item.ActSetup > 0) or item.ActSetup > float(item.Est2Setup) * OVER_EST_RATING else False)
            red_oper.append(True if (item.Est2Oper == 0 and item.ActOper > 0) or item.ActOper > float(item.Est2Oper) * OVER_EST_RATING else False)
            red_labor.append(True if (item.Est2Labor == 0 and item.ActLabor > 0) or item.ActLabor > float(item.Est2Labor) * OVER_EST_RATING else False)
            red_ai.append(True if (item.Est2Oper == 0 and item.AI > 0) or item.AI > float(item.Est2Oper) * OVER_EST_RATING else False)
            yellow_setup.append(True if item.ActSetup > item.Est2Setup else False)
            yellow_oper.append(True if item.ActOper > item.Est2Oper else False)
            yellow_labor.append(True if item.ActLabor > item.Est2Labor else False)
            yellow_ai.append(True if item.AI > item.Est2Oper else False)
    context = {
        'ffgcode': ffgcode,
        'fwc': fwc,
        'ftype': ftype,
        'fstartdate': fstartdate,
        'fstopdate': fstopdate,
        'fmonth': fmonth,
        'wc_list': wc_list,
        'planActualCycleTimeList': planActualCycleTimeList,
        'per_setups': per_setups,
        'per_opers': per_opers,
        'per_labors': per_labors,
        'acts': acts,
        'per_tars': per_tars,
        'per_ais': per_ais,
        'yellow_setup': yellow_setup,
        'yellow_oper': yellow_oper,
        'yellow_labor': yellow_labor,
        'yellow_ai': yellow_ai,
        'red_setup': red_setup,
        'red_oper': red_oper,
        'red_labor': red_labor,
        'red_ai': red_ai,
    }
    return render(request, 'report/plan_act_cycle_time.html', context)

def order_analysis(request):
    routings = getWorkCenterRoutingList()
    profit_centers = getProfitCenterList()
    token = generate_random_token(6)
    context = {
        'profit_centers': profit_centers,
        'routings': routings,
        'token': token,
    }
    return render(request, 'report/order_analysis.html', context)

#--------------------------------------------------------------------------- SAP

def sap_order(request, fdate, fhour):
    if fdate == "NOW":
        fdate = datetime.today().strftime('%Y-%m-%d')
    if fhour == "NOW":
        fhour = datetime.today().strftime('%H')
    sapOrderList = getSAPOrderList(fdate, fhour)
    context = {
        'fdate' : fdate,
        'fhour' : fhour,
        'sapOrderList' : sapOrderList,
    }
    return render(request, 'sap_log/data_from_sap/sap_order.html', context)

def sap_routing(request, fdate, fhour):
    if fdate == "NOW":
        fdate = datetime.today().strftime('%Y-%m-%d')
    if fhour == "NOW":
        fhour = datetime.today().strftime('%H')
    sapRoutingList = getSAPRoutingList(fdate, fhour)
    context = {
        'fdate' : fdate,
        'fhour' : fhour,
        'sapRoutingList' : sapRoutingList,
    }
    return render(request, 'sap_log/data_from_sap/sap_routing.html', context)

def sap_component(request, fdate, fhour):
    if fdate == "NOW":
        fdate = datetime.today().strftime('%Y-%m-%d')
    if fhour == "NOW":
        fhour = datetime.today().strftime('%H')
    sapComponentList = getSAPComponentList(fdate, fhour)
    context = {
        'fdate' : fdate,
        'fhour' : fhour,
        'sapComponentList' : sapComponentList,
    }
    return render(request, 'sap_log/data_from_sap/sap_component.html', context)

def sap_report(request, fdate, fhour):
    if fdate == "NOW":
        fdate = datetime.today().strftime('%Y-%m-%d')
    if fhour == "NOW":
        fhour = datetime.today().strftime('%H')
    sapReportList = getSAPReportList(fdate, fhour)
    context = {
        'fdate' : fdate,
        'fhour' : fhour,
        'sapReportList' : sapReportList,
    }
    return render(request, 'sap_log/data_to_sap/sap_report.html', context)

def sap_mod(request, fdate, fhour):
    if fdate == "NOW":
        fdate = datetime.today().strftime('%Y-%m-%d')
    if fhour == "NOW":
        fhour = datetime.today().strftime('%H')
    sapModifierList = getSAPModifierList(fdate, fhour)
    context = {
        'fdate' : fdate,
        'fhour' : fhour,
        'sapModifierList' : sapModifierList,
    }
    return render(request, 'sap_log/data_to_sap/sap_mod.html', context)

#------------------------------------------------------------------- ADMIN PANEL

def controller(request):
    userList = getUserList()
    empNameList = []
    for user in userList:
        empNameList.append(getEmpIDByUserID(user.UserID))
    overtimehour = getOvertimeHour()
    canMP = getManualReportAllow()
    refreshSecond = getRefreshSecond()
    drawingAppPath = getDrawingAppPath()
    profitCenterList = getProfitCenterList()
    # update_estimate()
    # update_target_cycle_time()
    # run_collect_ab_data('2023-10') Done
    # run_collect_ab_data('2023-09') Done 
    # run_collect_ab_data('2023-08') Done
    # run_collect_ab_data('2023-07') Done
    # run_collect_ab_data('2023-06') Done
    # run_collect_ab_data('2023-05') Done
    # run_collect_ab_data('2023-04') Done
    # run_collect_ab_data('2023-03') Done
    # run_collect_ab_data('2023-02') Done
    # run_collect_ab_data('2023-01') Done
    # run_collect_ab_data('2022-12') Done
    # run_collect_ab_data('2022-11') Done
    # run_collect_ab_data('2022-10') Done 
    # run_collect_ab_data('2022-09') Done
    # run_collect_ab_data('2022-08') Done
    # run_collect_ab_data('2022-07') Done
    # run_collect_ab_data('2022-06') Done
    # run_collect_ab_data('2022-05') Done
    # run_collect_ab_data('2022-04') Done
    # run_collect_ab_data('2022-03') Done
    # run_collect_ab_data('2022-02') Done
    # run_collect_ab_data('2022-01') Done
    context = {
        'userList': userList,
        'empNameList': empNameList,
        'overtimehour': overtimehour,
        'canMP': canMP,
        'refreshSecond': refreshSecond,
        'drawingAppPath': drawingAppPath,
        'profitCenterList': profitCenterList,
    }
    return render(request, 'admin_panel/controller.html', context)

def error_data(request):
    orderNoRoutingList = getSAPOrderNoRoutingList()
    duplicateRoutingList = getSAPDuplicateRoutingList()
    workCenterErrorList = getWorkCenterErrorList()
    orderStopNotStartList = getOrderStopNotStart()
    operationRemainQtyList = getOperationRemainQtyList()
    lastProcessStopOrderNotStop = getLastProcessStopOrderNotStop()
    context = {
        'orderNoRoutingList': orderNoRoutingList,
        'duplicateRoutingList': duplicateRoutingList,
        'workCenterErrorList': workCenterErrorList,
        'orderStopNotStartList': orderStopNotStartList,
        'operationRemainQtyList': operationRemainQtyList,
        'lastProcessStopOrderNotStop': lastProcessStopOrderNotStop,
    }
    return render(request, 'admin_panel/error_data.html', context)

def drawing(request, dir, fg_code):
    drawingAppPath = getDrawingAppPath()
    path = drawingAppPath + str(dir) + '&' + str(fg_code)
    print(path)
    return redirect(path)

def diskpulse(request):
    hours = [''] * 24
    mins = [''] * 60
    context = {
        'hours' : hours,
        'mins' : mins,
        'search_txt': "",
        'fopr': "ALL",
        'fdate' : None,
        'fhour' : None,
        'list': [],
        'fpaths': [],
        'fnames': [],
    }
    return render(request, 'diskpulse.html', context)

def diskpulse_search(request):
    # Created = 0, Deleted = 1, Modified = 2, Renamed = 3, Renamed To = 4
    hours = [''] * 24
    mins = [''] * 60
    exclude_tmp = True 
    exclude_mod = False
    ## 
    search_txt = request.POST['search_txt'].strip()
    fopr = request.POST['fopr']
    fdate = request.POST['fdate'] if request.POST['fdate'] != None else None
    fhour = request.POST['fhour'] if request.POST['fhour'] != None and request.POST['fhour'] != "" else None
    fmin = request.POST['fmin'] if request.POST['fmin'] != None and request.POST['fmin'] != "" else None
    list, fpaths, fnames, = [], [], []
    is_too_heavy = False
    if search_txt == "" and (not fdate or not fhour or not fmin):
        is_too_heavy = True
    if not is_too_heavy:
        conn = pyodbc.connect('Driver={SQL Server};''Server=SVCCS-SFR\SQLEXPRESS01;''Database=DiskPulse;''UID=sa;''PWD=$fr@2021;''Trusted_Connection=yes;')
        cursor = conn.cursor()
        sql = "SELECT * FROM [changes] WHERE (owner LIKE '%"+str(search_txt)+"%' OR fname LIKE '%"+str(search_txt)+"%' OR dt LIKE '%"+str(search_txt)+"%')"
        if exclude_tmp:
            sql += " AND fname NOT LIKE '%.tmp' AND fname NOT LIKE '%~$%'"
        if exclude_mod:
            sql += " AND ctype != '2'"
        if fopr != "ALL":
            sql += " AND ctype = "+str(fopr)
        if fdate:
            date = fdate.replace("-", "/")
            sql += " AND dt LIKE '"+str(date)+"%'"
        if fhour:
            hour = "0" + str(fhour) if int(fhour) < 10 else str(fhour)
            sql += " AND dt LIKE '% "+str(hour)+":%'"
        if fmin:
            min = "0" + str(fmin) if int(fmin) < 10 else str(fmin)
            sql += " AND dt LIKE '%:"+str(min)+":%'"
        cursor.execute(sql)
        list = cursor.fetchall()
        for item in list:
            fpath, fname = get_file_info(item.fname)
            fpaths.append(fpath)
            fnames.append(fname)
    context = {
        'hours' : hours,
        'mins' : mins,
        'search_txt': search_txt,
        'fopr': fopr,
        'fdate' : fdate,
        'fhour' : fhour,
        'fmin' : fmin,
        'list': list,
        'fpaths': fpaths,
        'fnames': fnames,
    }
    return render(request, 'diskpulse.html', context)

def get_file_info(path):
    fpath = path
    fname = '-'
    tmps = path.split("\\")
    if len(tmps) > 1:
        fpath = path.replace("\\" + tmps[-1],"")
        fname = tmps[-1]
    return fpath, fname

################################################################################
#################################### REQUEST ###################################
################################################################################

def fp_emp_search(request):
    operator_id = request.GET.get('operator_id')
    path = ""
    operator = getOperator(operator_id)
    if emp != None:
        if isOperatorOperating(operator_id):
            oopr = getOperatorOperatingByEmpID(operator_id)
            path = str(oopr.OperatorOrderNo) + "" + str(oopr.OperatorOperationNo)
    data = {
        'path': path,
    }
    return JsonResponse(data)

#-------------------------------------------------------------------- MAIN TABLE

def get_operating_workcenter_list(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    OWCList = [list(i) for i in getOperatingWorkCenterListForTable(order_no, operation_no)]
    hasOperatorOperatingList = []
    for owc in OWCList:
        hasOperatorOperatingList.append(hasOperatorOperating(owc[0]))
    data = {
        'OWCList': OWCList,
        'hasOperatorOperatingList': hasOperatorOperatingList,
    }
    return JsonResponse(data)

def get_operating_operator_list(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    OOPRList = [list(i) for i in getOperatingOperatorListForTable(order_no, operation_no)]
    data = {
        'OOPRList': OOPRList,
    }
    return JsonResponse(data)

def get_workcenter_data(request):
    routing_no = request.GET.get('routing_no') # Routing
    routing_wcg = request.GET.get('routing_wcg') # Routing
    workcenter_no = request.GET.get('workcenter_no') # Machine
    canAdd = False
    invalid_text = ''
    WorkCenterNo = None
    WorkCenterName = None
    workcenter = getWorkCenter(workcenter_no)
    if workcenter != None:
        WorkCenterNo = workcenter.WorkCenterNo
        WorkCenterName = workcenter.WorkCenterName
        if workcenter.IsRouting:
            invalid_text = WorkCenterNo + " is a rounting."
        elif workcenter.WorkCenterGroup != routing_wcg:
            invalid_text = WorkCenterNo + " is in " + workcenter.WorkCenterGroup + " Group."
        elif workcenter.IsActive == False:
            invalid_text = WorkCenterNo + " is In-Active."
        elif isWorkCenterOperating(workcenter_no):
            owc = getWorkCenterOperatingByWorkCenterNo(workcenter_no)
            invalid_text = WorkCenterNo + " is working at " + owc.OrderNo + "-" + owc.OperationNo + "."
        else:
            canAdd = True
    else:
        canAdd = False
        invalid_text = 'Work Center Not Found'
    data = {
        'canAdd': canAdd,
        'invalid_text' : invalid_text,
        'WorkCenterNo': WorkCenterNo,
        'WorkCenterName': WorkCenterName,
    }
    return JsonResponse(data)

def get_operator_data(request):
    operator_id = request.GET.get('operator_id')
    canAdd = False
    invalid_text = ''
    EmpID = None
    EmpName = None
    Section = None
    operator = getOperator(operator_id)
    if operator != None:
        EmpID = (operator.EmpID).strip()
        EmpName = operator.EmpName
        Section = operator.Section
        if isOperatorOperating(operator_id):
            oopr = getOperatorOperatingByEmpID(operator_id)
            invalid_text = operator_id + ' is working at ' + str(oopr.OperatorOrderNo) + "-" + str(oopr.OperatorOperationNo) + "."
        elif isNotFixedOvertime(operator_id):
            ot = getNotFixedOvertime(operator_id)
            invalid_text = operator_id + ' is not fixed overtime of ' + str(ot.OrderNo) + "-" + str(ot.OperationNo) + "."
        elif operator.IsActive == False:
            invalid_text = operator_id + " is In-Active."
        else:
            canAdd = True
    else:
        canAdd = False
        invalid_text = 'Operator Not Found'
    data = {
        'canAdd': canAdd,
        'invalid_text' : invalid_text,
        'EmpID': EmpID,
        'EmpName': EmpName,
        'Section': Section,
    }
    return JsonResponse(data)

#-------------------------------------------------------------- INNER MAIN TABLE

def add_operating_workcenter(request):
    #-- *** ONLY MACHINE TYPE ***
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    workcenter_no = request.GET.get('workcenter_no')
    #-- RECHECK QTY
    operation = getOperation(order_no, operation_no)
    remainQty = operation.ProcessQty - (operation.AcceptedQty + operation.RejectedQty)
    if remainQty > 0:
        #-- WORKCENTER : ADD
        insertOperatingWorkCenter(order_no, operation_no, workcenter_no)
    data = {
    }
    return JsonResponse(data)

def delete_operating_workcenter(request):
    #-- *** ONLY MACHINE TYPE ***
    id = request.GET.get('id')
    #-- WORKCENTER : DELETE
    deleteOperatingWorkCenter(id)
    data = {
    }
    return JsonResponse(data)

def idle_operating_workcenter(request):

    #-- *** ONLY MACHINE TYPE ***
    id = request.GET.get('id')
    code = request.GET.get('code')
    #-- WORKCENTER : IDLE
    updateOperatingWorkCenter(id, 'IDLE')
    updateIdleCodeOperatingWorkCenter(id, code)
    data = {
    }
    return JsonResponse(data)

def stop_idle_operating_workcenter(request):
    #-- *** ONLY MACHINE TYPE ***
    id = request.GET.get('id')
    #-- WORKCENTER : STOP IDLE
    updateOperatingWorkCenter(id, "COMPLETE")
    owc = getWorkCenterOperatingByID(id) 
    #-- SAVE DATA
    idleTime = int(((owc.StopDateTime - owc.StartDateTime).total_seconds())/60)
    if int(idleTime) > 0:
        insertHistoryOperate(owc.OrderNo, owc.OperationNo, "NULL", owc.WorkCenterNo, "IDLE", 0, 0, 0, idleTime, owc.StartDateTime, owc.StopDateTime, owc.IdleCode)
    #-- GO BACK TO WAITING
    updateOperatingWorkCenter(id, 'WAITING')
    data = {
    }
    return JsonResponse(data)

def add_operating_operator(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    operator_id = request.GET.get('operator_id')
    owc_id = request.GET.get('owc_id')
    status = request.GET.get('status')
    refresh = False
    #-- RECHECK QTY
    operation = getOperation(order_no, operation_no)
    remainQty = operation.ProcessQty - (operation.AcceptedQty + operation.RejectedQty)
    if remainQty > 0:
        #-- INSERT EMP AT COMPUTER (IF NOT EXT-WORK)
        if status != 'EXT-WORK':
            insertEmpAtComputer(operator_id, getClientIP(request))
        #-- OPERATOR : WORKING/SETUP/EXT-WORK
        insertOperatingOperator(order_no, operation_no, operator_id, owc_id, status)
        #-- IF OPERATION IS NOT LABOR TYPE
        if owc_id != "-1":
            #-- IF MACHINE NOT START YET
            owc = getWorkCenterOperatingByID(owc_id)
            if owc.StartDateTime == None:
                #-- WORKCENTER : WORKING/SETUP
                updateOperatingWorkCenter(owc_id, status)
        #-- IF NOT START OPERATION YET
        if hasNotStartOperation(order_no, operation_no):
            refresh = True
            #-- OPERATION : START
            updateOperationControl(order_no, operation_no, 0, 0, "START")
            #-- IF NOT START ORDER YET
            if hasNotStartOrder(order_no):
                #-- ORDER : START
                updateOrderControl(order_no, "START")
        #-- IF JOINING
        joinList = getJoinList(order_no, operation_no)
        for join in joinList:
            if hasNotStartOperation(join.OrderNo, join.OperationNo):
                #-- OPERATION : START
                updateOperationControl(join.OrderNo, join.OperationNo, 0, 0, "START")
                #-- IF NOT START ORDER YET
                if hasNotStartOrder(join.OrderNo):
                    #-- ORDER : START
                    updateOrderControl(join.OrderNo, "START")
    data = {
        'refresh' : refresh,
    }
    return JsonResponse(data)

def start_work_operating_operator(request):
    #-- *** ONLY MACHINE TYPE ***
    id = request.GET.get('id')
    oopr = getOperatorOperatingByID(id)
    if oopr.WorkCenterStatus.strip() == "SETUP":
        #-- OPERATOR : SAVE SETUP TIME
        updateOperatingOperator(id, "COMPLETE")
        #-- WORKCENTER : SAVE SETUP TIME
        oopr = getOperatorOperatingByID(id)
        updateOperatingWorkCenter(oopr.OperatingWorkCenterID, "COMPLETE")
        #-- SAP : SETUP TIME
        oopr = getOperatorOperatingByID(id)
        setuptime = int(((oopr.OperatorStopDateTime - oopr.OperatorStartDateTime).total_seconds())/60)
        #-- IF SETUP TIME IS LESS THAN 1 MIN DON'T SEND DATA TOP SAP
        if int(setuptime > 0):
            insertSFR2SAP_Report(oopr.WorkCenterNo,oopr.OrderNo,oopr.OperationNo,0,0,setuptime,0,0,oopr.OperatorStartDateTime,oopr.OperatorStopDateTime,oopr.EmpID)
        #-- OPERATOR : SETUP TIME LOG
        insertHistoryOperate(oopr.OrderNo,oopr.OperationNo, oopr.EmpID, oopr.WorkCenterNo, "SETUP", setuptime, 0, 0,0, oopr.OperatorStartDateTime, oopr.OperatorStopDateTime, "NULL")
        #-- WORKCENTER : WORKING
        updateOperatingWorkCenter(oopr.OperatingWorkCenterID, "WORKING")
    #-- OPERATOR : WORKING
    updateOperatingOperator(id, "WORKING")
    data = {
    }
    return JsonResponse(data)

def stop_setup_operating_operator(request):
    #-- *** ONLY MACHINE TYPE ***
    id = request.GET.get('id')
    #-- OPERATOR : SAVE SETUP TIME
    updateOperatingOperator(id, "COMPLETE")
    oopr = getOperatorOperatingByID(id)
    #-- DELETE EMP AT COMPUTER
    deleteEmpAtComputer(oopr.EmpID)
    #-- WORKCENTER : SAVE SETUP TIME
    updateOperatingWorkCenter(oopr.OperatingWorkCenterID, "COMPLETE")
    #-- SAP : SETUP TIME
    setuptime = int(((oopr.OperatorStopDateTime - oopr.OperatorStartDateTime).total_seconds())/60)
    #-- IF SETUP TIME IS LESS THAN 1 MIN DON'T SEND DATA TOP SAP
    if int(setuptime > 0):
        insertSFR2SAP_Report(oopr.WorkCenterNo,oopr.OrderNo,oopr.OperationNo,0,0,setuptime,0,0,oopr.OperatorStartDateTime,oopr.OperatorStopDateTime,oopr.EmpID)
    #-- OPERATOR : SETUP TIME LOG
    insertHistoryOperate(oopr.OrderNo,oopr.OperationNo, oopr.EmpID, oopr.WorkCenterNo, "SETUP", setuptime, 0, 0,0, oopr.OperatorStartDateTime, oopr.OperatorStopDateTime, "NULL")
    #-- WORKCENTER : WAITING
    updateOperatingWorkCenter(oopr.OperatingWorkCenterID, "WAITING")
    #-- OPERATOR : EXIT
    deleteOperatingOperator(id)
    #-- CHECK REMAINING IS OPERATING
    IsOperating = isOperatingOperation(oopr.OrderNo, oopr.OperationNo)
    #-- DELETE DOUBLE REPORT
    deleteDoubleRecord(oopr.OrderNo, oopr.OperationNo)
    data = {
        'IsOperating' : IsOperating,
    }
    return JsonResponse(data)

def stop_work_operating_operator(request):
    id = request.GET.get('id')
    oopr = getOperatorOperatingByID(id)
    status = oopr.OperatorStatus
    type = "WORKING"
    #-- DELETE EMP AT COMPUTER
    deleteEmpAtComputer(oopr.EmpID)
    #-- OPERATOR : SAVE WORKING TIME
    updateOperatingOperator(id, "COMPLETE")
    #-- SAP : WORKING TIME
    oopr = getOperatorOperatingByID(id)
    workcenter = oopr.WorkCenterNo
    worktimeOperator = str(int(((oopr.OperatorStopDateTime - oopr.OperatorStartDateTime).total_seconds())/60))
    worktimeMachine = worktimeOperator
    if oopr.WorkCenterNo == None:
        workcenter = getOperation(oopr.OperatorOrderNo, oopr.OperatorOperationNo).WorkCenterNo
        worktimeMachine = 0
    elif oopr.MachineType.strip() == 'Auto':
        worktimeMachine = 0
    elif oopr.MachineType.strip() == 'Manual':
        worktimeMachine = 0
        if hasOperatorOperating(oopr.OperatingWorkCenterID) == False:
            owc = getWorkCenterOperatingByID(oopr.OperatingWorkCenterID)
            worktimeMachine = str(int(((oopr.OperatorStopDateTime - owc.StartDateTime).total_seconds())/60))
    if status == "EXT-WORK":
        worktimeOperator = 0
    #-- IF EXTERNAL PROCESS DONT SEND DATA TO SAP (COMFIRMATION WILL HAVE ALL THIS INFO)
    #-- IF WORK TIME IS LESS THAN 1 MIN DON'T SEND DATA TO SAP
    if status != "EXT-WORK" and (int(worktimeMachine) > 0 or int(worktimeOperator) > 0):
        insertSFR2SAP_Report(workcenter,oopr.OperatorOrderNo,oopr.OperatorOperationNo,0,0,0,worktimeMachine,worktimeOperator,oopr.OperatorStartDateTime,oopr.OperatorStopDateTime,oopr.EmpID)
    #-- OPERATOR : OPERATING TIME LOG
    insertHistoryOperate(oopr.OperatorOrderNo,oopr.OperatorOperationNo, oopr.EmpID, workcenter, type, 0, worktimeMachine, worktimeOperator,0, oopr.OperatorStartDateTime, oopr.OperatorStopDateTime, "NULL")
    #-- IF OPERATION IS NOT LABOR TYPE & NO OPERATOR WORKING & WORKCENTER IS MANUAL
    if oopr.OperatingWorkCenterID != None and hasOperatorOperating(oopr.OperatingWorkCenterID) == False and oopr.MachineType.strip() == 'Manual':
        #-- WORKCENTER : STOP
        updateOperatingWorkCenter(oopr.OperatingWorkCenterID, "COMPLETE")
    #-- CHECK REMAINING IS OPERATING
    IsOperating = isOperatingOperation(oopr.OperatorOrderNo, oopr.OperatorOperationNo)
    #-- DELETE DOUBLE REPORT
    deleteDoubleRecord(oopr.OperatorOrderNo, oopr.OperatorOperationNo)
    data = {
        'IsOperating' : IsOperating,
    }
    return JsonResponse(data)

def stop_operating_workcenter(request):
    #-- *** ONLY AUTO MACHINE TYPE ***
    id = request.GET.get('id')
    #-- WORKCENTER : STOP
    updateOperatingWorkCenter(id, "COMPLETE")
    #-- SAP : WORKING TIME
    owc = getWorkCenterOperatingByID(id)
    worktimeMachine = int(((owc.StopDateTime - owc.StartDateTime).total_seconds())/60)
    #-- IF WORK TIME IS LESS THAN 1 MIN DON'T SEND DATA TOP SAP
    if int(worktimeMachine) > 0:
        insertSFR2SAP_Report(owc.WorkCenterNo,owc.OrderNo,owc.OperationNo,0,0,0,worktimeMachine,0,owc.StartDateTime,owc.StopDateTime,'9999')
    #-- WORKCENTER : OPERATING TIME LOG
    insertHistoryOperate(owc.OrderNo,owc.OperationNo, "NULL", owc.WorkCenterNo, "WORKING", 0, worktimeMachine, 0, 0, owc.StartDateTime, owc.StopDateTime, "NULL")
    #-- CHECK REMAINING IS OPERATING
    IsOperating = isOperatingOperation(owc.OrderNo, owc.OperationNo)
    #-- DELETE DOUBLE REPORT
    deleteDoubleRecord(owc.OrderNo, owc.OperationNo)
    data = {
        'IsOperating' : IsOperating,
    }
    return JsonResponse(data)

#-----------------------------------------CONFIRM & MANUAL REPORT & FIX OVERTIME

def get_data_for_confirm(request):
    id = request.GET.get('id')
    oopr = getOperatorOperatingByID(id)
    operator_text = oopr.EmpID.strip() + " | " + oopr.EmpName
    workcenter_text = ""
    if oopr.WorkCenterNo != None:
        workcenter = getWorkCenter(oopr.WorkCenterNo)
        workcenter_text = workcenter.WorkCenterNo + " | " + workcenter.WorkCenterName
    start_time = oopr.OperatorStartDateTime.strftime("%d-%m-%Y, %H:%M:%S")
    stop_time = oopr.OperatorStopDateTime.strftime("%d-%m-%Y, %H:%M:%S")
    data = {
        'operator_text': operator_text,
        'workcenter_text': workcenter_text,
        'start_time': start_time,
        'stop_time': stop_time,
    }
    return JsonResponse(data)

def confirm(request):
    confirm_id = request.GET.get('confirm_id')
    good_qty = request.GET.get('good_qty')
    reject_qty = request.GET.get('reject_qty')
    reject_reason = request.GET.get('reject_reason')
    other_reason = request.GET.get('other_reason')
    scrap_at = request.GET.get('scrap_at')
    reject_reason.replace("'", "")
    if reject_reason == "-1" or reject_qty == 0:
        reject_reason = ""
    elif reject_reason == "OTHER":
        reject_reason = other_reason
    if reject_reason != "SCRAP FROM PREVIOUS PROCESS":
        scrap_at = ""
    #-- RECHECK QTY
    oopr = getOperatorOperatingByID(confirm_id)
    orderNo = oopr.OperatorOrderNo
    operationNo = oopr.OperatorOperationNo
    workcenter = oopr.WorkCenterNo
    operation = getOperation(orderNo, operationNo)
    remainQty = operation.ProcessQty - (operation.AcceptedQty + operation.RejectedQty)
    if remainQty >= (int(good_qty) + int(reject_qty)):
        if oopr.WorkCenterNo == None:
            workcenter = getOperation(orderNo, operationNo).WorkCenterNo
        #-- UPDATE QTY OF CURRENT OPERATION
        updateOperationControl(orderNo,operationNo, good_qty, reject_qty, "UPDATEQTY")
        #-- UPDATE PROCESS QTY OF NEXT OPERATION
        nextOperation = getNextOperation(orderNo,operationNo)
        #-- CONFIRM : LOG
        insertHistoryConfirm(orderNo,operationNo, oopr.EmpID, workcenter, good_qty, reject_qty, reject_reason, scrap_at)
        #-- SAP : CONFIRM
        insertSFR2SAP_Report(workcenter,orderNo,operationNo,good_qty,reject_qty,0,0,0,oopr.OperatorStartDateTime,oopr.OperatorStopDateTime,oopr.EmpID)
        #-- TOOLLIST : CONFIRM
        runToolConfirm(orderNo, operationNo, workcenter, good_qty, reject_qty)
        if nextOperation != None:
            updateOperationControl(nextOperation.OrderNo,nextOperation.OperationNo, good_qty, 0, "PROCESSQTY")
        #-- NO MORE REMAINING QTY
        operation = getOperation(orderNo, operationNo)
        remainQty = operation.ProcessQty - (operation.AcceptedQty + operation.RejectedQty)
        if remainQty == 0:
            #-- IF AUTO MACHINE(S) STILL WORKING
            owcList = getOperatingWorkCenterList(orderNo, operationNo)
            for owc in owcList:
                if owc.Status == 'WORKING' and owc.MachineType.strip() == 'Auto':
                    #-- WORKCENTER : STOP
                    updateOperatingWorkCenter(owc.OperatingWorkCenterID, "COMPLETE")
                    #-- SAP : WORKING TIME
                    owc = getWorkCenterOperatingByID(owc.OperatingWorkCenterID)
                    worktimeMachine = str(int(((owc.StopDateTime - owc.StartDateTime).total_seconds())/60))
                    #-- IF WORK TIME IS LESS THAN 1 MIN DON'T SEND DATA TO SAP
                    if int(worktimeMachine) > 0:
                        insertSFR2SAP_Report(owc.WorkCenterNo,owc.OrderNo,owc.OperationNo,0,0,0,worktimeMachine,0,owc.StartDateTime,owc.StopDateTime,'9999')
                    #-- WORKCENTER : OPERATING TIME LOG
                    insertHistoryOperate(owc.OrderNo,owc.OperationNo, "NULL", owc.WorkCenterNo, "WORKING",0,worktimeMachine,0,0, owc.StartDateTime, owc.StopDateTime, "NULL")
            #-- CLEAR ALL CONTROL DATA
            deleteAllOperatingData(orderNo, operationNo)
            #-- STOP OPERATION
            updateOperationControl(orderNo, operationNo, 0, 0, "STOP")
            #-- IF LAST OPERATION IN ORDER or IF NO MORE REMAINING QTY IN ORDER
            hasNoMoreQty = True
            operationList = getOperationList(orderNo)
            for i in range(len(operationList)):
                tempRemainQty = operationList[i].ProcessQty - (operationList[i].AcceptedQty + operationList[i].RejectedQty)
                if tempRemainQty > 0:
                    hasNoMoreQty = False
                    break
            if hasNoMoreQty:
                #-- ORDER : STOP
                updateOrderControl(orderNo, "STOP")
                #-- PENDING PLN_FAI
                if operation.WorkCenterNo.strip() == 'PLN_FAI':
                    addPendingPLNFAI(operation.OrderNo, operation.OperationNo, operation.ProcessQty)
    data = {
    }
    return JsonResponse(data)

def manual_report(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    emp_id = request.GET.get('emp_id')
    workcenter_no = request.GET.get('workcenter_no')
    labor_time = request.GET.get('labor_time')
    operate_time = request.GET.get('operate_time')
    setup_time = request.GET.get('setup_time')
    start_time = parser.parse(request.GET.get('start_time'))
    stop_time = parser.parse(request.GET.get('stop_time'))
    good_qty = request.GET.get('good_qty')
    reject_qty = request.GET.get('reject_qty')
    reject_reason = request.GET.get('reject_reason')
    other_reason = request.GET.get('other_reason')
    scrap_at = request.GET.get('scrap_at')
    reject_reason.replace("'", "")
    if reject_reason == "-1" or reject_qty == 0:
        reject_reason = ""
    elif reject_reason == "OTHER":
        reject_reason = other_reason
    if reject_reason != "SCRAP FROM PREVIOUS PROCESS":
        scrap_at = ""
    #-- RECHECK QTY
    operation = getOperation(order_no, operation_no)
    remainQty = operation.ProcessQty - (operation.AcceptedQty + operation.RejectedQty)
    if remainQty >= (int(good_qty) + int(reject_qty)):
        #-- IF NOT START OPERATION YET
        if hasNotStartOperation(order_no, operation_no):
            #-- OPERATION : START
            updateOperationControl(order_no, operation_no, 0, 0, "START")
            #-- IF NOT START ORDER YET
            if hasNotStartOrder(order_no):
                #-- ORDER : START
                updateOrderControl(order_no, "START")
        #-- MANUAL REPORT : LOG
        insertHistoryOperate(order_no, operation_no, emp_id, workcenter_no, "MANUAL", setup_time, operate_time, labor_time,0, start_time, stop_time, "NULL")
        # SAP : CONFIRM TIME & QTY
        insertSFR2SAP_Report(workcenter_no,order_no,operation_no,good_qty,reject_qty,setup_time,operate_time,labor_time,start_time,stop_time,emp_id)
        #-- TOOLLIST : CONFIRM
        runToolConfirm(order_no, operation_no, workcenter_no, good_qty, reject_qty)
        #-- CLEAR OVERTIME IS FIXED
        fixOvertimeReported(emp_id)
        #-- CONFIRM : LOG
        if int(good_qty) > 0 or int(reject_qty) > 0:
            #-- UPDATE QTY OF CURRENT OPERATION
            updateOperationControl(order_no, operation_no, good_qty, reject_qty, "UPDATEQTY")
            #-- UPDATE PROCESS QTY OF NEXT OPERATION
            nextOperation = getNextOperation(order_no, operation_no)
            # -- ADD HISTORY LOG
            insertHistoryConfirm(order_no, operation_no, emp_id, workcenter_no, good_qty, reject_qty, reject_reason, scrap_at)
            if nextOperation != None:
                updateOperationControl(nextOperation.OrderNo,nextOperation.OperationNo, good_qty, 0, "PROCESSQTY")
            #-- NO MORE REMAINING QTY
            operation = getOperation(order_no, operation_no)
            remainQty = operation.ProcessQty - (operation.AcceptedQty + operation.RejectedQty)
            if remainQty == 0:
                #-- IF AUTO MACHINE(S) STILL WORKING
                owcList = getOperatingWorkCenterList(order_no, operation_no)
                for owc in owcList:
                    if owc.Status == 'WORKING':
                        #-- WORKCENTER : STOP
                        updateOperatingWorkCenter(owc.OperatingWorkCenterID, "COMPLETE")
                        #-- SAP : WORKING TIME
                        owc = getWorkCenterOperatingByID(owc.OperatingWorkCenterID)
                        worktimeMachine = str(int(((owc.StopDateTime - owc.StartDateTime).total_seconds())/60))
                        #-- IF WORK TIME IS LESS THAN 1 MIN DON'T SEND DATA TOP SAP
                        if int(worktimeMachine) > 0:
                            insertSFR2SAP_Report(owc.WorkCenterNo,owc.OrderNo,owc.OperationNo,0,0,0,worktimeMachine,0,owc.StartDateTime,owc.StopDateTime,'9999')
                        #-- WORKCENTER : OPERATING TIME LOG
                        insertHistoryOperate(owc.OrderNo,owc.OperationNo, "NULL", owc.WorkCenterNo, "WORKING",0,worktimeMachine,0, owc.StartDateTime, owc.StopDateTime, "NULL")
                #-- CLEAR ALL CONTROL DATA
                deleteAllOperatingData(order_no, operation_no)
                #-- STOP OPERATION
                updateOperationControl(order_no, operation_no, 0, 0, "STOP")
                #-- IF LAST OPERATION IN ORDER or IF NO MORE REMAINING QTY IN ORDER
                hasNoMoreQty = True
                operationList = getOperationList(order_no)
                for i in range(len(operationList)):
                    tempRemainQty = operationList[i].ProcessQty - (operationList[i].AcceptedQty + operationList[i].RejectedQty)
                    if tempRemainQty > 0:
                        hasNoMoreQty = False
                        break
                if hasNoMoreQty:
                    #-- ORDER : STOP
                    updateOrderControl(order_no, "STOP")
                    #-- PENDING PLN_FAI
                    if operation.WorkCenterNo.strip() == 'PLN_FAI':
                        addPendingPLNFAI(operation.OrderNo, operation.OperationNo, operation.ProcessQty)
    data = {
    }
    return JsonResponse(data)

def fix_overtime(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    operation = getOperation(order_no, operation_no)
    operatingOperatorList = getOperatingOperatorList(order_no, operation_no)
    for oopr in operatingOperatorList:
        if isOvertimeOperator(oopr.OperatingOperatorID):
            insertOvertimeOperator(oopr)
    if operation.WorkCenterType.strip() == 'Machine' and operation.MachineType.strip() == 'Auto':
        operatingWorkCenterList = getOperatingWorkCenterList(order_no, operation_no)
        for owc in operatingWorkCenterList:
            #-- ONLY WORKING WORKCENTER
            if isOvertimeWorkCenter(owc.OperatingWorkCenterID):
                insertOvertimeWorkCenter(owc)
    #-- DELETE ALL OPERATING DATA
    deleteAllOperatingData(order_no, operation_no)
    data = {

    }
    return JsonResponse(data)

#-------------------------------------------------------------------------- JOIN

def join(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    join_list = request.GET.getlist('join_list[]')
    #-- CLEAR ALL CONTROL DATA OF JOIN PROCESS (MAIN)
    deleteAllOperatingData(order_no, operation_no)
    for join_item in join_list:
        join_order_no = join_item[0:10]
        join_operation_no = join_item[10:14]
        #-- JOIN PROCESS
        joinProcess(order_no, operation_no, join_order_no, join_operation_no)
        #-- CLEAR ALL CONTROL DATA OF JOIN PROCESS
        deleteAllOperatingData(join_order_no, join_operation_no)
    data = {
    }
    return JsonResponse(data)

def break_join(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    operation = getOperation(order_no, operation_no)
    joinList = getJoinList(order_no, operation_no)
    operating_workcenter_list = getOperatingWorkCenterList(order_no, operation_no)
    operating_operator_list = getOperatingOperatorList(order_no, operation_no)
    for join in joinList:
        #-- COPY COMPLETE OPERATING OPERATION TO JOINING OPERATION
        for owc in operating_workcenter_list:
            owc_id = getInsertJoinOperatingWorkCenter(join.OrderNo, join.OperationNo, owc.WorkCenterNo, owc.StartDateTime, owc.StopDateTime)
            for oopr in operating_operator_list:
                if oopr.OperatingWorkCenterID == owc.OperatingWorkCenterID:
                    insertJoinOperatingOperator(join.OrderNo, join.OperationNo, oopr.EmpID, owc_id, oopr.StartDateTime, oopr.StopDateTime)
        #-- REMOVE JOIN
        joinProcessRemove(join.OrderNo, join.OperationNo)
        #-- HISTORY : JOIN
        insertHistoryJoin(order_no, operation_no, join.OrderNo, join.OperationNo, join.JoinStartDateTime)
    data = {
    }
    return JsonResponse(data)

#---------------------------------------------------------------------- MODIFIER

def inc_qty(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    emp_id = request.GET.get('emp_id')
    password = request.GET.get('password')
    amount = request.GET.get('amount')
    operation = getOperation(order_no, operation_no)
    #-- INCREASE QTY OF 1st OPERATION & ORDER
    increaseQty(order_no, operation_no, amount)
    #-- HISTORY : INCREASE QTY
    user = getUserByPassword(password)
    insertHistoryModifier("INC QTY", order_no, operation_no, emp_id, user.UserID)
    data = {
    }
    return JsonResponse(data)

def dec_qty(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    emp_id = request.GET.get('emp_id')
    password = request.GET.get('password')
    amount = request.GET.get('amount')
    operation = getOperation(order_no, operation_no)
    #-- DECREASE QTY OF 1st OPERATION & ORDER
    decreaseQty(order_no, operation_no, amount)
    #-- HISTORY : DECREASE QTY
    user = getUserByPassword(password)
    insertHistoryModifier("DEC QTY", order_no, operation_no, emp_id, user.UserID)
    data = {
    }
    return JsonResponse(data)

def delete_operation(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    emp_id = request.GET.get('emp_id')
    password = request.GET.get('password')
    operation = getOperation(order_no, operation_no)
    nextoolink = "0"
    #-- CLEAR DATA MIGHT LEFT (LIKE WAITING WORKCENTER)
    deleteAllOperatingData(order_no, operation_no)
    #-- TRANSFER PROCESS QTY TO NEXT OPERATION
    nextOperation = getNextOperation(order_no, operation_no)
    if nextOperation != None:
        nextoolink = nextOperation.OrderNo + nextOperation.OperationNo
        updateOperationControl(nextOperation.OrderNo,nextOperation.OperationNo, operation.ProcessQty, 0, "PROCESSQTY")
    else:
        #-- ORDER : STOP
        updateOrderControl(operation_no, "STOP")
    #-- SAP MODIFIER : DELETE OPERATION
    insertSFR2SAP_Modifier_Delete(order_no, operation_no)
    #-- HISTORY : DELETE OPERATION
    user = getUserByPassword(password)
    insertHistoryModifier("DELETE", order_no, operation_no, emp_id, user.UserID)
    #-- DELETE THIS OPERATION
    deleteOperationControl(order_no, operation_no)
    #-- IF NO MORE REMAINING QTY IN ORDER
    hasNoMoreQty = True
    operationList = getOperationList(order_no)
    for i in range(len(operationList)):
        tempRemainQty = operationList[i].ProcessQty - (operationList[i].AcceptedQty + operationList[i].RejectedQty)
        if tempRemainQty > 0:
            hasNoMoreQty = False
            break
    if hasNoMoreQty:
        #-- ORDER : STOP
        updateOrderControl(order_no, "STOP")
    #-- IF NO MORE ROUTING: DELETE ALL THING
    operationList = getOperationList(order_no)
    if len(operationList) == 0:
        deleteAllSFRAndSAPOrder(order_no)
    data = {
        'nextoolink' : nextoolink,
    }
    return JsonResponse(data)

def add_operation(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('new_operation_no')
    emp_id = request.GET.get('emp_id')
    work_center_no = request.GET.get('work_center_no')
    plan_start_date = request.GET.get('plan_start_date')
    plan_finish_date = request.GET.get('plan_finish_date')
    control_key = request.GET.get('control_key')
    est_setup_time = request.GET.get('est_setup_time')
    est_operate_time = request.GET.get('est_operate_time')
    est_labor_time = request.GET.get('est_labor_time')
    pdt = request.GET.get('pdt')
    purchasing_org = request.GET.get('purchasing_org')
    cost_element = request.GET.get('cost_element')
    mat_group = request.GET.get('mat_group')
    purchasing_group = request.GET.get('purchasing_group')
    price_unit = request.GET.get('price_unit')
    price = request.GET.get('price')
    currency = request.GET.get('currency')
    password = request.GET.get('password')
    #-- ADD OPERATION CONTROL
    if control_key == "PP01":
        insertOperationControl(order_no, operation_no, work_center_no, plan_start_date, plan_finish_date, est_setup_time, est_operate_time, est_labor_time)
    else:
        insertOperationControl(order_no, operation_no, work_center_no, plan_start_date, plan_finish_date, 0, 0, 0)
    #-- SAP : ADD OPERATION
    insertSFR2SAP_Modifier_Add(order_no, operation_no, control_key, work_center_no, pdt, cost_element, price_unit, price, currency, mat_group, purchasing_group, purchasing_org, est_setup_time, est_operate_time, est_labor_time)
    #-- HISTORY : ADD OPERATION
    user = getUserByPassword(password)
    insertHistoryModifier("ADD", order_no, operation_no, emp_id, user.UserID)
    #-- IF NEXT OPERATION HAS PROCESS QTY TRANSFER TO NEW OPERATION
    nextOperation = getNextOperation(order_no, operation_no)
    if nextOperation != None:
        updateOperationControl(order_no, operation_no, nextOperation.ProcessQty, 0, "PROCESSQTY")
        updateOperationControl(order_no, nextOperation.OperationNo, (nextOperation.ProcessQty * -1), 0, "PROCESSQTY")
        deleteAllOperatingData(order_no, nextOperation.OperationNo)
    data = {
    }
    return JsonResponse(data)

def change_operation(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    emp_id = request.GET.get('emp_id')
    work_center_no = request.GET.get('work_center_no')
    plan_start_date = request.GET.get('plan_start_date')
    plan_finish_date = request.GET.get('plan_finish_date')
    control_key = request.GET.get('control_key')
    est_setup_time = request.GET.get('est_setup_time')
    est_operate_time = request.GET.get('est_operate_time')
    est_labor_time = request.GET.get('est_labor_time')
    pdt = request.GET.get('pdt')
    purchasing_org = request.GET.get('purchasing_org')
    cost_element = request.GET.get('cost_element')
    mat_group = request.GET.get('mat_group')
    purchasing_group = request.GET.get('purchasing_group')
    price_unit = request.GET.get('price_unit')
    price = request.GET.get('price')
    currency = request.GET.get('currency')
    password = request.GET.get('password')
    #-- CLEAR DATA MIGHT LEFT (LIKE WAITING WORKCENTER)
    deleteAllOperatingData(order_no, operation_no)
    #-- SAP REQUIRED TO SEND ONLY 1 CHANGE OF THE SAME OPERATION PER HOURLY FILE
    deleteSFR2SAP_Modifier_Change(order_no, operation_no)
    #-- CHANGE OPERATION CONTROL
    if control_key == "PP01":
        changeOperationControl(order_no, operation_no, work_center_no, plan_start_date, plan_finish_date, est_setup_time, est_operate_time, est_labor_time)
    else:
        est_setup_time = 0
        est_operate_time = 0
        est_labor_time = 0
        changeOperationControl(order_no, operation_no, work_center_no, plan_start_date, plan_finish_date, 0, 0, 0)
    #-- SAP : CHANGE OPERATION
    insertSFR2SAP_Modifier_Change(order_no, operation_no, control_key, work_center_no, pdt, cost_element, price_unit, price, currency, mat_group, purchasing_group, purchasing_org, est_setup_time, est_operate_time, est_labor_time)
    #-- HISTORY : CHANGE OPERATION
    user = getUserByPassword(password)
    insertHistoryModifier("CHANGE", order_no, operation_no, emp_id, user.UserID)
    data = {
    }
    return JsonResponse(data)

#-------------------------------------------------------------------------- NOTE

def save_note(request):
    order_no = request.GET.get('order_no')
    order_note = request.GET.get('order_note')
    operation_note_list = request.GET.getlist('operation_note_list[]')
    weight = request.GET.get('weight')
    size = request.GET.get('size')
    operationList = getOperationList(order_no)
    updateOrderNote(order_no, order_note, weight, size)
    for i in range(len(operationList)):
        updateOperationNote(order_no, operationList[i].OperationNo, operation_note_list[i])
    data = {
    }
    return JsonResponse(data)

def save_over_est_note(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    over_est_note = request.GET.get('over_est_note').strip()
    updateOverEstNote(order_no, operation_no, over_est_note)
    data = {
    }
    return JsonResponse(data)

#-------------------------------------------------------------------- VALIDATION
def validate_new_operation(request):
    order_no = request.GET.get('order_no')
    new_operation_no = request.GET.get('new_operation_no')
    canAdd = True
    #1
    if isExistOperation(order_no, new_operation_no):
        canAdd = False
    #2
    operationList = getOperationList(order_no)
    for i in range(len(operationList)):
        if operationList[i].ProcessStart != None and new_operation_no < operationList[i].OperationNo:
            canAdd = False
            break
    #3
    if isExistDeletedOperation(order_no, new_operation_no):
        canAdd = False
    data = {
        'canAdd': canAdd,
    }
    return JsonResponse(data)

def validate_routing(request):
    work_center_no = request.GET.get('work_center_no')
    canUse = True
    invalidText = ""
    isExt = False
    work_center = getWorkCenter(work_center_no)
    if work_center == None:
        canUse = False
        invalidText = "Routing not found."
    elif work_center.IsRouting == False:
        canUse = False
        invalidText = "This Work Center is not a routing."
    elif work_center.IsActive == False:
        canUse = False
        invalidText = "This Work Center is In-Active."
    else:
        isExt = work_center.IsExternalProcess
    data = {
        'canUse': canUse,
        'invalidText': invalidText,
        'isExt': isExt,
    }
    return JsonResponse(data)

def validate_work_center(request):
    work_center_no = request.GET.get('work_center_no')
    work_center_group = request.GET.get('work_center_group')
    canUse = True
    invalidText = ''
    work_center = getWorkCenter(work_center_no)
    if work_center == None:
        canUse = False
        invalidText = 'Work Center Not Found.'
    elif work_center.IsRouting:
        canUse = False
        invalidText = 'Routing can not be used.'
    elif work_center.WorkCenterGroup != work_center_group:
        canUse = False
        invalidText = 'This Work Center is not in the same Work Center Group.'
    elif work_center.IsActive == False:
        canUse = False
        invalidText = 'This Work Center is In-Active.'
    data = {
        'canUse': canUse,
        'invalidText': invalidText,
    }
    return JsonResponse(data)

def validate_wc_no(request):
    wc_no = request.GET.get('wc_no').strip()
    canUse = True
    if isExistWorkCentrNo(wc_no):
        canUse = False
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_operator(request):
    emp_id = request.GET.get('emp_id')
    isExist = isExistOperator(emp_id)
    data = {
        'isExist': isExist,
    }
    return JsonResponse(data)

def validate_password(request):
    password = request.GET.get('password')
    isCorrect = False
    user = getUserByPassword(password)
    if user != None:
        isCorrect = True
    data = {
        'isCorrect': isCorrect,
    }
    return JsonResponse(data)

def validate_section_chief_password(request):
    password = request.GET.get('password')
    isCorrect = False
    user = getUserByPassword(password)
    if user != None and (user.UserRole.strip() == 'CHIEF' or user.UserRole.strip() == 'ADMIN' or user.UserRole.strip() == 'SUPERADMIN'):
        isCorrect = True
    data = {
        'isCorrect': isCorrect,
    }
    return JsonResponse(data)

def validate_admin_password(request):
    password = request.GET.get('password')
    isCorrect = False
    user = getUserByPassword(password)
    if user != None and (user.UserRole.strip() == 'ADMIN' or user.UserRole.strip() == 'SUPERADMIN'):
        isCorrect = True
    data = {
        'isCorrect': isCorrect,
    }
    return JsonResponse(data)

def validate_super_admin_password(request):
    password = request.GET.get('password')
    isCorrect = False
    user = getUserByPassword(password)
    if user != None and user.UserRole.strip() == 'SUPERADMIN':
        isCorrect = True
    data = {
        'isCorrect': isCorrect,
    }
    return JsonResponse(data)

def validate_new_user_id(request):
    user_id = request.GET.get('user_id')
    canUse = False
    user = getUserByUserID(user_id)
    if user == None:
        canUse = True
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_new_password(request):
    password = request.GET.get('password')
    canUse = False
    user = getUserByPassword(password)
    if user == None:
        canUse = True
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_upload_tool_list(request):
    order_no = request.GET.get('order_no').strip()
    operation_no = request.GET.get('operation_no').strip()
    wc_no = request.GET.get('wc_no').strip()
    canUse = False
    invalidText = ''
    if not isExistOperation(order_no, operation_no):
        invalidText = f'{order_no}-{operation_no} is not exist.'
    elif not isExistWorkCentrNo(wc_no):
        invalidText = 'Work Center is not exist.'
    elif not isMachine(wc_no):
        invalidText = 'Work Center is not a Machine No.'
    elif isExistToolHeader(order_no, operation_no, wc_no):
        invalidText = f'Tool List of {wc_no} is already exist in {order_no}-{operation_no}'
    else:
        canUse = True
    data = {
        'canUse': canUse,
        'invalidText': invalidText,
    }
    return JsonResponse(data)

def validate_transfer_tool(request):
    order_no = request.GET.get('order_no').strip()
    operation_no = request.GET.get('operation_no').strip()
    toolh_id = request.GET.get('toolh_id').strip()
    toolh = getToolHeaderByID(toolh_id)
    canUse = False
    invalidText = ''
    if not isExistOperation(order_no, operation_no):
        invalidText = f'{order_no}-{operation_no} is not exist.'
    elif isExistToolHeader(order_no, operation_no, toolh.WorkCenterNo):
        invalidText = f'Tool List of {toolh.WorkCenterNo} is already exist in {order_no}-{operation_no}'
    else:
        prod_1 = getOrder(toolh.OrderNo)
        prod_2 = getOrder(order_no)
        if prod_1.FG_MaterialCode != prod_2.FG_MaterialCode:
            invalidText = f'FG Material Code not match.'
        else:
            canUse = True
    data = {
        'canUse': canUse,
        'invalidText': invalidText,
    }
    return JsonResponse(data)

def validate_change_tool_wc(request):
    toolh_id = request.GET.get('toolh_id').strip()
    wc_no = request.GET.get('wc_no').strip()
    toolh = getToolHeaderByID(toolh_id)
    canUse = False
    invalidText = ''
    order_no = toolh.OrderNo
    operation_no = toolh.OperationNo
    if not isExistWorkCentrNo(wc_no):
        invalidText = 'Work Center is not exist.'
    elif not isMachine(wc_no):
        invalidText = 'Work Center is not a Machine No.'
    elif isExistToolHeader(order_no, operation_no, wc_no):
        invalidText = f'Tool List of {wc_no} is already exist in {order_no}-{operation_no}'
    else:
        canUse = True
    data = {
        'canUse': canUse,
        'invalidText': invalidText,
    }
    return JsonResponse(data)

def validate_add_new_tool(request):
    toolh_id = request.GET.get('toolh_id').strip()
    no = request.GET.get('no').strip()
    toolItemList = getToolItemList(toolh_id)
    canUse = True
    invalidText = ''
    for tooli in toolItemList:
        if tooli.No == float(no):
            canUse = False
            invalidText = f'No {no} already has tool'
            break
    data = {
        'canUse': canUse,
        'invalidText': invalidText,
    }
    return JsonResponse(data)

#------------------------------------------------------------------- ADMIN PANEL
def add_new_user(request):
    user_id = request.GET.get('user_id')
    user_password = request.GET.get('user_password')
    user_role = request.GET.get('user_role')
    insertUser(user_id, user_password, user_role)
    data = {
    }
    return JsonResponse(data)

def delete_user(request):
    user_id = request.GET.get('user_id')
    deleteUser(user_id)
    data = {
    }
    return JsonResponse(data)

def change_user_password(request):
    user_id = request.GET.get('user_id')
    user_password = request.GET.get('user_password')
    changeUserPassword(user_id, user_password)
    data = {
    }
    return JsonResponse(data)

def mpa(request):
    status = request.GET.get('status')
    updateManualReportAllowdance(status)
    data = {
    }
    return JsonResponse(data)

def add_wc(request):
    type = request.GET.get('type')
    wc_no = request.GET.get('wc_no').strip()
    wc_name = request.GET.get('wc_name').strip()
    wcg = request.GET.get('wcg').strip()
    on_rt = request.GET.get('on_rt').strip()
    target = request.GET.get('target').strip()
    capacity = request.GET.get('capacity').strip()
    pfc = request.GET.get('pfc').strip()
    active_date = datetime.today().strftime('%Y-%m') + "-01"
    insertWorkCenter(type, wc_no, wc_name, wcg, on_rt, target, capacity, pfc, active_date, DEFAULT_INACTIVE_DATE)
    data = {
    }
    return JsonResponse(data)

def reset_order(request):
    order_no = request.GET.get('order_no')
    start_operation_no = request.GET.get('start_operation_no')
    conn = get_connection()
    cursor = conn.cursor()
    sql = " DELETE FROM OperatingOperator WHERE OrderNo = "+order_no+" "
    sql += " DELETE FROM OperatingWorkCenter WHERE OrderNo = "+order_no+" "
    sql += " DELETE FROM HistoryConfirm WHERE OrderNo = "+order_no+" "
    sql += " DELETE FROM HistoryOperate WHERE OrderNo = "+order_no+" "
    sql += " DELETE FROM HistoryJoin WHERE JoinToOrderNo = "+order_no+" "
    sql += " DELETE FROM HistoryJoin WHERE JoinByOrderNo = "+order_no+" "
    sql += " DELETE FROM HistoryModifier WHERE OrderNo = "+order_no+" "
    sql += " DELETE FROM OvertimeOperator WHERE OrderNo = "+order_no+" "
    sql += " DELETE FROM OvertimeWorkCenter WHERE OrderNo = "+order_no+" "
    sql += " DELETE FROM OrderControl WHERE OrderNo = "+order_no+" "
    sql += " DELETE FROM OperationControl WHERE OrderNo = "+order_no+" "
    sql += " DELETE FROM PartialLotTraveller WHERE OrderNo = "+order_no+" "
    # sql += " DELETE FROM SFR2SAP_Report WHERE ProductionOrderNo = "+order_no+" "
    sql += " DELETE SAP_Routing WHERE ProductionOrderNo = '"+order_no+"' AND OperationNumber < " + str(start_operation_no)
    cursor.execute(sql)
    conn.commit()
    data = {
    }
    return JsonResponse(data)

def cancel_order(request):
    order_no = request.GET.get('order_no')
    emp_id = request.GET.get('emp_id').strip()
    reason = request.GET.get('reason').strip()
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO CanceledOrder ([OrderNo],[EmpID],[Reason],[DateTimeStamp]) VALUES ('"+order_no+"','"+emp_id+"','"+reason+"',CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    data = {
    }
    return JsonResponse(data)

def sqty_operation(request):
    order_no = request.GET.get('order_no')
    operation_no = request.GET.get('operation_no')
    process_qty = int(request.GET.get('process_qty'))
    accepted_qty = int(request.GET.get('accepted_qty'))
    rejected_qty = int(request.GET.get('rejected_qty'))
    if process_qty > 0:
        #-- UPDATE QTY OPERATION CONTROL
        setOperationQty(order_no, operation_no, process_qty, accepted_qty, rejected_qty)
        #-- IF NOT START OPERATION YET
        if hasNotStartOperation(order_no, operation_no):
            #-- OPERATION : START
            updateOperationControl(order_no, operation_no, 0, 0, "START")
            #-- IF NOT START ORDER YET
            if hasNotStartOrder(order_no):
                #-- ORDER : START
                updateOrderControl(order_no, "START")
        #-- CHECK IF FULL CONFIRM QTY
        if process_qty - (accepted_qty + rejected_qty) <= 0:
            #-- CLEAR ALL CONTROL DATA
            deleteAllOperatingData(order_no, operation_no)
            #-- STOP OPERATION
            updateOperationControl(order_no, operation_no, 0, 0, "STOP")
            #-- IF LAST OPERATION IN ORDER or IF NO MORE REMAINING QTY IN ORDER
            hasNoMoreQty = True
            operationList = getOperationList(order_no)
            for i in range(len(operationList)):
                tempRemainQty = operationList[i].ProcessQty - (operationList[i].AcceptedQty + operationList[i].RejectedQty)
                if tempRemainQty > 0:
                    hasNoMoreQty = False
                    break
            if isLastOperation(order_no, operation_no) and hasNoMoreQty:
                #-- ORDER : STOP
                updateOrderControl(order_no, "STOP")
    data = {
    }
    return JsonResponse(data)

#--------------------------------------------------------------------------- ETC

def increase_lot_no(request):
    order_no = request.GET.get('order_no')
    start_operation_no = request.GET.get('start_operation_no')
    final_lot_no = request.GET.get('final_lot_no')
    lot_qty = request.GET.get('lot_qty')
    type = request.GET.get('type')
    password = request.GET.get('password')
    lot_no = int(final_lot_no) + 1
    chief_id = getUserByPassword(password).UserID
    updateOrderLotNo(order_no)
    insertPLT(order_no, start_operation_no, lot_no, lot_qty, type, chief_id)
    data = {
    }
    return JsonResponse(data)

def fix_rm_mat_code(request):
    order_no = request.GET.get('order_no')
    rm_mat_code = request.GET.get('rm_mat_code')
    fixRMMaterialCode(order_no, rm_mat_code)
    data = {
    }
    return JsonResponse(data)

def set_wc_target(request):
    wc_no = request.GET.get('wc_no')
    target_hour = request.GET.get('target_hour')
    setWorkCenterTarget(wc_no, target_hour)
    data = {
    }
    return JsonResponse(data)

def set_wc_cap(request):
    wc_no = request.GET.get('wc_no')
    cap_hour = request.GET.get('cap_hour')
    setWorkCenterCapacity(wc_no, cap_hour)
    data = {
    }
    return JsonResponse(data)

def clear_pln_fai(request):
    order_no = request.GET.get('order_no')
    clearPLNFAI(order_no)
    data = {
    }
    return JsonResponse(data)

def remark_pln_fai(request):
    order_no = request.GET.get('order_no')
    remark = request.GET.get('remark')
    remarkPLNFAI(order_no, remark)
    data = {
    }
    return JsonResponse(data)

def save_abgraph_mn(request):
    date = request.GET.get('date')
    mc_no = request.GET.get('mc_no')
    hr = request.GET.get('hr')
    min = request.GET.get('min')
    states = request.GET.getlist('states[]')
    saveABGraphMunalData(mc_no, date, states, hr, min)
    data = {
    }
    return JsonResponse(data)

def get_wcs_of_profit_center(request):
    profit_center = request.GET.get('profit_center')
    workCenterList = getWorkCenterListInProfitCenter(profit_center)
    wcs = []
    for workCenter in workCenterList:
        wcs.append(workCenter.WorkCenterNo)
    data = {
        'wcs': wcs,
    }
    return JsonResponse(data)

def generate_order_analysis(request):
    token = request.GET.get('token')
    fg_matcode = request.GET.get('fg_matcode')
    start_date = request.GET.get('start_date') if request.GET.get('start_date') != '' else None
    stop_date = request.GET.get('stop_date') if request.GET.get('stop_date') != '' else None
    type_excel_data = request.GET.get('type_excel_data')
    include_other_operation = True if request.GET.get('include_other_operation') == 'true' else False
    is_consider_all = True if request.GET.get('is_consider_all') == 'true' else False
    consider_operations = request.GET.getlist('consider_operations[]')
    consider_operations_sql = get_list_for_sql(consider_operations)
    cursor = get_connection().cursor()
    sql = f"""
    SELECT
        FG_MaterialCode, TB1.OrderNo, TB1.OperationNo, WorkCenterNo, WorkCenterType, CAST(ProcessStart AS DATE) AS ProcessStart, CAST(ProcessStop AS DATE) AS ProcessStop, 
        OrderQty, ProcessQty, AcceptedQty, RejectedQty, TotalRejectedQty,
        CASE WHEN TB1.OperationNo <> LastOperationNo THEN NULL ELSE TotalRejectedQty END AS CummulateRejectQty,
        ROUND(COALESCE(((RejectedQty / NULLIF(OrderQty, 0)) * 100), 0), 2) AS RejectByOrderPercentage,
        ROUND(COALESCE(((RejectedQty / NULLIF(ProcessQty, 0)) * 100), 0), 2) AS RejectByProcessPercentage,
        CASE WHEN TB1.OperationNo <> LastOperationNo THEN NULL ELSE ROUND(COALESCE(((TotalRejectedQty / NULLIF(OrderQty, 0)) * 100), 0), 2) END AS CummulateRejectPercentage,
        EstSetupPerUnit, EstSetupTotal, 
        COALESCE(ActSetupTotal, 0) AS ActSetupTotal,
        ROUND(COALESCE(((EstSetupTotal / NULLIF(ActSetupTotal, 0)) * 100), 0), 2) AS SetupPercentage,
        EstOperPerUnit, EstOperTotal, MainProcess, TargetPerUnit, TargetTotal, COALESCE(ActOperTotal, 0) AS ActOperTotal, 
        ROUND(COALESCE(((EstOperTotal / NULLIF(ActOperTotal, 0)) * 100), 0), 2) AS OperByEstPercentage,
        ROUND(COALESCE(((TargetTotal / NULLIF(ActOperTotal, 0)) * 100), 0), 2) AS OperByTarPercentage,
        COALESCE(IdleTotal, 0) AS IdleTotal, COALESCE((IdleTotal + ActOperTotal), 0) AS ActIdleOperTotal,
        ROUND(COALESCE(((EstOperTotal / NULLIF((IdleTotal + ActOperTotal), 0)) * 100), 0), 2) AS IdleOperPercentage,
        EstLaborPerUnit, EstLaborTotal, COALESCE(ActLaborTotal, 0) AS ActLaborTotal,
        ROUND(COALESCE(((EstLaborTotal / NULLIF(ActLaborTotal, 0)) * 100), 0), 2) AS LaborPercentage
        FROM
        (
        SELECT OC.FG_MaterialCode, OPC.OrderNo, OPC.OperationNo, OPC.WorkCenterNo, WC.WorkCenterType, OPC.ProcessStart, OPC.ProcessStop, OC.ProductionOrderQuatity AS OrderQty, OPC.ProcessQty, OPC.AcceptedQty, OPC.RejectedQty,
        OPC.EstSetupTime AS EstSetupPerUnit, (OPC.EstSetupTime * OPC.ProcessQty) AS EstSetupTotal, OPC.EstOperationTime AS EstOperPerUnit, (OPC.EstOperationTime * OPC.ProcessQty) AS EstOperTotal,
        OPC.EstLaborTime AS EstLaborPerUnit, (OPC.EstLaborTime * OPC.ProcessQty) AS EstLaborTotal, COALESCE(WC.MainProcess, '') AS MainProcess, COALESCE(CTT.TargetValue, 0) AS TargetPerUnit, (COALESCE(CTT.TargetValue, 0) * OPC.ProcessQty) AS TargetTotal
        FROM OperationControl AS OPC
        INNER JOIN OrderControl AS OC ON OPC.OrderNo = OC.OrderNo
        INNER JOIN WorkCenter AS WC ON OPC.WorkCenterNo = WC.WorkCenterNo
        LEFT JOIN CycleTimeTarget AS CTT ON WC.MainProcess = CTT.MainProcess AND OC.FG_MaterialCode = CTT.FG_MaterialCode
        WHERE OC.FG_MaterialCode LIKE '{fg_matcode}%'"""
    if not include_other_operation and not is_consider_all:
        sql += f" AND OPC.WorkCenterNo IN {consider_operations_sql}"
    if start_date and stop_date:
        sql += f" AND OPC.OrderNo IN (SELECT OrderNo FROM OperationControl WHERE ProcessStop BETWEEN '{start_date}' AND '{stop_date}'"
        if not is_consider_all:
            sql += f" AND WorkCenterNo IN {consider_operations_sql}"
        sql += ")"
    elif start_date:
        sql += f" AND OPC.OrderNo IN (SELECT OrderNo FROM OperationControl WHERE ProcessStop >= '{start_date}'"
        if not is_consider_all:
            sql += f" AND WorkCenterNo IN {consider_operations_sql}"
        sql += ")"
    elif stop_date:
        sql += f" AND OPC.OrderNo IN (SELECT OrderNo FROM OperationControl WHERE ProcessStop <= '{stop_date}'"
        if not is_consider_all:
            sql += f" AND WorkCenterNo IN {consider_operations_sql}"
        sql += ")"
    sql += """
        ) AS TB1
        INNER JOIN (SELECT OrderNo, MAX(OperationNo) As LastOperationNo, SUM(RejectedQty) AS TotalRejectedQty FROM OperationControl GROUP BY OrderNo) AS TB3 ON TB1.OrderNo = TB3.OrderNo
        LEFT JOIN (
        SELECT OrderNo, OperationNo, COALESCE(SUM(Setup), 0) AS ActSetupTotal, COALESCE(SUM(Oper), 0) As ActOperTotal, COALESCE(SUM(Labor), 0) AS ActLaborTotal, COALESCE(SUM(Idle), 0) AS IdleTotal
        FROM HistoryOperate
        GROUP BY OrderNo, OperationNo
        ) AS TB2 ON TB1.OrderNo = TB2.OrderNo AND TB1.OperationNo = TB2.OperationNo
        ORDER BY FG_MaterialCode, TB1.OrderNo ASC, TB1.OperationNo ASC
    """
    print(sql)
    cursor.execute(sql)
    orders = cursor.fetchall()
    file_name = f'order_analysis_{token}.xlsx'
    file_path = f'media/order_analysis/{file_name}'
    try:
        dir_path = file_path
        shutil.rmtree(dir_path)
    except:
        error_msg = 'cant find path'
    source_file = settings.MEDIA_ROOT + f'/order_analysis_template.xlsx'
    destination_file = settings.MEDIA_ROOT + '/order_analysis/' + file_name
    shutil.copy(source_file, destination_file)
    wb = load_workbook(file_path)
    # Detail
    ws = wb['Detail']
    ws['B3'] = datetime.today()
    ws['B4'] = getClientIP(request)
    ws['B5'] = fg_matcode
    ws['B6'] = start_date if start_date else ''
    ws['B7'] = stop_date if stop_date else ''
    ws['B8'] = 'Yes' if include_other_operation else 'No'
    if is_consider_all:
        ws['B9'] = 'All'
    else:
        row = 9
        for consider_operation in consider_operations:
            ws['B' + str(row)] = consider_operation
            row = row + 1
    # Table
    ws = wb['Table']
    row = 4
    columns = [c for c in string.ascii_uppercase]
    columns += [a + b for a in string.ascii_uppercase for b in string.ascii_uppercase]
    for order in orders:
        values = [order.FG_MaterialCode, order.OrderNo, order.OperationNo, order.WorkCenterNo, order.WorkCenterType, order.ProcessStart,
          order.ProcessStop, order.OrderQty, order.ProcessQty, order.AcceptedQty, order.RejectedQty, order.CummulateRejectQty, 
          order.RejectByOrderPercentage, order.RejectByProcessPercentage, order.CummulateRejectPercentage, order.EstSetupPerUnit,
          order.EstSetupTotal, order.ActSetupTotal, order.SetupPercentage, order.EstOperPerUnit, order.EstOperTotal, order.MainProcess,
          order.TargetPerUnit, order.TargetTotal, order.ActOperTotal, order.OperByEstPercentage, order.OperByTarPercentage, order.IdleTotal,
          order.ActIdleOperTotal, order.IdleOperPercentage, order.EstLaborPerUnit, order.EstLaborTotal, order.ActLaborTotal, order.LaborPercentage]
        for col, value in zip(columns, values):
            ws[col + str(row)] = value
        row = row + 1

    if type_excel_data == 'Rej':
        columns_to_hide = ['E','F','G', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH']

        for column_letter in columns_to_hide:
            ws.column_dimensions[column_letter].hidden = True

    wb.save(file_path)
    data = {
        'new_token': generate_random_token(6),
        'file_path': '/' + file_path        
    }
    return JsonResponse(data)

#--------------------------------------------------------------------------- TOOL LIST

def upload_tool_list_save(request): 
    order_no = request.POST['order_no'].strip()
    operation_no = request.POST['operation_no'].strip()
    wc_no = request.POST['wc_no'].strip()
    file = request.FILES['file']
    sheets = request.POST.getlist('sheets')
    # upload file
    file_path = 'tool_list/' + file.name
    # remove file
    try:
        path = 'media/' + file_path
        os.remove(path)
    except Exception:
        print('Error : File Not Found')
    # save file
    fs = FileSystemStorage()
    fs.save(file_path, file)
    time.sleep(3)
    fg_matcode = getOrder(order_no).FG_MaterialCode
    fg_drawing = getOrder(order_no).FG_Drawing
    # run upload data to db
    if run_upload_tool_list(True, fg_matcode, fg_drawing, order_no, operation_no, wc_no, file, sheets):
        run_upload_tool_list(False, fg_matcode, fg_drawing, order_no, operation_no, wc_no, file, sheets)
    return redirect('/tool_store/')

def transfer_tool(request):
    order_no = request.GET['order_no'].strip()
    operation_no = request.GET['operation_no'].strip()
    toolh_id = request.GET.get('toolh_id')
    transferToolHeader(toolh_id, order_no, operation_no)
    data = {
    }
    return JsonResponse(data)

def return_tool_all(request):
    toolh_id = request.GET.get('toolh_id')
    toolh = getToolHeaderByID(toolh_id)
    toolItemList = getToolItemList(toolh_id)
    for tooli in toolItemList:
        if tooli.ConfirmedQty > 0:
            insertHistoryTool(tooli.ID, tooli.No, tooli.CTCode, tooli.ToolNo, tooli.ConfirmedQty, tooli.ToolLifeQty, tooli.ToolLifeMin, tooli.Remark, 'Return', toolh.FG_MaterialCode, toolh.OrderNo, toolh.OperationNo, toolh.WorkCenterNo)
    deleteTool(toolh_id)
    data = {
    }
    return JsonResponse(data)

def return_tool(request):
    tooli_id = request.GET.get('tooli_id')
    tooli = getToolItem(tooli_id)
    toolh = getToolHeaderByID(tooli.ToolHeaderID)
    if tooli.ConfirmedQty > 0:
            insertHistoryTool(tooli.ID, tooli.No, tooli.CTCode, tooli.ToolNo, tooli.ConfirmedQty, tooli.ToolLifeQty, tooli.ToolLifeMin, tooli.Remark, 'Return', toolh.FG_MaterialCode, toolh.OrderNo, toolh.OperationNo, toolh.WorkCenterNo)
    deleteToolItem(tooli_id)
    if len(getToolItemList(tooli.ToolHeaderID)) == 0:
        deleteTool(tooli.ToolHeaderID)
    data = {
    }
    return JsonResponse(data)

def change_tool_wc(request):
    toolh_id = request.GET.get('toolh_id').strip()
    wc_no = request.GET.get('wc_no').strip()
    changeToolHeaderWorkCenter(toolh_id, wc_no)
    data = {
    }
    return JsonResponse(data)

def change_tool(request):
    tooli_id = request.GET.get('tooli_id')
    change_type = request.GET.get('change_type')
    reason = request.GET.get('reason').strip()
    ct_code = request.GET.get('ct_code').strip()
    tool_no = request.GET.get('tool_no').strip()
    tool_life_qty = request.GET.get('tool_life_qty').strip()
    tool_life_min = request.GET.get('tool_life_min').strip()
    remark = request.GET.get('remark').strip()
    tooli = getToolItem(tooli_id)
    toolh = getToolHeaderByID(tooli.ToolHeaderID)
    if change_type == 'Same':
        InsertToolItem(tooli.ToolHeaderID, tooli.No, tooli.CTCode, tooli.ToolNo, tooli.ToolLifeQty, tooli.ToolLifeMin, tooli.Remark, 0)
    elif change_type == 'Diff':
        InsertToolItem(tooli.ToolHeaderID, tooli.No, ct_code, tool_no, tool_life_qty, tool_life_min, remark, 0)
    if tooli.ConfirmedQty > 0:
        insertHistoryTool(tooli.ID, tooli.No, tooli.CTCode, tooli.ToolNo, tooli.ConfirmedQty, tooli.ToolLifeQty, tooli.ToolLifeMin, tooli.Remark, reason, toolh.FG_MaterialCode, toolh.OrderNo, toolh.OperationNo, toolh.WorkCenterNo)
    deleteToolItem(tooli_id)
    data = {
    }
    return JsonResponse(data)

def runToolConfirm(order_no, operation_no, wc_no, accepted_qty, rejected_qty):
    confirmed_qty = int(accepted_qty) + int(rejected_qty)
    if confirmed_qty == 0:
        return
    if not isExistToolHeader(order_no, operation_no, wc_no):
        return
    toolh = getToolHeaderByOrder(order_no, operation_no, wc_no)
    toolItemList = getToolItemList(toolh.ID)
    for tooli in toolItemList:
        updateToolItemConfirm(tooli.ID, confirmed_qty)
    return

def add_new_tool(request):
    toolh_id = request.GET.get('toolh_id')
    no = request.GET.get('no').strip()
    ct_code = request.GET.get('ct_code').strip()
    tool_no = request.GET.get('tool_no').strip()
    tool_life_qty = request.GET.get('tool_life_qty').strip()
    tool_life_min = request.GET.get('tool_life_min').strip()
    remark = request.GET.get('remark').strip()
    InsertToolItem(toolh_id, no, ct_code, tool_no, tool_life_qty, tool_life_min, remark, 0)
    data = {
    }
    return JsonResponse(data)

################################################################################
################################### DATABASE ###################################
################################################################################

def get_connection():
    conn = pyodbc.connect('Driver={SQL Server};''Server=SVCCS-SFR\SQLEXPRESS01;''Database=SFR;''UID=sa;''PWD=$fr@2021;''Trusted_Connection=yes;')
    return conn

def get_connection_hr_focus():
    conn = pyodbc.connect('Driver={SQL Server};''Server=HR-FOCUS\SQLEXPRESS;''Database=HR_CCS;''UID=sa;''PWD=@MISccs55;''Trusted_Connection=yes;')
    return conn

#-------------------------------------------------------------------------- LIST

def getSAPOrderList(fdate, fhour):
    cursor = get_connection().cursor()
    sql = ""
    if fhour == "ALLDAY":
        sql = "SELECT * FROM [SAP_Order] WHERE DateGetFromSAP >= '" + fdate + " 00:00:00' AND DateGetFromSAP <= '" + fdate + " 23:59:59' ORDER BY DateGetFromSAP DESC"
    else:
        sql = "SELECT * FROM [SAP_Order] WHERE DateGetFromSAP >= '" + fdate + " " + fhour + ":00:00' AND DateGetFromSAP <= '" + fdate + " " + fhour + ":59:59' ORDER BY DateGetFromSAP DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getSAPRoutingList(fdate, fhour):
    cursor = get_connection().cursor()
    sql = ""
    if fhour == "ALLDAY":
        sql = "SELECT * FROM [SAP_Routing] WHERE DateGetFromSAP >= '" + fdate + " 00:00:00' AND DateGetFromSAP <= '" + fdate + " 23:59:59' ORDER BY DateGetFromSAP DESC"
    else:
        sql = "SELECT * FROM [SAP_Routing] WHERE DateGetFromSAP >= '" + fdate + " " + fhour + ":00:00' AND DateGetFromSAP <= '" + fdate + " " + fhour + ":59:59' ORDER BY DateGetFromSAP DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getSAPComponentList(fdate, fhour):
    cursor = get_connection().cursor()
    sql = ""
    if fhour == "ALLDAY":
        sql = "SELECT * FROM [SAP_Component] WHERE DateGetFromSAP >= '" + fdate + " 00:00:00' AND DateGetFromSAP <= '" + fdate + " 23:59:59' ORDER BY DateGetFromSAP DESC"
    else:
        sql = "SELECT * FROM [SAP_Component] WHERE DateGetFromSAP >= '" + fdate + " " + fhour + ":00:00' AND DateGetFromSAP <= '" + fdate + " " + fhour + ":59:59' ORDER BY DateGetFromSAP DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getSAPReportList(fdate, fhour):
    cursor = get_connection().cursor()
    sql = ""
    if fhour == "ALLDAY":
        sql = "SELECT * FROM [SFR2SAP_Report] WHERE DateTimeStamp >= '" + fdate + " 00:00:00' AND DateTimeStamp <= '" + fdate + " 23:59:59' ORDER BY DateTimeStamp DESC"
    else:
        sql = "SELECT * FROM [SFR2SAP_Report] WHERE DateTimeStamp >= '" + fdate + " " + fhour + ":00:00' AND DateTimeStamp <= '" + fdate + " " + fhour + ":59:59' ORDER BY DateTimeStamp DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getSAPModifierList(fdate, fhour):
    cursor = get_connection().cursor()
    if fhour == "ALLDAY":
        sql = "SELECT * FROM [SFR2SAP_Modifier] WHERE DateTimeStamp >= '" + fdate + " 00:00:00' AND DateTimeStamp <= '" + fdate + " 23:59:59' ORDER BY DateTimeStamp DESC"
    else:
        sql = "SELECT * FROM [SFR2SAP_Modifier] WHERE DateTimeStamp >= '" + fdate + " " + fhour + ":00:00' AND DateTimeStamp <= '" + fdate + " " + fhour + ":59:59' ORDER BY DateTimeStamp DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [WorkCenter]"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterRoutingList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [WorkCenter] WHERE IsRouting = 1"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterMachineList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [WorkCenter] WHERE WorkCenterType = 'Machine' AND IsRouting = 0"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterGroupList():
    cursor = get_connection().cursor()
    sql = "SELECT WorkCenterGroup FROM [WorkCenter] GROUP BY WorkCenterGroup"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterListInProfitCenter(profit_center):
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM [WorkCenter] WHERE ProfitCenter = '{profit_center}'"
    cursor.execute(sql)
    return cursor.fetchall()

def getOnRoutingList():
    cursor = get_connection().cursor()
    sql = "SELECT WC1.WorkCenterNo FROM WorkCenter AS WC1 INNER JOIN WorkCenter AS WC2 ON WC1.WorkCenterNo = WC2.OnRouting GROUP BY WC1.WorkCenterNo"
    cursor.execute(sql)
    return cursor.fetchall()

def getProfitCenterList():
    cursor = get_connection().cursor()
    sql = "SELECT ProfitCenter FROM WorkCenter WHERE IsActive = 1 GROUP BY ProfitCenter"
    cursor.execute(sql)
    return cursor.fetchall()

def getMachineWorkCenterGroupList():
    cursor = get_connection().cursor()
    sql = "SELECT WorkCenterGroup FROM [WorkCenter] WHERE WorkCenterType = 'Machine' GROUP BY WorkCenterGroup"
    cursor.execute(sql)
    return cursor.fetchall()

def getOperatorList():
    cursor = get_connection().cursor()
    sql = """
            SELECT EMP.EmpID, EmpName, Section, CostCenter, EmploymentType, Position, IsActive, ProfitCenter, JobFunction, MAX(StartDateTime) AS LastStartWorkingTime FROM Employee AS EMP
            LEFT JOIN HistoryOperate AS HO ON EMP.EmpID = HO.EmpID
            GROUP BY EMP.EmpId, EmpName, Section, CostCenter, EmploymentType, Position, IsActive, ProfitCenter, JobFunction
        """
    cursor.execute(sql)
    return cursor.fetchall()

def getBomList(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [SAP_Component] WHERE ProductionOrderNo = '" + order_no + "' ORDER BY DateGetFromSAP"
    cursor.execute(sql)
    return cursor.fetchall()

def getIdleTypeList():
    cursor = get_connection().cursor()
    sql = """
            SELECT * FROM IdleType
        """
    cursor.execute(sql)
    return cursor.fetchall()

def getRejectReasonList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [RejectReason] ORDER BY Name ASC"
    cursor.execute(sql)
    return cursor.fetchall()

def getMaterialGroupList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [MaterialGroup]"
    cursor.execute(sql)
    return cursor.fetchall()

def getPurchaseGroupList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [PurchaseGroup]"
    cursor.execute(sql)
    return cursor.fetchall()

def getCurrencyList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [Currency]"
    cursor.execute(sql)
    return cursor.fetchall()

def getOperationList(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperationControl] as OC"
    sql += " LEFT JOIN [WorkCenter] as WC ON OC.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE OrderNo = '" + order_no + "' ORDER BY OperationNo ASC"
    cursor.execute(sql)
    return cursor.fetchall()

def getModList(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [HistoryModifier] WHERE OrderNo = '" + order_no + "' ORDER BY ModifyDateTime DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getOverTimeOperatorList(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OvertimeOperator] WHERE OrderNo = '"+order_no+"' AND OperationNo = '"+operation_no+"' ORDER BY DateTimeStamp DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getOverTimeWorkCenterList(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OvertimeWorkCenter] WHERE OrderNo = '"+order_no+"' AND OperationNo = '"+operation_no+"' ORDER BY DateTimeStamp DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getUserList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [User]"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkingOrderList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OrderControl] WHERE ProcessStart IS NOT NULL AND ProcessStop IS NULL AND OrderNo NOT IN (SELECT OrderNo FROM CanceledOrder)"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkingWorkCenterList():
    cursor = get_connection().cursor()
    sql = "SELECT OWC.WorkCenterNo AS WCN, OC.Note AS OperationNote, * FROM [OperatingWorkCenter] as OWC INNER JOIN [WorkCenter] as WC ON OWC.WorkCenterNo = WC.WorkCenterNo"
    sql += " INNER JOIN [OperationControl] as OC ON OC.OrderNo = OWC.OrderNo AND OC.OperationNo = OWC.OperationNo"
    sql += " INNER JOIN [OrderControl] as ORDC ON OWC.OrderNo = ORDC.OrderNo"
    sql += " LEFT JOIN IdleType AS IT ON IT.Code = OWC.IdleCode"
    sql += " WHERE Status <> 'COMPLETE' ORDER BY OWC.OperatingWorkCenterID ASC"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkingOperatorList():
    cursor = get_connection().cursor()
    sql = "SELECT OOPR.EmpID, EMP.EmpName, EMP.Section, EMP.CostCenter, OOPR.Status, OOPR.OrderNo, OOPR.OperationNo, OOPR.StartDateTime, OWC.WorkCenterNo, OC.Note, ORDC.FG_MaterialCode, ORDC.FG_Drawing "
    sql += " FROM [OperatingOperator] as OOPR INNER JOIN [Employee] as EMP ON OOPR.EmpID = EMP.EmpID"
    sql += " INNER JOIN [OperationControl] as OC ON OC.OrderNo = OOPR.OrderNo AND OC.OperationNo = OOPR.OperationNo"
    sql += " INNER JOIN [OrderControl] as ORDC ON OOPR.OrderNo = ORDC.OrderNo"
    sql += " LEFT JOIN [OperatingWorkCenter] as OWC ON OOPR.OperatingWorkCenterID = OWC.OperatingWorkCenterID"
    sql += " LEFT JOIN [WorkCenter] as WC ON OWC.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE OOPR.Status <> 'COMPLETE' ORDER BY OOPR.OperatingOperatorID ASC"
    cursor.execute(sql)
    return cursor.fetchall()

def getNoneWorkingWorkCenterList():
    cursor = get_connection().cursor()
    sql = """
            SELECT WC.*, TB.StopDateTime FROM WorkCenter AS WC LEFT JOIN
            (SELECT WorkCenterNo, MAX(StopDateTime) AS StopDateTime FROM HistoryOperate GROUP BY WorkCenterNo) AS TB ON WC.WorkCenterNo = TB.WorkCenterNo
            WHERE WC.WorkCenterNo NOT IN (SELECT WorkCenterNo FROM OperatingWorkCenter WHERE Status <> 'COMPLETE' AND Status <> 'WAITING')
            AND WC.IsRouting = 0 AND WC.IsActive = 1 AND WC.WorkCenterType = 'Machine'
            """
    cursor.execute(sql)
    return cursor.fetchall()

def getOperatingWorkCenterList(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperatingWorkCenter] as OWC INNER JOIN [WorkCenter] as WC ON OWC.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "' ORDER BY OWC.OperatingWorkCenterID ASC"
    cursor.execute(sql)
    return cursor.fetchall()

def getOperatingWorkCenterListForTable(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT OWC.OperatingWorkCenterID, OWC.WorkCenterNo, WC.WorkCenterName, OWC.StartDateTime, OWC.StopDateTime, OWC.Status, WC.MachineType, IT.Description, *"
    sql += " FROM [OperatingWorkCenter] as OWC INNER JOIN [WorkCenter] as WC ON OWC.WorkCenterNo = WC.WorkCenterNo"
    sql += " LEFT JOIN IdleType AS IT ON IT.Code = OWC.IdleCode"
    sql += " WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "' ORDER BY OWC.OperatingWorkCenterID DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getOperatingOperatorList(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT OOPR.OperatingOperatorID, OOPR.OrderNo, OOPR.OperationNo, OOPR.EmpID, OOPR.StartDateTime, OOPR.StopDateTime, OOPR.Status, OOPR.OperatingWorkCenterID, OWC.WorkCenterNo"
    sql += " FROM [OperatingOperator] as OOPR INNER JOIN [Employee] as EMP ON OOPR.EmpID = EMP.EmpID"
    sql += " LEFT JOIN [OperatingWorkCenter] as OWC ON OOPR.OperatingWorkCenterID = OWC.OperatingWorkCenterID"
    sql += " LEFT JOIN [WorkCenter] as WC ON OWC.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE OOPR.OrderNo = '" + order_no + "' AND OOPR.OperationNo = '" + operation_no + "' ORDER BY OOPR.OperatingOperatorID ASC"
    cursor.execute(sql)
    return cursor.fetchall()

def getOperatingOperatorListForTable(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT OOPR.OperatingOperatorID, WC.WorkCenterNo, WC.WorkCenterName, OOPR.EmpID, EMP.EmpName, OOPR.StartDateTime, OOPR.StopDateTime, OOPR.Status,*"
    sql += " FROM [OperatingOperator] as OOPR INNER JOIN [Employee] as EMP ON OOPR.EmpID = EMP.EmpID"
    sql += " LEFT JOIN [OperatingWorkCenter] as OWC ON OOPR.OperatingWorkCenterID = OWC.OperatingWorkCenterID"
    sql += " LEFT JOIN [WorkCenter] as WC ON OWC.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE OOPR.OrderNo = '" + order_no + "' AND OOPR.OperationNo = '" + operation_no + "' ORDER BY OOPR.OperatingOperatorID DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getJoinList(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperationControl] WHERE JoinToOrderNo = '" + order_no + "' AND JoinToOperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    return cursor.fetchall()

def getJoinableList(order_no, work_center_group):
    cursor = get_connection().cursor()
    sql = "SELECT OC.OrderNo as OCOrderNo, OC.OperationNo as OCOperationNo, OC.WorkCenterNo as OCWorkCenterNo, *"
    sql += " FROM [OperationControl] as OC INNER JOIN [WorkCenter] as WC ON OC.WorkCenterNo = WC.WorkCenterNo"
    sql += " LEFT JOIN [OperatingOperator] as OOPR ON OC.OrderNo = OOPR.OrderNo AND OC.OperationNo = OOPR.OperationNo"
    sql += " WHERE "
    #1
    sql += " ProcessQty - (AcceptedQty + RejectedQty) <> 0"
    #2
    sql += " AND (EmpID IS NULL OR STATUS = 'COMPLETE')"
    #3
    sql += " AND WorkCenterType = 'Machine'"
    #4
    sql += " AND WC.WorkCenterGroup = '" + work_center_group + "'"
    #5
    sql += " AND OC.OrderNo <> '" + order_no + "'"
    #6
    sql += " AND JoinToOrderNo IS NULL AND JoinToOperationNo IS NULL"
    cursor.execute(sql)
    return cursor.fetchall()

def getHistoryOperateList(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [HistoryOperate] AS HO"
    sql += " INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " LEFT JOIN IdleType AS IT ON IT.Code = HO.IdleCode"
    sql += " LEFT JOIN SFR2SAP_Report AS SAP"
    sql += " ON HO.OrderNo = SAP.ProductionOrderNo AND HO.OperationNo = SAP.OperationNumber AND HO.Type = 'MANUAL'"
    sql += " AND Ho.Setup = SAP.SetupTime AND HO.Oper = SAP.OperTime AND HO.Labor = SAP.LaborTime AND HO.EmpID = SAP.EmployeeID"
    sql += " AND SAP.StartDate = CONVERT(NVARCHAR(MAX), HO.StartDateTime, 112) AND SAP.StartTime = REPLACE(CONVERT(NVARCHAR(MAX), HO.StartDateTime, 8), ':', '')"
    sql += " AND SAP.FinishDate = CONVERT(NVARCHAR(MAX), HO.StopDateTime, 112) AND SAP.FinishTime = REPLACE(CONVERT(NVARCHAR(MAX), HO.StopDateTime, 8), ':', '')"
    sql += " WHERE OrderNo = '"+order_no+"' AND OperationNo = '"+operation_no+"' ORDER BY StopDateTime DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getHistoryConfirmList(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT HC.*, OC.OperationNo AS OPN, OC.WorkCenterNo AS WCN"
    sql += " FROM HistoryConfirm AS HC LEFT JOIN OperationControl AS OC ON HC.OrderNo = OC.OrderNo AND HC.ScrapAt = OC.OperationNo"
    sql += " WHERE HC.OrderNo = '"+order_no+"' AND HC.OperationNo = '"+operation_no+"'"
    sql += " ORDER BY ConfirmDateTime DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getHistoryJoinList(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [HistoryJoin] "
    sql += " WHERE (JoinToOrderNo = '" + order_no + "' AND JoinToOperationNo = '" + operation_no + "')"
    sql += " OR (JoinByOrderNo = '" + order_no + "' AND JoinByOperationNo = '" + operation_no + "')"
    cursor.execute(sql)
    return cursor.fetchall()

def getHistoryToolList(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM HistoryTool WHERE OrderNo = '{order_no}' AND OperationNo = '{operation_no}' ORDER BY No ASC"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterHistoryTransactionList(work_center_no, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = f"""
            SELECT HO.*, IT.Description FROM [HistoryOperate] AS HO
            LEFT JOIN IdleType AS IT ON HO.IdleCode = IT.Code
            WHERE WorkCenterNo = '{work_center_no}' AND month(StartDateTime) = '{month}' AND year(StartDateTime) = '{year}'
        """
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterHistoryOvertimeList(work_center_no, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OvertimeWorkCenter] WHERE WorkCenterNo = '"+work_center_no+"' AND month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getOvertimeWorkCenterList(fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OvertimeWorkCenter] WHERE month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getEmployeeHistoryTransactionList(emp_id, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [HistoryOperate] AS HO"
    sql += " INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE EmpID = '"+emp_id+"'AND month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getEmployeeHistoryConfirmList(emp_id, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [HistoryConfirm] WHERE EmpID = '"+emp_id+"' AND month(ConfirmDateTime) = '"+month+"' AND year(ConfirmDateTime) = '"+year+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getEmployeeHistoryOvertimeList(emp_id, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OvertimeOperator] WHERE EmpID = '"+emp_id+"' AND month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getOvertimeOperatorList(fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OvertimeOperator] AS OO LEFT JOIN Employee AS EMP ON OO.EmpID = EMP.EmpID"
    sql += " WHERE month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getOvertimeCounterList(fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT O.EmpID, E.EmpName, E.Section, E.CostCenter, COUNT(*) AS Counter From OvertimeOperator AS O INNER JOIN Employee AS E ON O.EmpID = E.EmpID"
    sql += " WHERE MONTH(StartDateTime) = '"+month+"' AND YEAR(StartDateTime) = '"+year+"' GROUP BY O.EmpID, E.EmpName, E.Section, E.CostCenter"
    cursor.execute(sql)
    return cursor.fetchall()

def getOperationNoTimeList(fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = f"""
            SELECT * 
            FROM OperationControl AS OC
            INNER JOIN WorkCenter AS WC ON OC.WorkCenterNo = WC.WorkCenterNo
            WHERE OrderNo NOT IN (SELECT OrderNo FROM CanceledOrder) /* Cancel Order */
            AND CONCAT(OrderNo, OperationNo) NOT IN (SELECT CONCAT(JoinToOrderNo, JoinToOperationNo) FROM HistoryJoin) /* Was Join To */
            AND CONCAT(OrderNo, OperationNo) NOT IN (SELECT CONCAT(JoinByOrderNo, JoinByOperationNo) FROM HistoryJoin) /* Was Join By */
            AND CONCAT(OrderNo, OperationNo) NOT IN (SELECT CONCAT(OrderNo, OperationNo) FROM HistoryOperate WHERE Oper + Labor != 0) /* Has Time Report */
            AND ProcessStop IS NOT NULL /* Only Stop Process */
            AND IsExternalProcess = 0 /* Exclude External Process */
            AND OC.WorkCenterNo NOT IN ('MT_AS', 'MT', 'PK', 'PK_AS', 'MT_CNF', 'MT-MD', 'JP_AS', 'HT', 'P-PE') /* Exclude Some Workcenter */
            AND Note NOT LIKE '%Combine%' AND Note <> 'PP' /* Exclude Some Note */
            AND month(ProcessStop) = '{month}' AND year(ProcessStop) = '{year}'
            ORDER BY ProcessStop DESC
          """
    cursor.execute(sql)
    return cursor.fetchall()

def getEmpWorkTimeList(fstartdate, fstopdate, emp_type):
    fstopdate = datetime.strptime(fstopdate, '%Y-%m-%d') + timedelta(days=1)
    fstopdate = fstopdate.strftime('%Y-%m-%d')
    cursor = get_connection().cursor()
    emp_type_sql = ''
    if emp_type != 'All':
        emp_type_sql = f"AND EmploymentType = '{emp_type}' "
    sql = f"""
            SELECT Section, CostCenter, EMP.EmpID, EmpName, EmploymentType, Position, ProfitCenter, JobFunction, ROUND(SUM(Setup), 0) AS Setup, ROUND(SUM(Labor), 0) AS Labor, ROUND(SUM(Setup), 0) + ROUND(SUM(Labor), 0) AS Total
            FROM HistoryOperate AS HO
            INNER JOIN Employee AS EMP ON HO.EmpID = EMP.EmpID
            WHERE (Labor <> 0 OR Setup <> 0) 
            AND DATEADD(s, DATEDIFF(s, StartDateTime, StopDateTime)/2, StartDateTime) >= '{fstartdate} 07:00:000' 
            AND DATEADD(s, DATEDIFF(s, StartDateTime, StopDateTime)/2, StartDateTime) <= '{fstopdate} 07:00:000'
            {emp_type_sql}
            GROUP BY Section, CostCenter, EMP.EmpID, EmpName, EmploymentType, Position, ProfitCenter, JobFunction
          """
    # sql = f"""
    #         SELECT EMP.*, TB1.Setup, TB1.Labor, TB1.Total 
    #         FROM Employee AS EMP
    #         LEFT JOIN
    #         (
    #         SELECT EmpID, ROUND(SUM(Setup), 0) AS Setup, ROUND(SUM(Labor), 0) AS Labor, ROUND(SUM(Setup), 0) + ROUND(SUM(Labor), 0) AS Total 
    #         FROM HistoryOperate
    #         WHERE (Labor <> 0 OR Setup <> 0) 
    #         AND DATEADD(s, DATEDIFF(s, StartDateTime, StopDateTime)/2, StartDateTime) >= '{fstartdate} 07:00:000'
    #         AND DATEADD(s, DATEDIFF(s, StartDateTime, StopDateTime)/2, StartDateTime) <= '{fstopdate} 07:00:000'
    #         {emp_type_sql}
    #         GROUP BY EmpID
    #         ) AS TB1 ON EMP.EmpID = TB1.EmpID
    #         WHERE IsActive = 1
    #     """
    cursor.execute(sql)
    return cursor.fetchall()

def getPTLList(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [PartialLotTraveller] WHERE OrderNo = '" + order_no + "' ORDER BY LotNo DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getSAPOrderNoRoutingList():
    cursor = get_connection().cursor()
    sql = "SELECT OD.ProductionOrderNo, OD.DateGetFromSAP"
    sql += " From SAP_Order AS OD LEFT JOIN SAP_Routing AS RT ON OD.ProductionOrderNo = RT.ProductionOrderNo WHERE RT.ProductionOrderNo IS NULL"
    cursor.execute(sql)
    return cursor.fetchall()

def getSAPDuplicateRoutingList():
    cursor = get_connection().cursor()
    sql = """
            SELECT RT1.ProductionOrderNo, RT1.OperationNumber, RT1.DateGetFromSAP AS RT1DateGetFromSAP, RT2.DateGetFromSAP AS RT2DateGetFromSAP
            FROM SAP_Routing AS RT1 INNER JOIN SAP_Routing AS RT2 ON RT1.ProductionOrderNo = RT2.ProductionOrderNo AND RT1.OperationNumber = RT2.OperationNumber
            WHERE RT1.DateGetFromSAP < RT2.DateGetFromSAP
        """
    cursor.execute(sql)
    return cursor.fetchall()

def getNoneStartOrderList():
    cursor = get_connection().cursor()
    sql = """
            SELECT ProductionOrderNo, OrderNo, SO.FG_MaterialCode, SO.RequestDate, SO.FG_Drawing, SO.DateGetFromSAP FROM SAP_Order AS SO
            LEFT JOIN OrderControl AS OC ON SO.ProductionOrderNo = OC.OrderNo
            WHERE (OC.OrderNo IS NULL OR (OC.ProcessStart IS NULL)) AND ProductionOrderNo NOT IN (SELECT OrderNo From CanceledOrder)
          """
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterErrorList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM SAP_Routing AS RT LEFT JOIN WorkCenter AS WC ON RT.WorkCenter = WC.WorkCenterNo WHERE WC.WorkCenterNo IS NULL"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterInGroupList(work_center_group):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM WorkCenter WHERE IsRouting = 0 AND WorkCenterGroup = '"+work_center_group+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterInGroupActiveList(work_center_group, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    fdate = year + '-' + month + '-15'
    cursor = get_connection().cursor()
    sql = "SELECT * FROM WorkCenter WHERE IsRouting = 0 AND WorkCenterGroup = '"+work_center_group+"' AND IsActive = 1"
    sql += " AND '"+str(fdate)+"' BETWEEN ActiveDate AND InActiveDate"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterOnRoutingActiveList(on_rt, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    fdate = year + '-' + month + '-15'
    cursor = get_connection().cursor()
    sql = "SELECT * FROM WorkCenter WHERE IsRouting = 0 AND OnRouting = '"+on_rt+"' AND IsActive = 1"
    sql += " AND '"+str(fdate)+"' BETWEEN ActiveDate AND InActiveDate"
    cursor.execute(sql)
    return cursor.fetchall()

def getMachineOnRoutingIsActiveList(frt):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM WorkCenter WHERE IsRouting = 0 AND OnRouting = '"+str(frt)+"' AND IsActive = 1"
    cursor.execute(sql)
    return cursor.fetchall()

def getAutoMachineManualReportOvertimeList(fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT DateTimeStamp, HO.OrderNo, HO.OperationNo, EmpID, HO.WorkCenterNo, Setup, Oper, Labor"
    sql += " FROM HistoryOperate AS HO "
    sql += " INNER JOIN OvertimeWorkCenter AS OTWC ON HO.OrderNo = OTWC.OrderNo AND HO.OperationNo = OTWC.OperationNo "
    sql += " INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo "
    sql += " WHERE Type = 'MANUAL' AND WC.MachineType = 'Auto' AND (Setup != 0 OR Oper != 0 OR Labor != 0) "
    sql += " AND month(DateTimeStamp) = '"+month+"' AND year(DateTimeStamp) = '"+year+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getEmpWorkRecordsList(ftype, fdate, fmonth, fstartdate, fstopdate):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT HO.EmpID, EMP.EmpName, EMP.Section, EMP.CostCenter, SUM(Setup) AS Setup, SUM(Labor) AS Labor"
    sql += " FROM HistoryOperate AS HO INNER JOIN Employee AS EMP ON HO.EmpID = EMP.EmpID"
    sql += " WHERE HO.Setup + HO.Labor > 0 AND"
    if ftype == "DAILY":
        sql += " StartDateTime >= '" + fdate + " 00:00:00' AND StartDateTime <= '" + fdate + " 23:59:59'"
    if ftype == "MONTHLY":
        sql += " month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    if ftype == "RANGE":
        sql += " StartDateTime >= '" + fstartdate + " 00:00:00' AND StartDateTime <= '" + fstopdate + " 23:59:59'"
    sql += " GROUP BY HO.EmpID, EMP.EmpName, EMP.Section, EMP.CostCenter"
    cursor.execute(sql)
    return cursor.fetchall()

def getWorkCenterWorkRecordsList(ftype, fdate, fmonth, fstartdate, fstopdate):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT WorkCenterGroup, WC.WorkCenterNo, WorkCenterName, SUM(Setup) AS Setup, SUM(Oper) AS Oper"
    sql += " FROM HistoryOperate AS HO INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE HO.Setup + HO.Oper > 0 AND"
    if ftype == "DAILY":
        sql += " StartDateTime >= '" + fdate + " 00:00:00' AND StartDateTime <= '" + fdate + " 23:59:59'"
    if ftype == "MONTHLY":
        sql += " month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    if ftype == "RANGE":
        sql += " StartDateTime >= '" + fstartdate + " 00:00:00' AND StartDateTime <= '" + fstopdate + " 23:59:59'"
    sql += " GROUP BY WC.WorkCenterGroup, WC.WorkCenterNo, WC.WorkCenterName"
    cursor.execute(sql)
    return cursor.fetchall()

def getCompletedOrderList(ftype, fdate, fmonth, fstartdate, fstopdate):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT *, DATEDIFF(DAY, CONVERT(DATE, ProcessStart), CONVERT(DATE, ProcessStop)) AS 'Day' FROM OrderControl"
    sql += " WHERE ProcessStop IS NOT NULL"
    if ftype == "DAILY":
        sql += " AND ProcessStop >= '" + fdate + " 00:00:00' AND ProcessStop <= '" + fdate + " 23:59:59'"
    if ftype == "MONTHLY":
        sql += " AND month(ProcessStop) = '"+month+"' AND year(ProcessStop) = '"+year+"'"
    if ftype == "RANGE":
        sql += " AND ProcessStop >= '" + fstartdate + " 00:00:00' AND ProcessStop <= '" + fstopdate + " 23:59:59'"
    cursor.execute(sql)
    return cursor.fetchall()

def getRejectedOrderList(ftype, fdate, fmonth, fstartdate, fstopdate):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT OC.OrderNo, OC.FG_MaterialCode, OC.FG_Drawing, OC.Note, OC.LotNo, OC.ProcessStart, OC.ProcessStop, DATEDIFF(DAY, CONVERT(DATE, OC.ProcessStart), CONVERT(DATE, OC.ProcessStop)) AS 'Day'"
    sql += " FROM OperationControl AS OPC INNER JOIN OrderControl AS OC ON OPC.OrderNo = OC.OrderNo"
    sql += " WHERE (OPC.ProcessStart IS NULL OR OPC.ProcessStop IS NULL ) AND OC.ProcessStop IS NOT NULL"
    if ftype == "DAILY":
        sql += " AND OC.ProcessStop >= '" + fdate + " 00:00:00' AND OC.ProcessStop <= '" + fdate + " 23:59:59'"
    if ftype == "MONTHLY":
        sql += " AND month(OC.ProcessStop) = '"+month+"' AND year(OC.ProcessStop) = '"+year+"'"
    if ftype == "RANGE":
        sql += " AND OC.ProcessStop >= '" + fstartdate + " 00:00:00' AND OC.ProcessStop <= '" + fstopdate + " 23:59:59'"
    sql += " GROUP BY OC.OrderNo, OC.FG_MaterialCode, OC.FG_Drawing, OC.Note, OC.LotNo, OC.ProcessStart, OC.ProcessStop"
    cursor.execute(sql)
    return cursor.fetchall()

def getCanceledOrderList(ftype, fdate, fmonth, fstartdate, fstopdate):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT * FROM CanceledOrder AS CO INNER JOIN OrderControl AS OC ON CO.OrderNo = OC.OrderNo"
    sql += " WHERE"
    if ftype == "DAILY":
        sql += " CO.DateTimeStamp >= '" + fdate + " 00:00:00' AND CO.DateTimeStamp <= '" + fdate + " 23:59:59'"
    if ftype == "MONTHLY":
        sql += " month(CO.DateTimeStamp) = '"+month+"' AND year(CO.DateTimeStamp) = '"+year+"'"
    if ftype == "RANGE":
        sql += " CO.DateTimeStamp >= '" + fstartdate + " 00:00:00' AND CO.DateTimeStamp <= '" + fstopdate + " 23:59:59'"
    cursor.execute(sql)
    return cursor.fetchall()

def getSAPDelayOperationList(fwc):
    cursor = get_connection().cursor()
    sql = """
            SELECT RT.ProductionOrderNo, RT.OperationNumber, RT.WorkCenter, SO.FG_MaterialCode, SO.ProductionOrderQuatity, SO.SalesOrderNo, SO.DrawingNo, RequestDate,
			CASE RequestDate WHEN '00.00.0000' THEN 9999 ELSE DATEDIFF(DAY, CONVERT(DATE, CONVERT(DATETIME, RequestDate, 104)), GETDATE()) END AS DelayFromRequestDate,
            DATEDIFF(DAY, CONVERT(DATE, SO.ReleaseDate, 104), GETDATE()) AS Actual_Work, SO.FG_Drawing
            FROM SAP_Routing AS RT INNER JOIN (SELECT ProductionOrderNo, MIN(OperationNumber) AS OperationNumber FROM SAP_Routing
            WHERE ProductionOrderNo IN (SELECT ProductionOrderNo FROM SAP_Order AS SO LEFT JOIN OrderControl AS OC ON SO.ProductionOrderNo = OC.OrderNo WHERE OC.OrderNo IS NULL)
            GROUP BY ProductionOrderNo) AS TB ON RT.ProductionOrderNo = TB.ProductionOrderNo AND RT.OperationNumber = TB.OperationNumber
            INNER JOIN SAP_Order AS SO ON SO.ProductionOrderNo = RT.ProductionOrderNo
        """
    if fwc != 'ALL':
        sql += f" WHERE RT.WorkCenter = '{fwc}'"
    cursor.execute(sql)
    return cursor.fetchall()

def getSFRDelayOperationList(fwc):
    cursor = get_connection().cursor()
    sql = """
            SELECT OPC.OrderNo, OPC.OperationNo, OPC.WorkCenterNo, (ProcessQty - (AcceptedQty + RejectedQty)) AS RemainingQty, OC.Note AS OrderNote, OPC.Note AS OperationNote,
            CASE RequestDate WHEN NULL THEN 9999 ELSE DATEDIFF(DAY, CONVERT(DATE, CONVERT(DATETIME, RequestDate)), GETDATE()) END AS DelayFromRequestDate,
            DATEDIFF(DAY, CONVERT(DATE, OPC.ProcessStart), GETDATE()) AS Actual_Work, OPC.ProcessStart, FG_MaterialCode, FG_Drawing, SalesOrderNo, DrawingNo, ProcessQty, RequestDate, OPC.DateGetFromSAP, OC.DateGetFromSAP AS Order_DGFS
            FROM OperationControl AS OPC INNER JOIN OrderControl AS OC ON OPC.OrderNo = OC.OrderNo
            WHERE OPC.OrderNo NOT IN (SELECT OrderNo FROM CanceledOrder) AND (ProcessQty - (AcceptedQty + RejectedQty) > 0)
          """
    if fwc != 'ALL':
        sql += f" AND OPC.WorkCenterNo = '{fwc}'"
    cursor.execute(sql)
    return cursor.fetchall()

def getMonthlyWorkCenterOperForABGraph(fwctype, fkey, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    fdate = year + '-' + month + '-15'
    cursor = get_connection().cursor()
    sql = "SELECT day(CONVERT(DATE, Fdate)) AS Fday, CAST(ROUND(SUM(Foper)/60, 0) AS Int) AS Foper FROM"
    sql += " ("
    # DIFF MONTH (CURRENT MONTH = STOP)
    sql += " (SELECT StopDateTime As Fdate, (Oper - ((DATEDIFF(MINUTE, StartDateTime, CONVERT(DATE, StopDateTime))/Oper) * Oper)) AS Foper"
    sql += " FROM HistoryOperate AS HO INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE Oper != 0 AND month(StartDateTime) != month(StopDateTime)"
    sql += " AND month(StopDateTime) = '"+month+"' AND year(StopDateTime) = '"+year+"'"
    sql += " AND '"+str(fdate)+"' BETWEEN ActiveDate AND InActiveDate"
    if fwctype == "WC":
        sql += " AND HO.WorkCenterNo = '"+fkey+"'"
    elif fwctype == "WCG":
        sql += " AND WorkCenterGroup = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1 AND IsActive = 1"
    elif fwctype == "RT":
        sql += " AND OnRouting = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1 AND IsActive = 1"
    sql += ") UNION"
    # DIFF DAY (CURRENT MONTH = START)
    sql += " (SELECT StartDateTime As Fdate, ((DATEDIFF(MINUTE, StartDateTime, CONVERT(DATE, StopDateTime))/Oper) * Oper) AS Foper"
    sql += " FROM HistoryOperate AS HO INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE Oper != 0 AND CONVERT(DATE, StartDateTime) != CONVERT(DATE, StopDateTime)"
    sql += " AND month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    sql += " AND '"+str(fdate)+"' BETWEEN ActiveDate AND InActiveDate"
    if fwctype == "WC":
        sql += " AND HO.WorkCenterNo = '"+fkey+"'"
    elif fwctype == "WCG":
        sql += " AND WorkCenterGroup = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1 AND IsActive = 1"
    elif fwctype == "RT":
        sql += " AND OnRouting = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1 AND IsActive = 1"
    sql += ") UNION"
    # SAME MONTH , DIFF DAY (FDAY = STOP)
    sql += " (SELECT StopDateTime As Fdate, (Oper - ((DATEDIFF(MINUTE, StartDateTime, CONVERT(DATE, StopDateTime))/Oper) * Oper)) AS Foper"
    sql += " FROM HistoryOperate AS HO INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE Oper != 0 AND CONVERT(DATE, StartDateTime) != CONVERT(DATE, StopDateTime) AND month(StartDateTime) = month(StopDateTime)"
    sql += " AND month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    sql += " AND '"+str(fdate)+"' BETWEEN ActiveDate AND InActiveDate"
    if fwctype == "WC":
        sql += " AND HO.WorkCenterNo = '"+fkey+"'"
    elif fwctype == "WCG":
        sql += " AND WorkCenterGroup = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    elif fwctype == "RT":
        sql += " AND OnRouting = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    sql += ") UNION"
    # SAME MONTH , SAME DAY
    sql += " (SELECT StartDateTime As Fdate, Oper AS Foper"
    sql += " FROM HistoryOperate AS HO INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE Oper != 0 AND CONVERT(DATE, StartDateTime) = CONVERT(DATE, StopDateTime)"
    sql += " AND month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    sql += " AND '"+str(fdate)+"' BETWEEN ActiveDate AND InActiveDate"
    if fwctype == "WC":
        sql += " AND HO.WorkCenterNo = '"+fkey+"'"
    elif fwctype == "WCG":
        sql += " AND WorkCenterGroup = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    elif fwctype == "RT":
        sql += " AND OnRouting = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    sql += ")) AS TB"
    sql += " GROUP BY day(CONVERT(DATE, Fdate))"
    cursor.execute(sql)
    return cursor.fetchall()

def getMonthlyWorkCenterSetupForABGraph(fwctype, fkey, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    fdate = year + '-' + month + '-15'
    cursor = get_connection().cursor()
    sql = "SELECT day(CONVERT(DATE, Fdate)) AS Fday, CAST(ROUND(SUM(Fsetup)/60, 0) AS Int) AS Fsetup FROM"
    sql += " ((SELECT StopDateTime As Fdate, (Setup - ((DATEDIFF(MINUTE, StartDateTime, CONVERT(DATE, StopDateTime))/Setup) * Setup)) AS Fsetup"
    sql += " FROM HistoryOperate AS HO INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE Setup != 0 AND month(StartDateTime) != month(StopDateTime)"
    sql += " AND month(StopDateTime) = '"+month+"' AND year(StopDateTime) = '"+year+"'"
    sql += " AND '"+str(fdate)+"' BETWEEN ActiveDate AND InActiveDate"
    if fwctype == "WC":
        sql += " AND HO.WorkCenterNo = '"+fkey+"'"
    elif fwctype == "WCG":
        sql += " AND WorkCenterGroup = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    elif fwctype == "RT":
        sql += " AND OnRouting = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    sql += ") UNION"
    sql += " (SELECT StartDateTime As Fdate, ((DATEDIFF(MINUTE, StartDateTime, CONVERT(DATE, StopDateTime))/Setup) * Setup) AS Fsetup"
    sql += " FROM HistoryOperate AS HO INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE Setup != 0 AND CONVERT(DATE, StartDateTime) != CONVERT(DATE, StopDateTime)"
    sql += " AND month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    sql += " AND '"+str(fdate)+"' BETWEEN ActiveDate AND InActiveDate"
    if fwctype == "WC":
        sql += " AND HO.WorkCenterNo = '"+fkey+"'"
    elif fwctype == "WCG":
        sql += " AND WorkCenterGroup = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    elif fwctype == "RT":
        sql += " AND OnRouting = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    sql += ") UNION"
    sql += " (SELECT StopDateTime As Fdate, (Setup - ((DATEDIFF(MINUTE, StartDateTime, CONVERT(DATE, StopDateTime))/Setup) * Setup)) AS Fsetup"
    sql += " FROM HistoryOperate AS HO INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE Setup != 0 AND CONVERT(DATE, StartDateTime) != CONVERT(DATE, StopDateTime) AND month(StartDateTime) = month(StopDateTime)"
    sql += " AND month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    sql += " AND '"+str(fdate)+"' BETWEEN ActiveDate AND InActiveDate"
    if fwctype == "WC":
        sql += " AND HO.WorkCenterNo = '"+fkey+"'"
    elif fwctype == "WCG":
        sql += " AND WorkCenterGroup = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    elif fwctype == "RT":
        sql += " AND OnRouting = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    sql += ") UNION"
    sql += " (SELECT StartDateTime As Fdate, Setup AS Fsetup"
    sql += " FROM HistoryOperate AS HO INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE Setup != 0 AND CONVERT(DATE, StartDateTime) = CONVERT(DATE, StopDateTime)"
    sql += " AND month(StartDateTime) = '"+month+"' AND year(StartDateTime) = '"+year+"'"
    sql += " AND '"+str(fdate)+"' BETWEEN ActiveDate AND InActiveDate"
    if fwctype == "WC":
        sql += " AND HO.WorkCenterNo = '"+fkey+"'"
    elif fwctype == "WCG":
        sql += " AND WorkCenterGroup = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    elif fwctype == "RT":
        sql += " AND OnRouting = '"+fkey+"' AND WorkCenterType = 'Machine' AND IsActive = 1"
    sql += ")) AS TB"
    sql += " GROUP BY day(CONVERT(DATE, Fdate))"
    cursor.execute(sql)
    return cursor.fetchall()

def getMonthlyWorkCenterManualForABGraph(fwctype, fkey, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT DAY(Date) Fday, SUM(WorkingHour) AS WorkingHour"
    sql += " FROM ABGraphManualData AS AB"
    sql += " INNER JOIN WorkCenter AS WC ON AB.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE MONTH(Date) = '"+str(month)+"' AND YEAR(Date) = '"+str(year)+"'"
    if fwctype == "WC":
        sql += " AND AB.WorkCenterNo = '"+str(fkey)+"'"
    elif fwctype == "WCG":
        sql += " AND WC.WorkCenterGroup = '"+str(fkey)+"'"
    elif fwctype == "RT":
        sql += " AND WC.OnRouting = '"+str(fkey)+"'"
    sql += " GROUP BY Date"
    cursor.execute(sql)
    return cursor.fetchall()

def getOrderStopNotStart():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM OrderControl AS OC WHERE OC.ProcessStart IS NULL AND OC.ProcessStop IS NOT NULL"
    cursor.execute(sql)
    return cursor.fetchall()

def getOperationRemainQtyList():
    cursor = get_connection().cursor()
    sql = """
            SELECT * FROM OrderControl AS OC INNER JOIN OperationControl AS OPC ON OC.OrderNo = OPC.OrderNo
            WHERE OC.ProcessStop IS NOT NULL AND ProcessQty - (AcceptedQty + RejectedQty) > 0
        """
    cursor.execute(sql)
    return cursor.fetchall()

def getLastProcessStopOrderNotStop():
    cursor = get_connection().cursor()
    sql = """
            SELECT *, OPC.ProcessStop AS LastProcessStop FROM OperationControl AS OPC INNER JOIN OrderControl AS OC ON OPC.OrderNo = OC.OrderNo
            WHERE CONCAT(OPC.OrderNo, OPC.OperationNo) IN
            (SELECT CONCAT(OrderNo, MAX(OperationNo)) AS PROD FROM OperationControl
            GROUP BY OrderNo)
            AND OPC.ProcessStop IS NOT NULL AND OC.ProcessStop IS NULL
        """
    cursor.execute(sql)
    return cursor.fetchall()

def getConfirmOperationList(fwc, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT ConfirmDateTime, OPC1.WorkCenterNo, ORD.FG_MaterialCode, ORD.FG_Drawing, HC.OrderNo, HC.OperationNo, EmpID, OPC1.ProcessQty, HC.AcceptedQty, HC.RejectedQty, RejectReason, ScrapAt, OPC2.WorkCenterNo  As ScrapAtWorkCenter"
    sql += " FROM HistoryConfirm AS HC INNER JOIN OperationControl AS OPC1 ON HC.OrderNo = OPC1.OrderNo AND HC.OperationNo = OPC1.OperationNo"
    sql += " INNER JOIN OrderControl AS ORD ON HC.OrderNo = ORD.OrderNo"
    sql += " LEFT JOIN OperationControl AS OPC2 ON HC.OrderNo = OPC2.OrderNo AND HC.ScrapAt = OPC2.OperationNo"
    sql += " WHERE month(ConfirmDateTime) = '"+month+"' AND year(ConfirmDateTime) = '"+year+"'"
    if fwc != 'ALL':
        sql += " AND OPC1.WorkCenterNo = '"+fwc+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getOverEstOperationList(fwc, fweek):
    start_date, end_date = get_date_between_week(fweek)
    start_date = str(start_date)[0:10]
    end_date = str(end_date)[0:10]
    cursor = get_connection().cursor()
    sql = "SELECT TB1.*, TB2.ActualSetup, TB2.ActualOper, TB2.ActualLabor, TB3.FG_MaterialCode, TB3.FG_Drawing, CONVERT(int, TB3.ProductionOrderQuatity) AS OrderQty, CT.TargetValue FROM"
    sql += " (SELECT * FROM OperationControl WHERE OrderNo NOT IN (SELECT OrderNo FROM CanceledOrder)"
    sql += f" AND (ProcessStop BETWEEN '{start_date}' AND '{end_date}')"
    if fwc != 'ALL':
        sql += " AND WorkCenterNo = '"+fwc+"'"
    sql += ") AS TB1 LEFT JOIN"
    sql += " (SELECT OrderNo, OperationNo, SUM(Setup) AS ActualSetup , SUM(Oper) AS ActualOper, SUM(Labor) AS ActualLabor FROM HistoryOperate GROUP BY OrderNo, OperationNo) AS TB2"
    sql += " ON TB1.OrderNo = TB2.OrderNo AND TB1.OperationNo = TB2.OperationNo INNER JOIN OrderControl AS TB3 ON TB1.OrderNo = TB3.OrderNo"
    sql += " INNER JOIN WorkCenter AS WC ON TB1.WorkCenterNo = WC.WorkCenterNo"
    sql += " LEFT JOIN CycleTimeTarget AS CT ON WC.MainProcess = CT.MainProcess AND TB3.FG_MaterialCode = CT.FG_MaterialCode"
    cursor.execute(sql)
    # print(sql)
    return cursor.fetchall()

def getEmpAtComputerList(ip_address):
    cursor = get_connection().cursor()
    sql = f"""SELECT * FROM EmpAtComputer AS EAC INNER JOIN Employee AS EMP ON EAC.EmpID = EMP.EmpID WHERE IPAddress = '{ip_address}' ORDER BY DateTimeStamp DESC"""
    cursor.execute(sql)
    return cursor.fetchall()

def getPendingPLNFAIList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM PendingPLNFAI AS PPF LEFT JOIN OrderControl AS OC ON PPF.OrderNo = OC.OrderNo"
    cursor.execute(sql)
    return cursor.fetchall()

def getOrderListBySaleNo(so_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM SAP_Order WHERE SalesOrderNo = '" + str(so_no) + "'"
    cursor.execute(sql)
    return cursor.fetchall()

def getOperatingHistoryConfirm(fwc, fmc, ftype, fstartdate, fstopdate, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT HC.WorkCenterNo AS Machine, OPC.WorkCenterNo AS WorkCenter, HC.AcceptedQty AS Yeild, HC.RejectedQty AS Scrap, OPC.AcceptedQty AS AcceptedAll, OPC.RejectedQty AS RejectedAll, * FROM HistoryConfirm AS HC"
    sql += " INNER JOIN OperationControl AS OPC ON HC.OrderNo = OPC.OrderNo AND HC.OperationNo = OPC.OperationNo"
    sql += " INNER JOIN OrderControl AS OC ON HC.OrderNo = OC.OrderNo"
    sql += " INNER JOIN Employee AS EMP ON HC.EmpID = EMP.EmpID"
    sql += " WHERE"
    if ftype == "MONTHLY":
        sql += " month(HC.ConfirmDateTime) = '"+str(month)+"' AND year(HC.ConfirmDateTime) = '"+str(year)+"'"
    else:
        if ftype == "DAILY":
            fstopdate = fstartdate
        sql += " HC.ConfirmDateTime >= '" + fstartdate + " 00:00:00' AND HC.ConfirmDateTime <= '" + fstopdate + " 23:59:59'"
    if fwc != 'ALL':
        sql += " AND OPC.WorkCenterNo = '"+str(fwc)+"'"
    if fmc != 'ALL':
        sql += " AND HC.WorkCenterNo = '"+str(fmc)+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getOperatingHistoryOperate(fwc, fmc, ftype, fstartdate, fstopdate, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT HO.WorkCenterNo AS Machine, OPC.WorkCenterNo AS WorkCenter, * FROM HistoryOperate AS HO"
    sql += " INNER JOIN OperationControl AS OPC ON HO.OrderNo = OPC.OrderNo AND HO.OperationNo = OPC.OperationNo"
    sql += " INNER JOIN OrderControl AS OC ON HO.OrderNo = OC.OrderNo"
    sql += " LEFT JOIN Employee AS EMP ON HO.EmpID = EMP.EmpID"
    sql += " LEFT JOIN IdleType AS IT ON HO.IdleCode = IT.Code"
    sql += " WHERE HO.Setup + HO.Oper + HO.Labor + HO.Idle > 0 AND"
    if ftype == "MONTHLY":
        sql += " month(HO.StartDateTime) = '"+str(month)+"' AND year(HO.StartDateTime) = '"+str(year)+"'"
    else:
        if ftype == "DAILY":
            fstopdate = fstartdate
        sql += " HO.StartDateTime >= '" + fstartdate + " 00:00:00' AND HO.StartDateTime <= '" + fstopdate + " 23:59:59'"
    if fwc != 'ALL':
        sql += " AND OPC.WorkCenterNo = '"+str(fwc)+"'"
    if fmc != 'ALL':
        sql += " AND HO.WorkCenterNo = '"+str(fmc)+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getABGraphCollectedData(type, wc, opr, month, year):
    cursor = get_connection().cursor()
    sql = "SELECT *, day(CONVERT(DATE, Date)) AS Fday FROM ABGraphData WHERE Type = '"+str(type)+"' AND WorkCenterNo = '"+str(wc)+"'"
    sql += " AND Operation = '"+str(opr)+"' AND month(Date) = '"+str(month)+"' AND year(Date) = '"+str(year)+"' ORDER BY Date ASC"
    cursor.execute(sql)
    return cursor.fetchall()

def getMachineWorkOnOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM HistoryOperate AS HO INNER JOIN WorkCenter AS WC ON HO.WorkCenterNo = WC.WorkCenterNo WHERE IsRouting = 0 AND OrderNo = '"+str(order_no)+"' AND OperationNo = '"+str(operation_no)+"'"
    cursor.execute(sql)
    return cursor.fetchall()

def getRecordRejectionPFC(pfc, month, year):
    cursor = get_connection().cursor()
    sql = f"""
        SELECT HC.*, WC.*, OC.*, OPC.ProcessQty FROM HistoryConfirm AS HC 
        INNER JOIN (SELECT * FROM WorkCenter WHERE ProfitCenter = '{pfc}') AS WC ON HC.WorkCenterNo = WC.WorkCenterNo
        INNER JOIN OrderControl AS OC ON HC.OrderNo = OC.OrderNo 
        INNER JOIN OperationControl AS OPC ON HC.OrderNo = OPC.OrderNo AND HC.OperationNo = OPC.OperationNo
        WHERE HC.RejectReason != 'MATERIAL QUANTITY ADJUSTMENT' AND HC.RejectedQty > 0 AND MONTH(ConfirmDateTime) = '{month}' AND YEAR(ConfirmDatetime) = '{year}'
        """
    cursor.execute(sql)
    return cursor.fetchall()

def getToolMasterList():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [ToolMaster]"
    cursor.execute(sql)
    return cursor.fetchall()

def getToolHeaderListByOrder(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM ToolHeader WHERE OrderNo = '{order_no}' AND OperationNo = '{operation_no}' ORDER BY DateTimeStamp"
    cursor.execute(sql)
    return cursor.fetchall()

def getToolStepList(id):
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM ToolStep WHERE ToolHeaderID = {id} ORDER BY StepOrder"
    cursor.execute(sql)
    return cursor.fetchall()

def getToolItemList(id):
    cursor = get_connection().cursor()
    sql = f"SELECT TI.CTCode AS CTCode1, * FROM ToolItem AS TI LEFT JOIN ToolMaster AS TM ON TI.CTCode = TM.CTCode WHERE ToolHeaderID = {id} ORDER BY No ASC"
    cursor.execute(sql)
    return cursor.fetchall()

def getToolHeaderAll():
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM ToolHeader ORDER BY DateTimeStamp DESC"
    cursor.execute(sql)
    return cursor.fetchall()

def getPlanActualCycleTimeList(ffgcode, fwc, ftype, fstartdate, fstopdate, fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    cursor = get_connection().cursor()
    sql = "SELECT TB.ProcessStart, TB.ProcessStop, HO.OrderNo, HO.OperationNo, TB.FG_MaterialCode, TB.FG_Drawing, TB.WorkCenterNo, TB.TargetValue, TB.ProcessQty, TB.AcceptedQty, TB.RejectedQty,"
    sql += "TB.EstSetupTime AS EstSetup, TB.EstOperationTime AS EstOper, TB.EstLaborTime AS EstLabor," 
    sql += "(TB.EstSetupTime * TB.ProcessQty) AS Est2Setup, (TB.EstOperationTime * TB.ProcessQty) AS Est2Oper, (TB.EstLaborTime * TB.ProcessQty) AS Est2Labor,"
    sql += "SUM(HO.Setup) AS ActSetup, SUM(Ho.Oper) AS ActOper, SUM(HO.Labor) AS ActLabor, SUM(HO.Idle) AS Idle, (SUM(Ho.Oper) + SUM(HO.Idle)) AS AI"
    sql += " FROM HistoryOperate AS HO INNER JOIN"
    sql += " (SELECT OC.FG_MaterialCode, OC.FG_Drawing, OPC.* , CT.TargetValue FROM OrderControl AS OC"
    sql += " INNER JOIN OperationControl AS OPC ON OC.OrderNo = OPC.OrderNo"
    sql += " INNER JOIN WorkCenter AS WC ON OPC.WorkCenterNo = WC.WorkCenterNo"
    sql += " LEFT JOIN CycleTimeTarget AS CT ON WC.MainProcess = CT.MainProcess AND OC.FG_MaterialCode = CT.FG_MaterialCode"
    sql += f" WHERE OPC.ProcessStop IS NOT NULL AND OC.FG_MaterialCode LIKE '{ffgcode}' AND OPC.WorkCenterNo = '{fwc}'"
    if ftype == 'DAILY':
        fstopdate = fstartdate
        sql += f" AND OPC.ProcessStart >= '{fstartdate} 00:00:00' AND OPC.ProcessStart <= '{fstopdate} 23:59:59'"
    elif ftype == 'RANGE':
        sql += f" AND OPC.ProcessStart >= '{fstartdate} 00:00:00' AND OPC.ProcessStart <= '{fstopdate} 23:59:59'"
    elif ftype == 'MONTHLY':
        sql += f" AND month(OPC.ProcessStart) = '{month}' AND year(OPC.ProcessStart) = '{year}'"
    sql += ") AS TB ON HO.OrderNo = TB.OrderNo AND HO.OperationNo = TB.OperationNo"
    sql += " GROUP BY TB.ProcessStart, TB.ProcessStop, HO.OrderNo, HO.OperationNo, TB.FG_MaterialCode, TB.FG_Drawing, TB.WorkCenterNo, TB.TargetValue,"
    sql += " TB.EstSetupTime, TB.EstOperationTime, TB.EstLaborTime, TB.ProcessQty, TB.ProcessQty, TB.AcceptedQty, TB.RejectedQty"
    # print(sql)
    cursor.execute(sql)
    return cursor.fetchall()

def getIdleList(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = f"""SELECT OrderNo, OperationNo, IdleCode, Description, SUM(Idle) AS Idle
            FROM HistoryOperate AS HO 
            INNER JOIN IdleType AS IT ON HO.IdleCode = IT.Code 
            WHERE Idle > 0 AND OrderNo = '{order_no}' AND OperationNo = '{operation_no}'
            GROUP BY OrderNo, OperationNo, IdleCode, Description"""
    cursor.execute(sql)
    return cursor.fetchall()

def getCycleTimeTargetList():
    cursor = get_connection().cursor()
    sql = """
      SELECT
            C1.FG_MaterialCode,
            C2.TargetValue AS 'MCA3Axis',
            C3.TargetValue AS 'HMC',
            C4.TargetValue AS 'TMA',
            C5.TargetValue AS 'MCA5Axis',
            C6.TargetValue AS 'CLA',
            C7.TargetValue AS 'BTA',
            C8.TargetValue AS 'OG',
            C9.TargetValue AS 'EDA',
            C10.TargetValue AS 'WCA',
            C11.TargetValue AS 'CLN_AS',
            C12.TargetValue AS 'TBL_AS',
            C13.TargetValue AS 'IQC_AS',
            C14.TargetValue AS 'FPI',
            C15.TargetValue AS 'MPI',
            C16.TargetValue AS 'STC',
            C17.TargetValue AS 'STD',
            C18.TargetValue AS 'STM',
            C19.TargetValue AS 'STH',
            C20.TargetValue AS 'IDHEN_OEM',
            C21.TargetValue AS 'MSM_OEM',
            C22.TargetValue AS 'VIB_OEM',
            C23.TargetValue AS 'STR_OEM',
            C24.TargetValue AS 'STR_VIB',
            C25.TargetValue AS 'ASSY_OEM',
            C26.TargetValue AS 'PNT_OEM',
            C27.TargetValue AS 'INK_OEM',
            C28.TargetValue AS 'STP',
            C29.TargetValue AS 'LS_AS',
            C30.TargetValue AS 'FQC_AS',
            C31.TargetValue AS 'PK_AS',
            C32.TargetValue AS 'IG'
        FROM
            CycleTimeTarget AS C1
            LEFT JOIN CycleTimeTarget AS C2 ON C1.FG_MaterialCode = C2.FG_MaterialCode AND C2.MainProcess = 'MCA 3 Axis'
            LEFT JOIN CycleTimeTarget AS C3 ON C1.FG_MaterialCode = C3.FG_MaterialCode AND C3.MainProcess = 'HMC'
            LEFT JOIN CycleTimeTarget AS C4 ON C1.FG_MaterialCode = C4.FG_MaterialCode AND C4.MainProcess = 'TMA'
            LEFT JOIN CycleTimeTarget AS C5 ON C1.FG_MaterialCode = C5.FG_MaterialCode AND C5.MainProcess = '5 Axis'
            LEFT JOIN CycleTimeTarget AS C6 ON C1.FG_MaterialCode = C6.FG_MaterialCode AND C6.MainProcess = 'CLA'
            LEFT JOIN CycleTimeTarget AS C7 ON C1.FG_MaterialCode = C7.FG_MaterialCode AND C7.MainProcess = 'BTA'
            LEFT JOIN CycleTimeTarget AS C8 ON C1.FG_MaterialCode = C8.FG_MaterialCode AND C8.MainProcess = 'OG'
            LEFT JOIN CycleTimeTarget AS C9 ON C1.FG_MaterialCode = C9.FG_MaterialCode AND C9.MainProcess = 'EDA'
            LEFT JOIN CycleTimeTarget AS C10 ON C1.FG_MaterialCode = C10.FG_MaterialCode AND C10.MainProcess = 'WCA'
            LEFT JOIN CycleTimeTarget AS C11 ON C1.FG_MaterialCode = C11.FG_MaterialCode AND C11.MainProcess = 'CLN_AS'
            LEFT JOIN CycleTimeTarget AS C12 ON C1.FG_MaterialCode = C12.FG_MaterialCode AND C12.MainProcess = 'TBL_AS'
            LEFT JOIN CycleTimeTarget AS C13 ON C1.FG_MaterialCode = C13.FG_MaterialCode AND C13.MainProcess = 'IQC_AS'
            LEFT JOIN CycleTimeTarget AS C14 ON C1.FG_MaterialCode = C14.FG_MaterialCode AND C14.MainProcess = 'FPI'
            LEFT JOIN CycleTimeTarget AS C15 ON C1.FG_MaterialCode = C15.FG_MaterialCode AND C15.MainProcess = 'MPI'
            LEFT JOIN CycleTimeTarget AS C16 ON C1.FG_MaterialCode = C16.FG_MaterialCode AND C16.MainProcess = 'STC'
            LEFT JOIN CycleTimeTarget AS C17 ON C1.FG_MaterialCode = C17.FG_MaterialCode AND C17.MainProcess = 'STD'
            LEFT JOIN CycleTimeTarget AS C18 ON C1.FG_MaterialCode = C18.FG_MaterialCode AND C18.MainProcess = 'STM'
            LEFT JOIN CycleTimeTarget AS C19 ON C1.FG_MaterialCode = C19.FG_MaterialCode AND C19.MainProcess = 'STH'
            LEFT JOIN CycleTimeTarget AS C20 ON C1.FG_MaterialCode = C20.FG_MaterialCode AND C20.MainProcess = 'IDHEN_OEM'
            LEFT JOIN CycleTimeTarget AS C21 ON C1.FG_MaterialCode = C21.FG_MaterialCode AND C21.MainProcess = 'MSM_OEM'
            LEFT JOIN CycleTimeTarget AS C22 ON C1.FG_MaterialCode = C22.FG_MaterialCode AND C22.MainProcess = 'VIB_OEM'
            LEFT JOIN CycleTimeTarget AS C23 ON C1.FG_MaterialCode = C23.FG_MaterialCode AND C23.MainProcess = 'STR_OEM'
            LEFT JOIN CycleTimeTarget AS C24 ON C1.FG_MaterialCode = C24.FG_MaterialCode AND C24.MainProcess = 'STR_VIB'
            LEFT JOIN CycleTimeTarget AS C25 ON C1.FG_MaterialCode = C25.FG_MaterialCode AND C25.MainProcess = 'ASSY_OEM'
            LEFT JOIN CycleTimeTarget AS C26 ON C1.FG_MaterialCode = C26.FG_MaterialCode AND C26.MainProcess = 'PNT_OEM'
            LEFT JOIN CycleTimeTarget AS C27 ON C1.FG_MaterialCode = C27.FG_MaterialCode AND C27.MainProcess = 'INK_OEM'
            LEFT JOIN CycleTimeTarget AS C28 ON C1.FG_MaterialCode = C28.FG_MaterialCode AND C28.MainProcess = 'STP'
            LEFT JOIN CycleTimeTarget AS C29 ON C1.FG_MaterialCode = C29.FG_MaterialCode AND C29.MainProcess = 'LS_AS'
            LEFT JOIN CycleTimeTarget AS C30 ON C1.FG_MaterialCode = C30.FG_MaterialCode AND C30.MainProcess = 'FQC_AS'
            LEFT JOIN CycleTimeTarget AS C31 ON C1.FG_MaterialCode = C31.FG_MaterialCode AND C31.MainProcess = 'PK_AS'
            LEFT JOIN CycleTimeTarget AS C32 ON C1.FG_MaterialCode = C32.FG_MaterialCode AND C32.MainProcess = 'IG'
        GROUP BY
            C1.FG_MaterialCode,
            C2.TargetValue,
            C3.TargetValue,
            C4.TargetValue,
            C5.TargetValue,
            C6.TargetValue,
            C7.TargetValue,
            C8.TargetValue,
            C9.TargetValue,
            C10.TargetValue,
            C11.TargetValue,
            C12.TargetValue,
            C13.TargetValue,
            C14.TargetValue,
            C15.TargetValue,
            C16.TargetValue,
            C17.TargetValue,
            C18.TargetValue,
            C19.TargetValue,
            C20.TargetValue,
            C21.TargetValue,
            C22.TargetValue,
            C23.TargetValue,
            C24.TargetValue,
            C25.TargetValue,
            C26.TargetValue,
            C27.TargetValue,
            C28.TargetValue,
            C29.TargetValue,
            C30.TargetValue,
            C31.TargetValue,
            C32.TargetValue
        ORDER BY
            C1.FG_MaterialCode
    """
    cursor.execute(sql)
    return cursor.fetchall()

def getCumulativeRejectionList(profitcenter, month_no, year):
    # Get WorkCenter of MT_AS to Improve Performance
    cursor = get_connection().cursor()
    sql = "SELECT WorkCenterNo FROM WorkCenter WHERE OnRouting = 'MT_AS'"
    cursor.execute(sql)
    tmps = cursor.fetchall()
    wcs = "(" + ", ".join(["'" + tmp.WorkCenterNo + "'" for tmp in tmps]) + ")" if len(tmps) > 0 else "('')"
    # Main
    cursor = get_connection().cursor()
    sql = f"""
        SELECT DAY(OrderStop) AS Day, ProfitCenter, SUM(OrderQty) AS OrderQty, SUM(RejectQty) AS RejectQty, ROUND(((SUM(RejectQty) / SUM(OrderQty)) * 100), 3) AS RejectPercentage FROM
        (
        SELECT OPC.OrderNo, MAX(WC.ProfitCenter) AS ProfitCenter, OC.ProcessStop AS OrderStop, OC.ProductionOrderQuatity AS OrderQty
        FROM OperationControl AS OPC
        INNER JOIN WorkCenter AS WC ON OPC.WorkCenterNo = WC.WorkCenterNo
        INNER JOIN OrderControl AS OC ON OPC.OrderNo = OC.OrderNo
        WHERE YEAR(OC.ProcessStop) = '{year}' AND MONTH(OC.ProcessStop) = '{month_no}' AND OPC.OrderNo NOT IN (SELECT OrderNo From CanceledOrder) 
        GROUP BY OPC.OrderNo, OC.ProcessStop, OC.ProductionOrderQuatity
        ) AS TB1
        INNER JOIN
        (
        SELECT OrderNo, CAST(SUM(RejectedQty) AS INT) AS RejectQty FROM HistoryConfirm WHERE WorkCenterNo NOT IN {wcs} GROUP BY OrderNo
        ) AS TB2
        ON TB1.OrderNo = TB2.OrderNo
        WHERE ProfitCenter = '{profitcenter}'
        GROUP BY ProfitCenter, DAY(OrderStop)
    """
    cursor.execute(sql)
    return cursor.fetchall()

def getCumulativeRejectionRecordList(profitcenter, month_no, year, is_rej_only):
    # Get WorkCenter of MT_AS to Improve Performance
    cursor = get_connection().cursor()
    sql = "SELECT WorkCenterNo FROM WorkCenter WHERE OnRouting = 'MT_AS'"
    cursor.execute(sql)
    tmps = cursor.fetchall()
    wcs = "(" + ", ".join(["'" + tmp.WorkCenterNo + "'" for tmp in tmps]) + ")" if len(tmps) > 0 else "('')"
    # Main
    cursor = get_connection().cursor()
    rej_only_sql = 'AND RejectQty > 0' if is_rej_only else ''
    sql = f"""
        SELECT OrderStop, TB1.OrderNo, OrderQty, RejectQty, FG_MaterialCode, FG_Drawing, DrawingNo FROM
        (
        SELECT OPC.OrderNo, MAX(WC.ProfitCenter) AS ProfitCenter, OC.ProcessStop AS OrderStop, OC.ProductionOrderQuatity AS OrderQty, OC.FG_MaterialCode, OC.FG_Drawing, OC.DrawingNo
        FROM OperationControl AS OPC
        INNER JOIN WorkCenter AS WC ON OPC.WorkCenterNo = WC.WorkCenterNo
        INNER JOIN OrderControl AS OC ON OPC.OrderNo = OC.OrderNo
        WHERE YEAR(OC.ProcessStop) = '{year}' AND MONTH(OC.ProcessStop) = '{month_no}' AND OPC.OrderNo NOT IN (SELECT OrderNo From CanceledOrder) 
        GROUP BY OPC.OrderNo, OC.ProcessStop, OC.ProductionOrderQuatity, OC.FG_MaterialCode, OC.FG_Drawing, OC.DrawingNo
        ) AS TB1
        INNER JOIN
        (
        SELECT OrderNo, CAST(SUM(RejectedQty) AS INT) AS RejectQty FROM HistoryConfirm WHERE WorkCenterNo NOT IN {wcs} GROUP BY OrderNo
        ) AS TB2
        ON TB1.OrderNo = TB2.OrderNo
        WHERE ProfitCenter = '{profitcenter}' {rej_only_sql}
    """
    cursor.execute(sql)
    return cursor.fetchall()

#-------------------------------------------------------------------------- ITEM
def getOrder(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OrderControl] WHERE OrderNo = '" + order_no + "'"
    cursor.execute(sql)
    return cursor.fetchone()

def getOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperationControl] as OPT"
    sql += " LEFT JOIN [WorkCenter] as WC ON OPT.WorkCenterNo = WC.WorkCenterNo"
    # sql += " LEFT JOIN [SAP_Routing] as SAPRT ON OPT.OrderNo = SAPRT.ProductionOrderNo AND OPT.OperationNo = SAPRT.OperationNumber"
    sql += " WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    return cursor.fetchone()

def getPreviousOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM OperationControl WHERE OrderNo = '" + order_no + "' AND OperationNo < '" + operation_no + "' ORDER BY OperationNo DESC"
    cursor.execute(sql)
    return cursor.fetchone()

def getFirstOperation(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM OperationControl WHERE OrderNo = '" + order_no + "' ORDER BY OperationNo ASC"
    cursor.execute(sql)
    return cursor.fetchone()

def getLastOperation(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM OperationControl WHERE OrderNo = '" + order_no + "' ORDER BY OperationNo DESC"
    cursor.execute(sql)
    return cursor.fetchone()

def getWorkCenter(workcenter_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [WorkCenter] WHERE WorkCenterNo = '" + str(workcenter_no) + "'"
    cursor.execute(sql)
    return cursor.fetchone()

def getOperator(operator_id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [Employee] WHERE EmpID = '" + str(operator_id) +"'"
    cursor.execute(sql)
    return cursor.fetchone()

def getWorkCenterOperatingByWorkCenterNo(workcenter_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperatingWorkCenter] WHERE WorkCenterNo = '" + workcenter_no + "' AND Status <> 'COMPLETE'"
    cursor.execute(sql)
    return cursor.fetchone()

def getOperatorOperatingByEmpID(operator_id):
    cursor = get_connection().cursor()
    sql = "SELECT OOPR.OrderNo as OperatorOrderNo, OOPR.OperationNo as OperatorOperationNo, OOPR.StartDateTime as OperatorStartDateTime, OOPR.StopDateTime as OperatorStopDateTime, OWC.Status as WorkCenterStatus, *"
    sql += " FROM [OperatingOperator] as OOPR INNER JOIN [Employee] as EMP ON OOPR.EmpID = EMP.EmpID"
    sql += " LEFT JOIN [OperatingWorkCenter] as OWC ON OOPR.OperatingWorkCenterID = OWC.OperatingWorkCenterID"
    sql += " LEFT JOIN [WorkCenter] as WC ON OWC.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE OOPR.EmpID = " + str(operator_id) + " AND OOPR.Status <> 'COMPLETE' AND OOPR.Status <> 'EXT-WORK'"
    cursor.execute(sql)
    return cursor.fetchone()

def getWorkCenterOperatingByID(id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperatingWorkCenter] WHERE OperatingWorkCenterID = " + str(id)
    cursor.execute(sql)
    return cursor.fetchone()

def getOperatorOperatingByID(id):
    cursor = get_connection().cursor()
    sql = "SELECT OOPR.OrderNo as OperatorOrderNo, OOPR.OperationNo as OperatorOperationNo, OOPR.StartDateTime as OperatorStartDateTime, OOPR.StopDateTime as OperatorStopDateTime, OOPR.Status as OperatorStatus, OWC.Status as WorkCenterStatus, *"
    sql += " FROM [OperatingOperator] as OOPR INNER JOIN [Employee] as EMP ON OOPR.EmpID = EMP.EmpID"
    sql += " LEFT JOIN [OperatingWorkCenter] as OWC ON OOPR.OperatingWorkCenterID = OWC.OperatingWorkCenterID"
    sql += " LEFT JOIN [WorkCenter] as WC ON OWC.WorkCenterNo = WC.WorkCenterNo"
    sql += " WHERE OOPR.OperatingOperatorID = " + str(id)
    cursor.execute(sql)
    return cursor.fetchone()

def getNextOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperationControl] WHERE OrderNo = '" + order_no + "' AND OperationNo > '" + operation_no + "' ORDER BY OperationNo ASC"
    cursor.execute(sql)
    return cursor.fetchone()

def getUserByPassword(password):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [dbo].[User] WHERE PasswordHash = HASHBYTES('SHA2_512', '"+ password + "')"
    cursor.execute(sql)
    return cursor.fetchone()

def getUserByUserID(user_id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [dbo].[User] WHERE UserID = '"+ user_id + "'"
    cursor.execute(sql)
    return cursor.fetchone()

def getPTL(order_no, lot_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [PartialLotTraveller] WHERE OrderNo = '" + order_no + "' AND LotNo = " + lot_no
    cursor.execute(sql)
    return cursor.fetchone()

def getEmpIDByUserID(user_id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [dbo].[Employee] WHERE EmpID = '"+ str(user_id) + "'"
    cursor.execute(sql)
    return cursor.fetchone()

def getOvertimeHour():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [dbo].[AdminConfig] WHERE KeyText = 'OVERTIME_HOUR'"
    cursor.execute(sql)
    return int((cursor.fetchone()).Value)

def getManualReportAllow():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [AdminConfig] WHERE KeyText = 'MANUAL_REPORT_ALLOWDANCE'"
    cursor.execute(sql)
    return cursor.fetchone().Value.strip() == 'True'

def getRefreshSecond():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [dbo].[AdminConfig] WHERE KeyText = 'REFRESH_SECOND'"
    cursor.execute(sql)
    return int((cursor.fetchone()).Value)

def getDrawingAppPath():
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [dbo].[AdminConfig] WHERE KeyText = 'DRAWING_APP_PATH'"
    cursor.execute(sql)
    return str((cursor.fetchone()).Value).strip()

def getSizeOfMachineWorkCenterByGroup(fwcg, factive):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM WorkCenter WHERE WorkCenterGroup = '"+fwcg+"' AND WorkCenterType = 'Machine'"
    if factive == "ACTIVE":
        sql += " AND IsActive = 1"
    cursor.execute(sql)
    return len(cursor.fetchall())

def getNotFixedOvertime(operator_id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM OvertimeOperator WHERE EmpID = '" + operator_id + "' AND isFixed = 0"
    cursor.execute(sql)
    return cursor.fetchone()

def getFirstConfirmTime(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM HistoryConfirm WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "' ORDER BY ConfirmDateTime"
    cursor.execute(sql)
    return cursor.fetchone()

def getActualTime(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT SUM(Setup) AS Setup , SUM(Oper) AS Oper, SUM(Labor) AS Labor, SUM(COALESCE(Idle, 0)) AS Idle FROM HistoryOperate WHERE OrderNo = '" + str(order_no) + "' AND OperationNo = '" + str(operation_no) + "' GROUP BY OrderNo"
    cursor.execute(sql)
    return cursor.fetchone()

def getCanceledOrderInfo(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT *, CO.EmpID as Requester FROM CanceledOrder AS CO LEFT JOIN Employee AS EM ON CO.EmpID = EM.EmpID WHERE OrderNo = '"+order_no+"'"
    cursor.execute(sql)
    return cursor.fetchone()

def getABGraphManualData(wc, fdate):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM ABGraphManualData WHERE WorkCenterNo = '"+str(wc)+"' AND Date = '"+str(fdate)+"'"
    cursor.execute(sql)
    return cursor.fetchone()

def getEfficiencyDataPFC(fpfc, day, month, year):
    cursor = get_connection().cursor()
    sql = "SELECT SUM(ActualSetup) AS ActSetup, SUM(ActualOper) AS ActOper, SUM(ActualLabor) AS ActLabor, SUM(EstSetup) AS EstSetup, SUM(EstOper) AS EstOper, SUM(EstLabor) AS EstLabor FROM"
    sql += " (SELECT TB2.ActualSetup, TB2.ActualOper, TB2.ActualLabor, (EstSetupTime * ProcessQty) AS EstSetup, (EstOperationTime * ProcessQty) AS EstOper, (EstLaborTime * ProcessQty) AS EstLabor FROM"
    sql += " (SELECT OC.* FROM OperationControl AS OC INNER JOIN WorkCenter AS WC ON OC.WorkCenterNo = WC.WorkCenterNo WHERE OrderNo NOT IN (SELECT OrderNo FROM CanceledOrder)"
    sql += " AND day(ProcessStop) = '"+str(day)+"' AND month(ProcessStop) = '"+str(month)+"' AND year(ProcessStop) = '"+str(year)+"' AND ProfitCenter = '"+str(fpfc)+"') AS TB1 LEFT JOIN"
    sql += " (SELECT OrderNo, OperationNo, SUM(Setup) AS ActualSetup , SUM(Oper) AS ActualOper, SUM(Labor) AS ActualLabor FROM HistoryOperate GROUP BY OrderNo, OperationNo) AS TB2"
    sql += " ON TB1.OrderNo = TB2.OrderNo AND TB1.OperationNo = TB2.OperationNo INNER JOIN OrderControl AS TB3 ON TB1.OrderNo = TB3.OrderNo) AS TB4"
    cursor.execute(sql)
    # print(sql)
    return cursor.fetchone()

def getEfficiencyDataWCG(fwcg, day, month, year):
    cursor = get_connection().cursor()
    sql = "SELECT SUM(ActualSetup) AS ActSetup, SUM(ActualOper) AS ActOper, SUM(ActualLabor) AS ActLabor, SUM(EstSetup) AS EstSetup, SUM(EstOper) AS EstOper, SUM(EstLabor) AS EstLabor FROM"
    sql += " (SELECT TB2.ActualSetup, TB2.ActualOper, TB2.ActualLabor, (EstSetupTime * ProcessQty) AS EstSetup, (EstOperationTime * ProcessQty) AS EstOper, (EstLaborTime * ProcessQty) AS EstLabor FROM"
    sql += " (SELECT OC.* FROM OperationControl AS OC INNER JOIN WorkCenter AS WC ON OC.WorkCenterNo = WC.WorkCenterNo WHERE OrderNo NOT IN (SELECT OrderNo FROM CanceledOrder)"
    sql += " AND day(ProcessStop) = '"+str(day)+"' AND month(ProcessStop) = '"+str(month)+"' AND year(ProcessStop) = '"+str(year)+"' AND WorkCenterGroup = '"+str(fwcg)+"') AS TB1 LEFT JOIN"
    sql += " (SELECT OrderNo, OperationNo, SUM(Setup) AS ActualSetup , SUM(Oper) AS ActualOper, SUM(Labor) AS ActualLabor FROM HistoryOperate GROUP BY OrderNo, OperationNo) AS TB2"
    sql += " ON TB1.OrderNo = TB2.OrderNo AND TB1.OperationNo = TB2.OperationNo INNER JOIN OrderControl AS TB3 ON TB1.OrderNo = TB3.OrderNo) AS TB4"
    cursor.execute(sql)
    return cursor.fetchone()

def getEfficiencyDataWC(fwc, day, month, year):
    cursor = get_connection().cursor()
    sql = "SELECT SUM(ActualSetup) AS ActSetup, SUM(ActualOper) AS ActOper, SUM(ActualLabor) AS ActLabor, SUM(EstSetup) AS EstSetup, SUM(EstOper) AS EstOper, SUM(EstLabor) AS EstLabor FROM"
    sql += " (SELECT TB2.ActualSetup, TB2.ActualOper, TB2.ActualLabor, (EstSetupTime * ProcessQty) AS EstSetup, (EstOperationTime * ProcessQty) AS EstOper, (EstLaborTime * ProcessQty) AS EstLabor FROM"
    sql += " (SELECT * FROM OperationControl WHERE OrderNo NOT IN (SELECT OrderNo FROM CanceledOrder)"
    sql += " AND day(ProcessStop) = '"+str(day)+"' AND month(ProcessStop) = '"+str(month)+"' AND year(ProcessStop) = '"+str(year)+"' AND WorkCenterNo = '"+str(fwc)+"') AS TB1 LEFT JOIN"
    sql += " (SELECT OrderNo, OperationNo, SUM(Setup) AS ActualSetup , SUM(Oper) AS ActualOper, SUM(Labor) AS ActualLabor FROM HistoryOperate GROUP BY OrderNo, OperationNo) AS TB2"
    sql += " ON TB1.OrderNo = TB2.OrderNo AND TB1.OperationNo = TB2.OperationNo INNER JOIN OrderControl AS TB3 ON TB1.OrderNo = TB3.OrderNo) AS TB4"
    cursor.execute(sql)
    # print(sql)
    return cursor.fetchone()

def getRejectPerDataPFC(fpfc, day, month, year):
    cursor = get_connection().cursor()
    sql = "SELECT SUM(HC.AcceptedQty) + SUM(HC.RejectedQty) AS ProcessQty, SUM(HC.RejectedQty) AS RejectedQty"
    sql += " FROM HistoryConfirm AS HC INNER JOIN OperationControl AS OC ON HC.OrderNo = OC.OrderNo AND HC.OperationNo = OC.OperationNo"
    sql += " INNER JOIN WorkCenter AS WC ON WC.WorkCenterNo = OC.WorkCenterNo"
    sql += " WHERE HC.RejectReason != 'MATERIAL QUANTITY ADJUSTMENT' AND WC.ProfitCenter = '"+str(fpfc)+"' AND day(ConfirmDateTime) = '"+str(day)+"' AND month(ConfirmDateTime) = '"+str(month)+"' AND year(ConfirmDateTime) = '"+str(year)+"'"
    cursor.execute(sql)
    return cursor.fetchone()

def getRejectPerDataWCG(fwcg, day, month, year):
    cursor = get_connection().cursor()
    sql = "SELECT SUM(HC.AcceptedQty) + SUM(HC.RejectedQty) AS ProcessQty, SUM(HC.RejectedQty) AS RejectedQty"
    sql += " FROM HistoryConfirm AS HC INNER JOIN OperationControl AS OC ON HC.OrderNo = OC.OrderNo AND HC.OperationNo = OC.OperationNo"
    sql += " INNER JOIN WorkCenter AS WC ON WC.WorkCenterNo = OC.WorkCenterNo"
    sql += " WHERE HC.RejectReason != 'MATERIAL QUANTITY ADJUSTMENT' AND WC.WorkCenterGroup = '"+str(fwcg)+"' AND day(ConfirmDateTime) = '"+str(day)+"' AND month(ConfirmDateTime) = '"+str(month)+"' AND year(ConfirmDateTime) = '"+str(year)+"'"

    cursor.execute(sql)
    return cursor.fetchone()

def getRejectPerDataRT(fwc, day, month, year):
    cursor = get_connection().cursor()
    sql = "SELECT SUM(HC.AcceptedQty) + SUM(HC.RejectedQty) AS ProcessQty, SUM(HC.RejectedQty) AS RejectedQty"
    sql += " FROM HistoryConfirm AS HC INNER JOIN OperationControl AS OC ON HC.OrderNo = OC.OrderNo AND HC.OperationNo = OC.OperationNo"
    sql += " WHERE HC.RejectReason != 'MATERIAL QUANTITY ADJUSTMENT' AND OC.WorkCenterNo = '"+str(fwc)+"' AND day(ConfirmDateTime) = '"+str(day)+"' AND month(ConfirmDateTime) = '"+str(month)+"' AND year(ConfirmDateTime) = '"+str(year)+"'"
    cursor.execute(sql)
    return cursor.fetchone()

def getToolHeaderByOrder(order_no, operation_no, wc_no):
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM [ToolHeader] WHERE WorkCenterNo = '{wc_no}' AND OrderNo = '{order_no}' AND OperationNo = '{operation_no}'"
    cursor.execute(sql)
    return cursor.fetchone()

def getToolHeaderByID(toolh_id):
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM ToolHeader WHERE ID = '{toolh_id}'"
    cursor.execute(sql)
    return cursor.fetchone()

def getToolItem(tooli_id):
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM ToolItem WHERE ID = '{tooli_id}'"
    cursor.execute(sql)
    return cursor.fetchone()

def getTotalIdleTime(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = f"SELECT SUM(Idle) AS Idle FROM HistoryOperate AS HO WHERE Idle > 0 AND OrderNo = '{order_no}' AND OperationNo = '{operation_no}'"
    cursor.execute(sql)
    return cursor.fetchone()

def getTotalEmpWorkTimeFromHRFocus(emp_id, start_date, stop_date):
    cursor = get_connection_hr_focus().cursor()
    sql = f"""
            SELECT TotalTime, TotalTime - (TotalTime % 60) As OptTime FROM
            (
            SELECT 
            COALESCE(SUM(
            (COALESCE(datediff(minute, Ch1, Ch2), 0) 
            + CASE WHEN COALESCE(datediff(minute, Ch3, Ch4), 0) >= 300 THEN datediff(minute, Ch3, Ch4) - 60 ELSE COALESCE(datediff(minute, Ch3, Ch4), 0) END
            + COALESCE(datediff(minute, Ch5, Ch6), 0) 
            + COALESCE(datediff(minute, Ch7, Ch8), 0) 
            + COALESCE(datediff(minute, Ch9, Ch10), 0)
            )
            ),0) As TotalTime
            FROM tbTRTimeTransaction WHERE EmpID = '{emp_id}' AND WorkDate BETWEEN '{start_date}' AND '{stop_date}' AND WtCode <> '8THNOT'
            ) AS TB1
        """
    cursor.execute(sql)
    return cursor.fetchone()

def getLastCanceledOrder():
    cursor = get_connection().cursor()
    sql = f"SELECT TOP(1) * FROM CanceledOrder ORDER BY DateTimeStamp DESC"
    cursor.execute(sql)
    return cursor.fetchone()

#----------------------------------------------------------------------- BOOLEAN

def isExistSAPOrder(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [SAP_Order] WHERE ProductionOrderNo = '" + order_no + "'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isExistSAPOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [SAP_Routing] WHERE ProductionOrderNo = '" + order_no + "' AND OperationNumber = '" + operation_no + "'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isExistOrder(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OrderControl] WHERE OrderNo = '" + order_no + "'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isExistOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperationControl] WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isExistOperator(emp_id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [Employee] WHERE EmpID = '" + emp_id + "'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isWorkCenterOperating(workcenter_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperatingWorkCenter] WHERE WorkCenterNo = '" + workcenter_no + "' AND Status <> 'COMPLETE'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isOperatorOperating(operator_id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperatingOperator] WHERE EmpID = " + str(operator_id) + " AND Status <> 'COMPLETE' AND Status <> 'EXT-WORK'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def hasOperatorOperating(owc_id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperatingOperator] WHERE OperatingWorkCenterID = " + str(owc_id) + " AND Status <> 'COMPLETE'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isOperatingOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperatingOperator] WHERE OrderNo = '" + order_no + "' and OperationNo = '" + operation_no + "' AND Status <> 'COMPLETE'"
    cursor.execute(sql)
    if len(cursor.fetchall()) > 0:
        return True
    sql = "SELECT * FROM [OperatingWorkCenter] WHERE OrderNo = '" + order_no + "' and OperationNo = '" + operation_no + "' AND Status <> 'COMPLETE'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def hasNotStartOrder(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OrderControl] WHERE OrderNo = '" + order_no + "' AND ProcessStart IS NULL"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def hasNotStartOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperationControl] WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "' AND ProcessStart IS NULL"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isFirstOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperationControl] WHERE OrderNo = '" + order_no + "' ORDER BY OperationNo ASC"
    cursor.execute(sql)
    result = cursor.fetchall()
    return (result[0].OperationNo.strip() == operation_no)

def isLastOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperationControl] WHERE OrderNo = '" + order_no + "' ORDER BY OperationNo DESC"
    cursor.execute(sql)
    result = cursor.fetchall()
    return (result[0].OperationNo.strip() == operation_no)

def isExistDeletedOperation(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [HistoryModifier] WHERE Type = 'DELETE' AND OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isOvertimeOperation(order_no, operation_no):
    ot = str(getOvertimeHour())
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperatingOperator] WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "' AND (Status = 'WORKING' OR Status = 'SETUP') AND (DATEADD(HOUR, "+ot+", StartDateTime) < CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    if len(cursor.fetchall()) > 0:
        return True
    else:
        cursor = get_connection().cursor()
        sql = "SELECT * FROM [OperatingWorkCenter] WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "' AND (Status = 'WORKING' OR Status = 'SETUP') AND (DATEADD(HOUR, "+ot+", StartDateTime) < CURRENT_TIMESTAMP)"
        cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isOvertimeOperator(oopr_id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperatingOperator] WHERE OperatingOperatorID = '"+str(oopr_id)+"' AND (Status = 'WORKING' OR Status = 'SETUP') AND (DATEADD(HOUR, 12, StartDateTime) < CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

#-- ONLY WORKING WORKCENTER
def isOvertimeWorkCenter(owc_id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OperatingWorkCenter] WHERE OperatingWorkCenterID = '"+str(owc_id)+"' AND Status = 'WORKING' AND (DATEADD(HOUR, 12, StartDateTime) < CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def hasReportTimeToSAP(order_no, operation_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [SFR2SAP_Report] WHERE ProductionOrderNo = '"+order_no+"' AND OperationNumber = '"+operation_no+"' AND LaborTime > 0"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isCanceledOrder(order_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [CanceledOrder] WHERE OrderNo = '" + order_no + "'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isNotFixedOvertime(operator_id):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [OvertimeOperator] WHERE EmpID = '" + operator_id + "'  AND isFixed = 0"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isExistWorkCentrNo(wc_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [WorkCenter] WHERE WorkCenterNo = '" + wc_no + "'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isCollectedABGraphData(type, wc, month, year):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [ABGraphData] WHERE Type = '"+str(type)+"' AND WorkCenterNo = '"+str(wc)+"' AND month(Date) = '"+str(month)+"' AND year(Date) = '"+str(year)+"'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isMachine(wc_no):
    cursor = get_connection().cursor()
    sql = "SELECT * FROM [WorkCenter] WHERE WorkCenterNo = '" + wc_no + "' AND WorkCenterType = 'Machine' AND IsRouting = 0"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isExistToolMaster(ct_code):
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM [ToolMaster] WHERE CTCode = '{ct_code}'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isExistToolHeader(order_no, operation_no, wc_no):
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM [ToolHeader] WHERE WorkCenterNo = '{wc_no}' AND OrderNo = '{order_no}' AND OperationNo = '{operation_no}'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

def isExistHistoryTool(tooli_id, order_no, operation_no):
    cursor = get_connection().cursor()
    sql = f"SELECT * FROM [HistoryTool] WHERE ToolItemID = {tooli_id} AND OrderNo = '{order_no}' AND OperationNo = '{operation_no}'"
    cursor.execute(sql)
    return (len(cursor.fetchall()) > 0)

#--------------------------------------------------------------------------- SET

def setDataFromSAP(order_no):
    setOrderControlFromSAP(order_no)
    setOperationControlFromSAP(order_no)
    return

def setOrderControlFromSAP(order_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "SELECT * FROM [SAP_Order] WHERE ProductionOrderNo = '" + order_no + "'"
    cursor.execute(sql)
    order = cursor.fetchone()
    SalesOrderNo = ""
    RequestDate = ""
    ReleaseDate = ""
    FG_Drawing = ""
    Aero = ""
    DateGetFromSAP = order.DateGetFromSAP.strftime("%Y-%m-%d %H:%M:%S")
    if order.SalesCreateDate == "00.00.0000":
        SalesOrderNo = "NULL"
    else:
        SalesOrderNo = "CONVERT(DATETIME,'"+order.SalesCreateDate+"',104)"
    if order.RequestDate == "00.00.0000":
        RequestDate = "NULL"
    else:
        RequestDate = "CONVERT(DATETIME,'"+order.RequestDate+"',104)"
    if order.ReleaseDate == "00.00.0000":
        ReleaseDate = "NULL"
    else:
        ReleaseDate = "CONVERT(DATETIME,'"+order.ReleaseDate+"',104)"
    if order.FG_Drawing != None:
        FG_Drawing = order.FG_Drawing
    if order.AeroSpace != None:
        Aero = order.AeroSpace
    else:
        FG_Drawing = order.FG_Drawing
    cursor = conn.cursor()
    sql = "INSERT INTO [OrderControl] ([OrderNo],[LotNo],[CustomerPONo],[PartNo],[PartName],[SalesOrderNo],[SalesCreateDate],[SalesOrderQuantity],[ProductionOrderQuatity],[FG_MaterialCode],[RM_MaterialCode],[MRP_Controller],[RequestDate],[ReleaseDate],[DrawingNo],[AeroSpace],[RoutingGroup],[RoutingGroupCounter],[Plant],[DateGetFromSAP],[FG_Drawing]) VALUES "
    sql += "('"+order_no+"',0,'"+order.CustomerPONo+"','"+order.PartNo+"','"+order.PartName.replace("'", " ")+"','"+order.SalesOrderNo+"',"
    sql += SalesOrderNo+","+str(order.SalesOrderQuantity)+","+str(order.ProductionOrderQuatity)+",'"+order.FG_MaterialCode+"','"+order.RM_MaterialCode+"',"
    sql += "'"+order.MRP_Controller+"',"+RequestDate+","+ReleaseDate+",'"+order.DrawingNo+"','"+Aero+"',"
    sql += "'"+order.RoutingGroup+"','"+order.RoutingGroupCounter+"','"+order.Plant+"','"+DateGetFromSAP+"','"+FG_Drawing+"')"
    cursor.execute(sql)
    conn.commit()
    return

def setOperationControlFromSAP(order_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "SELECT * FROM [SAP_Order] WHERE ProductionOrderNo = '" + order_no + "'"
    cursor.execute(sql)
    order = cursor.fetchone()
    sql = "SELECT * FROM [SAP_Routing] WHERE ProductionOrderNo = '" + order_no + "' ORDER BY OperationNumber ASC"
    cursor.execute(sql)
    operations = cursor.fetchall()
    inserted_operation_no_list = []
    for i in range(len(operations)):
        operationNo = frontZero(operations[i].OperationNumber.strip(), 4)
        #--
        date_get_from_sap = ""
        if operations[i].DateGetFromSAP != None:
            date_get_from_sap = str(operations[i].DateGetFromSAP)
        else:
            date_get_from_sap = str(datetime.now())
        date_get_from_sap = str(date_get_from_sap[0:19])
        #--
        if operationNo not in inserted_operation_no_list:
            #-- Special Case ST_CLLI to ST_CL LI
            wc = operations[i].WorkCenter
            if wc == 'ST_CLLI':
                wc = 'ST_CL LI'
            #--
            inserted_operation_no_list.append(operationNo)
            sql = "INSERT INTO [OperationControl] ([OrderNo],[OperationNo],[WorkCenterNo],[ProcessQty],[AcceptedQty],[RejectedQty],[PlanStartDate],[PlanFinishDate],[EstSetupTime],[EstOperationTime],[EstLaborTime],[DateGetFromSAP])"
            if i == 0:
                sql += " VALUES ('"+order_no+"','"+operationNo+"','"+wc+"',"+str(order.ProductionOrderQuatity)+",0,0,CONVERT(DATETIME, '"+str(operations[i].PlanStartDate)+"', 104),CONVERT(DATETIME, '"+str(operations[i].PlanFinishDate)+"', 104),"+str(operations[i].EstimateSetTime)+","+str(operations[i].EstimateOperationTime)+","+str(operations[i].EstimateLaborTime)+",'"+date_get_from_sap+"')"
            else:
                sql += " VALUES ('"+order_no+"','"+operationNo+"','"+wc+"',0,0,0,CONVERT(DATETIME, '"+str(operations[i].PlanStartDate)+"', 104),CONVERT(DATETIME, '"+str(operations[i].PlanFinishDate)+"', 104),"+str(operations[i].EstimateSetTime)+","+str(operations[i].EstimateOperationTime)+","+str(operations[i].EstimateLaborTime)+",'"+date_get_from_sap+"')"
            cursor.execute(sql)

        conn.commit()
    return

def setOperationQty(order_no, operation_no, process_qty, accepted_qty, rejected_qty):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OperationControl] SET ProcessQty = "+str(process_qty)+", AcceptedQty = "+str(accepted_qty)+", RejectedQty = "+str(rejected_qty)
    sql += " WHERE OrderNo = '"+order_no+"' AND OperationNo = '"+operation_no+"'"
    cursor.execute(sql)
    conn.commit()
    return

#------------------------------------------------------------------------ INSERT

def insertOperatingWorkCenter(order_no, operation_no, workcenter_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [OperatingWorkCenter] ([OrderNo],[OperationNo],[WorkCenterNo],[Status])"
    sql += " VALUES ('" + order_no + "','" + operation_no + "','" + workcenter_no + "','WAITING')"
    cursor.execute(sql)
    conn.commit()
    return

def insertOperatingOperator(order_no, operation_no, operator_id, owc_id, status):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [OperatingOperator] ([OrderNo],[OperationNo],[EmpID],[OperatingWorkCenterID],[Status],[StartDateTime])"
    sql += " VALUES ('" + order_no + "','" + operation_no + "','" + str(operator_id) + "','" + str(owc_id) + "','" + status + "',CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    return

def getInsertJoinOperatingWorkCenter(order_no, operation_no, workcenter_no, start_date_time, stop_date_time):
    startDateTime = start_date_time.strftime("%Y-%m-%d %H:%M:%S")
    stopDateTime = stop_date_time.strftime("%Y-%m-%d %H:%M:%S")
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [OperatingWorkCenter] ([OrderNo],[OperationNo],[WorkCenterNo],[StartDateTime],[StopDateTime],[Status])"
    sql += " VALUES ('" + order_no + "','" + operation_no + "','" + workcenter_no + "','" + startDateTime + "','" + stopDateTime + "','COMPLETE')"
    cursor.execute(sql)
    conn.commit()
    sql = "SELECT SCOPE_IDENTITY()"
    cursor.execute(sql)
    result = cursor.fetchall()
    if(len(result) == 0):
        return None
    return result[0][0]

def insertJoinOperatingOperator(order_no, operation_no, operator_id, owc_id, start_date_time, stop_date_time):
    startDateTime = start_date_time.strftime("%Y-%m-%d %H:%M:%S")
    stopDateTime = stop_date_time.strftime("%Y-%m-%d %H:%M:%S")
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [OperatingOperator] ([OrderNo],[OperationNo],[EmpID],[OperatingWorkCenterID],[StartDateTime],[StopDateTime],[Status])"
    sql += " VALUES ('" + order_no + "','" + operation_no + "','" + str(operator_id) + "','" + str(owc_id) + "','" + startDateTime + "','" + stopDateTime + "','COMPLETE')"
    cursor.execute(sql)
    conn.commit()
    return

def insertSFR2SAP_Report(workcenter, order_no, operation_no, yiled, scrap, setup_time, oper_time, labor_time, start_date_time, stop_date_time, emp_id):
    start_date = start_date_time.strftime("%Y%m%d")
    start_time = start_date_time.strftime("%H%M%S")
    stop_date = stop_date_time.strftime("%Y%m%d")
    stop_time = stop_date_time.strftime("%H%M%S")
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [SFR2SAP_Report] ([DateTimeStamp],[WorkCenter],[ProductionOrderNo],[OperationNumber],[Yiled],[Scrap],[SetupTime],[OperTime],[LaborTime],[StartDate],[StartTime],[FinishDate],[FinishTime],[EmployeeID])"
    sql += " VALUES (CURRENT_TIMESTAMP,"
    sql += "'" + str(workcenter) + "',"
    sql += "'" + str(order_no) + "',"
    sql += "'" + str(operation_no) + "',"
    sql += "'" + str(yiled) + "',"
    sql += "'" + str(scrap) + "',"
    sql += "'" + str(setup_time) + "',"
    sql += "'" + str(oper_time) + "',"
    sql += "'" + str(labor_time) + "',"
    sql += "'" + str(start_date) + "',"
    sql += "'" + str(start_time) + "',"
    sql += "'" + str(stop_date) + "',"
    sql += "'" + str(stop_time) + "',"
    sql += "'" + str(emp_id) + "')"
    cursor.execute(sql)
    conn.commit()
    return

def insertSFR2SAP_Modifier_Delete(order_no, operation_no):
    DateTimeStamp = "CURRENT_TIMESTAMP"
    # if datetime.now().hour == 23:
    #     DateTimeStamp = "DATEADD(HOUR,1,CURRENT_TIMESTAMP)"
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [SFR2SAP_Modifier] ([DateTimeStamp],[Mode],[OrderNo],[OperationNo])"
    sql += " VALUES (CURRENT_TIMESTAMP,'31',"
    sql += "'" + str(order_no) + "',"
    sql += "'" + str(operation_no) + "')"
    cursor.execute(sql)
    conn.commit()
    return

def insertSFR2SAP_Modifier_Add(order_no, operation_no, control_key, work_center_no, pdt, cost_element, price_unit, price, currency, mat_group, purchasing_group, purchasing_org, est_setup_time, est_operate_time, est_labor_time):
    DateTimeStamp = "CURRENT_TIMESTAMP"
    # if datetime.now().hour == 23:
    #     DateTimeStamp = "DATEADD(HOUR,1,CURRENT_TIMESTAMP)"
    conn = get_connection()
    cursor = conn.cursor()
    mode = ""
    sql = ""
    if control_key == "PP01":
        if int(float(est_setup_time)) == 0:
            est_setup_time = 'NULL'
        if int(float(est_operate_time)) == 0:
            est_operate_time = 'NULL'
        if int(float(est_labor_time)) == 0:
            est_labor_time = 'NULL'
        mode = "11"
        sql = "INSERT INTO [SFR2SAP_Modifier]"
        sql += " ([DateTimeStamp],[Mode],[OrderNo],[OperationNo],[ControlKey],[WorkCenter],[SetupTime],[OperTime],[LaborTime])"
        sql += " VALUES (CURRENT_TIMESTAMP,"
        sql += "'" + str(mode) + "',"
        sql += "'" + str(order_no) + "',"
        sql += "'" + str(operation_no) + "',"
        sql += "'" + str(control_key) + "',"
        sql += "'" + str(work_center_no) + "',"
        sql += "" + str(est_setup_time) + ","
        sql += "" + str(est_operate_time) + ","
        sql += "" + str(est_labor_time) + ")"
    elif control_key == "PP02":
        mode = "12"
        sql = "INSERT INTO [SFR2SAP_Modifier]"
        sql += " ([DateTimeStamp],[Mode],[OrderNo],[OperationNo],[ControlKey],[WorkCenter],[PlannedDeliveryTime],[CostElement],[PriceUnit],[Price],[Currency],[MaterialGroup],[PurchasingGroup],[PurchasingOrg])"
        sql += " VALUES (CURRENT_TIMESTAMP,"
        sql += "'" + str(mode) + "',"
        sql += "'" + str(order_no) + "',"
        sql += "'" + str(operation_no) + "',"
        sql += "'" + str(control_key) + "',"
        sql += "'" + str(work_center_no) + "',"
        sql += "" + str(pdt) + ","
        sql += "'" + str(cost_element) + "',"
        sql += "" + str(price_unit) + ","
        sql += "" + str(price) + ","
        sql += "'" + str(currency) + "',"
        sql += "'" + str(mat_group) + "',"
        sql += "'" + str(purchasing_group) + "',"
        sql += "'" + str(purchasing_org) + "')"
    cursor.execute(sql)
    conn.commit()
    return

def insertSFR2SAP_Modifier_Change(order_no, operation_no, control_key, work_center_no, pdt, cost_element, price_unit, price, currency, mat_group, purchasing_group, purchasing_org, est_setup_time, est_operate_time, est_labor_time):
    purchasing_group = purchasing_group.strip()
    conn = get_connection()
    cursor = conn.cursor()
    mode = ""
    sql = ""
    if control_key == "PP01":
        if int(float(est_setup_time)) == 0:
            est_setup_time = 'NULL'
        if int(float(est_operate_time)) == 0:
            est_operate_time = 'NULL'
        if int(float(est_labor_time)) == 0:
            est_labor_time = 'NULL'
        mode = "21"
        sql = "INSERT INTO [SFR2SAP_Modifier]"
        sql += " ([DateTimeStamp],[Mode],[OrderNo],[OperationNo],[ControlKey],[WorkCenter],[SetupTime],[OperTime],[LaborTime])"
        sql += " VALUES (CURRENT_TIMESTAMP,"
        sql += "'" + str(mode) + "',"
        sql += "'" + str(order_no) + "',"
        sql += "'" + str(operation_no) + "',"
        sql += "'" + str(control_key) + "',"
        sql += "'" + str(work_center_no) + "',"
        sql += "" + str(est_setup_time) + ","
        sql += "" + str(est_operate_time) + ","
        sql += "" + str(est_labor_time) + ")"
    elif control_key == "PP02":
        mode = "22"
        sql = "INSERT INTO [SFR2SAP_Modifier]"
        sql += " ([DateTimeStamp],[Mode],[OrderNo],[OperationNo],[ControlKey],[WorkCenter],[PlannedDeliveryTime],[CostElement],[PriceUnit],[Price],[Currency],[MaterialGroup],[PurchasingGroup],[PurchasingOrg])"
        sql += " VALUES (CURRENT_TIMESTAMP,"
        sql += "'" + str(mode) + "',"
        sql += "'" + str(order_no) + "',"
        sql += "'" + str(operation_no) + "',"
        sql += "'" + str(control_key) + "',"
        sql += "'" + str(work_center_no) + "',"
        sql += "" + str(pdt) + ","
        sql += "'" + str(cost_element) + "',"
        sql += "" + str(price_unit) + ","
        sql += "" + str(price) + ","
        sql += "'" + str(currency) + "',"
        sql += "'" + str(mat_group) + "',"
        sql += "'" + str(purchasing_group) + "',"
        sql += "'" + str(purchasing_org) + "')"
    cursor.execute(sql)
    conn.commit()
    return

def insertOperationControl(order_no, operation_no, work_center_no, plan_start_date, plan_finish_date, est_setup_time, est_operate_time, est_labor_time):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [OperationControl] ([OrderNo],[OperationNo],[WorkCenterNo],[ProcessQty],[AcceptedQty],[RejectedQty],[PlanStartDate],[PlanFinishDate],[EstSetupTime],[EstOperationTime],[EstLaborTime])"
    sql += " VALUES ('" + str(order_no) + "', '" + str(operation_no) + "', '" + str(work_center_no) + "', 0, 0, 0,'" + str(plan_start_date) + "','" + str(plan_finish_date) + "'," + str(est_setup_time) + "," + str(est_operate_time) + "," + str(est_labor_time) + ")"
    cursor.execute(sql)
    conn.commit()
    return

def insertHistoryOperate(order_no, operation_no, operator_id, workcenter_no, type, setup, oper, labor, idle, start_date_time, stop_date_time, code):
    startDateTime = start_date_time.strftime("%Y-%m-%d %H:%M:%S")
    stopDateTime = stop_date_time.strftime("%Y-%m-%d %H:%M:%S")
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [HistoryOperate] ([OrderNo],[OperationNo],[EmpID],[WorkCenterNo],[Type],[Setup],[Oper],[Labor],[Idle],[StartDateTime],[StopDateTime],[IdleCode])"
    sql += " VALUES ('"+order_no+"','"+operation_no+"',"+str(operator_id)+",'"+workcenter_no+"','"+type+"',"+str(setup)+","+str(oper)+","+str(labor)+","+str(idle)+",'"+startDateTime+"','"+stopDateTime+"',"+str(code)+")"
    cursor.execute(sql)
    conn.commit()
    return

def insertHistoryConfirm(order_no, operation_no, operator_id, workcenter_no, accept, reject, reason, scrap_at):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [HistoryConfirm] ([OrderNo],[OperationNo],[EmpID],[WorkCenterNo],[AcceptedQty],[RejectedQty],[RejectReason],[ScrapAt],[ConfirmDateTime])"
    sql += " VALUES ('"+order_no+"','"+operation_no+"','"+str(operator_id)+"','"+workcenter_no+"','"+str(accept)+"','"+str(reject)+"','"+reason+"','"+str(scrap_at)+"',CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    return

def insertHistoryJoin(order_no, operation_no, join_order_no, join_operation_no, start_date_time):
    startDateTime = start_date_time.strftime("%Y-%m-%d %H:%M:%S")
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [HistoryJoin] ([JoinToOrderNo],[JoinToOperationNo],[JoinByOrderNo],[JoinByOperationNo],[StartDateTime],[StopDateTime])"
    sql += " VALUES ('" + order_no + "','" + operation_no + "','" + join_order_no + "','" + join_operation_no + "','" + startDateTime + "',CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    return

def insertHistoryModifier(type, order_no, operation_no, emp_id, chief_id):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[HistoryModifier] ([Type],[OrderNo],[OperationNo],[EmpID],[ChiefID],[ModifyDateTime])"
    sql += " VALUES ('" + str(type) + "','" + str(order_no) + "','" + str(operation_no) + "','" + str(emp_id) + "','" + str(chief_id) + "',CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    return

def insertUser(user_id, user_password, user_role):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO dbo.[User] (UserID, PasswordHash, UserRole) VALUES('"+user_id+"', HASHBYTES('SHA2_512', '"+user_password+"'), '"+user_role+"')"
    cursor.execute(sql)
    conn.commit()
    return

def insertPLT(order_no, operation_no, lot_no, lot_qty, type, chief_id):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[PartialLotTraveller] ([OrderNo],[StartOperationNo],[LotNo],[LotQty],[Type],[ChiefID],[StartDateTime]) VALUES "
    sql += " ('"+order_no+"','"+operation_no+"',"+str(lot_no)+","+str(lot_qty)+",'"+type+"','"+chief_id+"',CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    return

def insertOvertimeWorkCenter(owc):
    startDateTime = owc.StartDateTime.strftime("%Y-%m-%d %H:%M:%S")
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [OvertimeWorkCenter] ([DateTimeStamp],[OrderNo],[OperationNo],[WorkCenterNo],[StartDateTime])"
    sql += " VALUES (CURRENT_TIMESTAMP,'"+owc.OrderNo+"','"+owc.OperationNo+"','"+owc.WorkCenterNo+"','"+startDateTime+"')"
    cursor.execute(sql)
    conn.commit()
    return

def insertOvertimeOperator(oopr):
    startDateTime = oopr.StartDateTime.strftime("%Y-%m-%d %H:%M:%S")
    WorkCenter = oopr.WorkCenterNo
    if WorkCenter == None:
        WorkCenter = '-1'
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [OvertimeOperator] ([DateTimeStamp],[OrderNo],[OperationNo],[EmpID],[WorkCenterNo],[Status],[StartDateTime])"
    sql += " VALUES (CURRENT_TIMESTAMP,'"+oopr.OrderNo+"','"+oopr.OperationNo+"','"+oopr.EmpID+"','"+WorkCenter+"','"+oopr.Status+"','"+startDateTime+"')"
    cursor.execute(sql)
    conn.commit()
    return

def insertEmpAtComputer(emp_id, ip_address):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[EmpAtComputer] ([DateTimeStamp],[EmpID],[IPAddress]) VALUES "
    sql += " (CURRENT_TIMESTAMP,'"+str(emp_id)+"','"+str(ip_address)+"')"
    cursor.execute(sql)
    conn.commit()
    return

def insertWorkCenter(type, wc_no, wc_name, wcg, on_rt, target, capacity, pfc, active_date, inactive_date):
    is_rt = 0 if type == 'auto_mc' or type == 'manual_mc' else 1
    wc_type = 'Labor' if type == 'labor_rt' or type == 'ext_rt' else 'Machine'
    is_ext = 1 if type == 'ext_rt' else 0
    mc_type = "NULL"
    if wc_type == 'Machine':
        mc_type = 'Auto' if type == 'auto_mc' or type == 'auto_mc_rt' else 'Manual'
    hour_rate = 600 if wc_type == 'Machine' else 150
    target = target if wc_type == 'Machine' else 0
    capacity = capacity if wc_type == 'Machine' else 0
    # NULLABLE
    if mc_type != "NULL":
        mc_type = "'"+str(mc_type)+"'"
    on_rt = "NULL" if on_rt == "" else "'"+str(on_rt)+"'"
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[WorkCenter]([WorkCenterNo],[WorkCenterName],[WorkCenterGroup],[IsRouting],[WorkCenterType],[IsExternalProcess],[MachineType],[HourlyRate],[IsActive],[Noted],[Target],[Capacity],[OnRouting],[ProfitCenter]) VALUES "
    sql += "('"+str(wc_no)+"','"+str(wc_name)+"','"+str(wcg)+"',"+str(is_rt)+",'"+str(wc_type)+"',"+str(is_ext)+","+str(mc_type)+","+str(hour_rate)+",1,NULL,"+str(target)+","+str(capacity)+","+str(on_rt)+",'"+str(pfc)+"')"
    cursor.execute(sql)
    conn.commit()
    return

def addPendingPLNFAI(order_no, operation_no, qty):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[PendingPLNFAI] ([OrderNo],[OperationNo],[Quantity],[DateTimeStamp]) VALUES "
    sql += " ('"+str(order_no)+"','"+str(operation_no)+"',"+str(qty)+",CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    return

def saveABGraphData(type, wc, opr, hr, day, month, year):
    date = frontZero(str(day), 2)
    fdate = year + '-' + month + '-' + date
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[ABGraphData] ([Type],[WorkCenterNo],[Operation],[Hour],[Date],[DateTimeStamp]) VALUES"
    sql += " ('"+str(type)+"','"+str(wc)+"','"+str(opr)+"',"+str(hr)+",'"+str(fdate)+"',CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    return

def saveABGraphMunalData(mc, date, states,hr,min):
    deleteABGraphManualData(mc, date)
    working_hr = 0
    day_hr = 0
    night_hr = 0
    stop_hr = 0
    empty_hr = 0
    if not hr:
        hr = 0
    if not min:
        min = 0
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[ABGraphManualData] ([WorkCenterNo],[Date],[DateTimeStamp],[H00],[H01],[H02],[H03],[H04],[H05],[H06],[H07],[H08],[H09],[H10],[H11],[H12],[H13],[H14],[H15],[H16],[H17],[H18],[H19],[H20],[H21],[H22],[H23],[WorkingHour],[DayHour],[NightHour],[StopHour],[EmptyHour],[Hour],[Min]) VALUES"
    sql += " ('"+str(mc)+"','"+str(date)+"',CURRENT_TIMESTAMP"
    for st in states:
        sql += ",'"+str(st)+"'"
        if st == 'A':
            day_hr = day_hr + 1
            working_hr = working_hr + 1
        elif st == 'B':
            night_hr = night_hr + 1
            working_hr = working_hr + 1
        elif st == 'S':
            stop_hr = stop_hr + 1
        elif st == 'N':
            empty_hr = empty_hr + 1
    sql += ","+str(working_hr)+","+str(day_hr)+","+str(night_hr)+","+str(stop_hr)+","+str(empty_hr)+","+str(hr)+","+str(min)+")"
    # print(sql)
    cursor.execute(sql)
    conn.commit()
    return

def getInsertToolHeader(drawing_no, drawing_rev, part_name, programmer, tool_list_no, tool_list_rev, customer, file_location, fg_matcode, fg_drawing, order_no, operation_no, wc_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[ToolHeader] ([DrawingNo],[DrawingRev],[PartName],[Programmer],[ToolListNo],[ToolListRev],[Customer],[FileLocation],[FG_MaterialCode],[FG_Drawing],[OrderNo],[OperationNo],[WorkCenterNo],[DateTimeStamp]) VALUES"
    sql += f" ('{drawing_no}','{drawing_rev}','{part_name}','{programmer}','{tool_list_no}','{tool_list_rev}','{customer}','{file_location}','{fg_matcode}','{fg_drawing}','{order_no}','{operation_no}','{wc_no}',CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    sql = "SELECT SCOPE_IDENTITY()"
    cursor.execute(sql)
    result = cursor.fetchall()
    if(len(result) == 0):
        return None
    return result[0][0]

def InsertToolStep(tool_header_id, step_order, step_name, operation, remark):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[ToolStep] ([ToolHeaderID],[StepOrder],[StepName],[Operation],[Remark],[DateTimeStamp]) VALUES"
    sql += f" ({tool_header_id},{step_order},'{step_name}','{operation}','{remark}',CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    return

def InsertToolItem(tool_header_id, no, ct_code, tool_no, tool_life_qty, tool_life_min, remark, confirmed_qty):
    tool_life_qty = tool_life_qty if tool_life_qty else 0
    tool_life_min = tool_life_min if tool_life_min else 0
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[ToolItem] ([ToolHeaderID],[No],[CTCode],[ToolNo],[ToolLifeQty],[ToolLifeMin],[Remark],[ConfirmedQty],[DateTimeStamp]) VALUES"
    sql += f" ({tool_header_id},{no},'{ct_code}','{tool_no}',{tool_life_qty},{tool_life_min},'{remark}',{confirmed_qty},CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()

def InsertToolMaster(ct_code, type, material, dia, flute, ma_code, proj_len, flute_len):
    type = type if type else ''
    material = material if material else ''
    dia = dia if dia else 0
    flute = flute if flute else 0
    ma_code = ma_code if ma_code else ''
    proj_len = proj_len if proj_len else 0
    flute_len = flute_len if flute_len else 0
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[ToolMaster] ([CTCode],[Type],[Material],[Dia],[Flute],[MACode],[ProjLen],[FluteLen],[DateTimeStamp]) VALUES"
    sql += f" ('{ct_code}','{type}','{material}',{dia},{flute},'{ma_code}',{proj_len},{flute_len},CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()

def insertHistoryTool(tooli_id, no, ct_code, tool_no, confirmed_qty, tool_life_qty, tool_life_min, remark, inactive_reason, fg_matcode, order_no, operation_no, wc_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO [dbo].[HistoryTool] ([ToolItemID],[No],[CTCode],[ToolNo],[ConfirmedQty],[ToolLifeQty],[ToolLifeMin],[Remark],[InActiveReason],[FG_MaterialCode],[OrderNo],[OperationNo],[WorkCenterNo],[DateTimeStamp])"
    sql += f" VALUES ({tooli_id},{no},'{ct_code}','{tool_no}',{confirmed_qty},{tool_life_qty},{tool_life_min},'{remark}','{inactive_reason}','{fg_matcode}','{order_no}','{operation_no}','{wc_no}',CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    return

#------------------------------------------------------------------------ UPDATE

def updateOperatingWorkCenter(id, status):
    conn = get_connection()
    cursor = conn.cursor()
    sql = ""
    if status == "WAITING":
        sql = "UPDATE [OperatingWorkCenter] SET [StartDateTime] = NULL, [StopDateTime] = NULL, [Status] = 'WAITING', IdleCode = NULL WHERE OperatingWorkCenterID = " + str(id)
    if status == "WORKING":
        sql = "UPDATE [OperatingWorkCenter] SET [StartDateTime] = CURRENT_TIMESTAMP, [StopDateTime] = NULL, [Status] = 'WORKING' WHERE OperatingWorkCenterID = " + str(id)
    if status == "SETUP":
        sql = "UPDATE [OperatingWorkCenter] SET [StartDateTime] = CURRENT_TIMESTAMP, [StopDateTime] = NULL, [Status] = 'SETUP' WHERE OperatingWorkCenterID = " + str(id)
    if status == "IDLE":
        sql = "UPDATE [OperatingWorkCenter] SET [StartDateTime] = CURRENT_TIMESTAMP, [StopDateTime] = NULL, [Status] = 'IDLE' WHERE OperatingWorkCenterID = " + str(id)
    if status == "COMPLETE":
        sql = "UPDATE [OperatingWorkCenter] SET [StopDateTime] = CURRENT_TIMESTAMP, [Status] = 'COMPLETE' WHERE OperatingWorkCenterID = " + str(id)
    cursor.execute(sql)
    conn.commit()
    return

def updateOperatingOperator(id, status):
    conn = get_connection()
    cursor = conn.cursor()
    sql = ""
    if status == "WAITING":
        sql = "UPDATE [OperatingOperator] SET [StartDateTime] = NULL, [StopDateTime] = NULL, [Status] = 'WAITING' WHERE OperatingOperatorID = " + str(id)
    if status == "WORKING":
        sql = "UPDATE [OperatingOperator] SET [StartDateTime] = CURRENT_TIMESTAMP, [StopDateTime] = NULL, [Status] = 'WORKING' WHERE OperatingOperatorID = " + str(id)
    if status == "SETUP":
        sql = "UPDATE [OperatingOperator] SET [StartDateTime] = CURRENT_TIMESTAMP, [StopDateTime] = NULL, [Status] = 'SETUP' WHERE OperatingOperatorID = " + str(id)
    if status == "COMPLETE":
        sql = "UPDATE [OperatingOperator] SET [StopDateTime] = CURRENT_TIMESTAMP, [Status] = 'COMPLETE' WHERE OperatingOperatorID = " + str(id)
    cursor.execute(sql)
    conn.commit()
    return

def updateIdleCodeOperatingWorkCenter(id, code):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OperatingWorkCenter] SET IdleCode = '"+str(code)+"' WHERE OperatingWorkCenterID = " + str(id)
    cursor.execute(sql)
    conn.commit()
    return

def updateOrderControl(order_no, status):
    conn = get_connection()
    cursor = conn.cursor()
    sql = ""
    if status == "START":
        sql = "UPDATE [OrderControl] SET [ProcessStart] = CURRENT_TIMESTAMP WHERE OrderNo = '" + order_no + "'"
    if status == "STOP":
        sql = "UPDATE [OrderControl] SET [ProcessStop] = CURRENT_TIMESTAMP WHERE OrderNo = '" + order_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def updateOperationControl(order_no, operation_no, accept, reject, status):
    conn = get_connection()
    cursor = conn.cursor()
    sql = ""
    if status == "START":
        sql = "UPDATE [OperationControl] SET [ProcessStart] = CURRENT_TIMESTAMP WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    if status == "STOP":
        sql = "UPDATE [OperationControl] SET [ProcessStop] = CURRENT_TIMESTAMP WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    if status == "UPDATEQTY":
        sql = "UPDATE [OperationControl] SET [AcceptedQty] += " + str(accept) + ", [RejectedQty] += " + str(reject) + " WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no  + "'"
    if status == "PROCESSQTY":
        sql = "UPDATE [OperationControl] SET [ProcessQty] += " + str(accept) + " WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def changeOperationControl(order_no, operation_no, work_center_no, plan_start_date, plan_finish_date, est_setup_time, est_operate_time, est_labor_time):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OperationControl] SET WorkCenterNo = '"+ str(work_center_no) + "', PlanStartDate = '" + str(plan_start_date) + "', PlanFinishDate = '" + str(plan_finish_date) + "', EstSetupTime = '" + str(est_setup_time) + "', EstOperationTime = '" + str(est_operate_time) + "', EstLaborTime = '" + str(est_labor_time) + "' WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def joinProcess(order_no, operation_no, join_order_no, join_operation_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OperationControl] SET [JoinToOrderNo] = '" + order_no + "', [JoinToOperationNo] = '" + operation_no + "', [JoinStartDateTime] = CURRENT_TIMESTAMP"
    sql += " WHERE OrderNo = '" + join_order_no + "' AND OperationNo = '" + join_operation_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def joinProcessRemove(order_no, operation_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OperationControl] SET [JoinToOrderNo] = NULL, [JoinToOperationNo] = NULL, [JoinStartDateTime] = NULL WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def changeUserPassword(user_id, user_password):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [User] SET [PasswordHash] = HASHBYTES('SHA2_512', '"+user_password+"') WHERE UserID = '"+user_id+"'"
    cursor.execute(sql)
    conn.commit()
    return

def updateOrderLotNo(order_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OrderControl] SET [LotNo] += 1 WHERE OrderNo = '" + order_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def updateManualReportAllowdance(status):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [AdminConfig] SET [Value] = '"+status+"' WHERE KeyText = 'MANUAL_REPORT_ALLOWDANCE'"
    cursor.execute(sql)
    conn.commit()
    return

def increaseQty(order_no, operation_no, amount):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OrderControl] SET [ProductionOrderQuatity] += '" + amount + "' WHERE OrderNo = '" + order_no + "'"
    cursor.execute(sql)
    conn.commit()
    sql = "UPDATE [OperationControl] SET [ProcessQty] += '" + amount + "' WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def decreaseQty(order_no, operation_no, amount):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OrderControl] SET [ProductionOrderQuatity] -= '" + amount + "' WHERE OrderNo = '" + order_no + "'"
    cursor.execute(sql)
    conn.commit()
    sql = "UPDATE [OperationControl] SET [ProcessQty] -= '" + amount + "' WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def updateOrderNote(order_no, note, weight, size):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OrderControl] SET [Note] = '"+note+"',[Weight] = '"+str(weight)+"',[Size] = '"+str(size)+"' WHERE OrderNo = '"+order_no+"'"
    cursor.execute(sql)
    conn.commit()
    return

def updateOperationNote(order_no, operation_no, note):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OperationControl] SET [Note] = '"+note+"' WHERE OrderNo = '"+order_no+"' AND OperationNo = '"+operation_no+"'"
    cursor.execute(sql)
    conn.commit()
    return

def updateOverEstNote(order_no,  operation_no, note):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OperationControl] SET [OverEstNote] = '"+ str(note) +"' WHERE OrderNo = '"+ str(order_no) +"' AND OperationNo = '" + str(operation_no) + "'"
    cursor.execute(sql)
    conn.commit()
    return

def setWorkCenterTarget(wc_no, target_hour):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE WorkCenter SET Target = "+target_hour+" WHERE WorkCenterNo = '"+wc_no+"'"
    cursor.execute(sql)
    conn.commit()
    return

def setWorkCenterCapacity(wc_no, cap_hour):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE WorkCenter SET Capacity = "+cap_hour+" WHERE WorkCenterNo = '"+wc_no+"'"
    cursor.execute(sql)
    conn.commit()
    return

def updateMailDate(date):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [AdminConfig] SET [Value] = '"+date+"' WHERE KeyText = 'MAIL_DATE'"
    cursor.execute(sql)
    conn.commit()
    return

def fixRMMaterialCode(order_no, rm_mat_code):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [OrderControl] SET RM_MaterialCode = '"+str(rm_mat_code)+"' WHERE OrderNo = '"+str(order_no)+"'"
    cursor.execute(sql)
    conn.commit()
    return

def fixOvertimeReported(emp_id):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE OvertimeOperator SET isFixed = 1 WHERE EmpID = '"+str(emp_id)+"'"
    cursor.execute(sql)
    conn.commit()
    return

def remarkPLNFAI(order_no, remark):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "UPDATE [PendingPLNFAI] SET Remark = '"+remark+"' WHERE OrderNo = '"+order_no+"'"
    cursor.execute(sql)
    conn.commit()
    return

def increaseHistoyToolConfirm(tooli_id, order_no, operation_no, confirmed_qty):
    conn = get_connection()
    cursor = conn.cursor()
    sql = f"UPDATE HistoryTool SET ConfirmedQty = ConfirmedQty + {confirmed_qty} WHERE ToolItemID = {tooli_id} AND OrderNo = '{order_no}' AND OperationNo = '{operation_no}'"
    cursor.execute(sql)
    conn.commit()
    return

def transferToolHeader(toolh_id, order_no, operation_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = f"UPDATE ToolHeader SET OrderNo = '{order_no}', OperationNo = '{operation_no}' WHERE ID = {toolh_id}"
    cursor.execute(sql)
    conn.commit()
    return

def updateToolItemConfirm(tooli_id, confirmed_qty):
    conn = get_connection()
    cursor = conn.cursor()
    sql = f"UPDATE ToolItem SET ConfirmedQty = ConfirmedQty + {confirmed_qty} WHERE ID = {tooli_id}"
    cursor.execute(sql)
    conn.commit()
    return

def changeToolHeaderWorkCenter(toolh_id, wc_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = f"UPDATE ToolHeader SET WorkCenterNo = '{wc_no}' WHERE ID = {toolh_id}"
    cursor.execute(sql)
    conn.commit()
    return
#------------------------------------------------------------------------ DELETE

def deleteOperatingWorkCenter(id):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [OperatingWorkCenter] WHERE OperatingWorkCenterID = " + str(id)
    cursor.execute(sql)
    conn.commit()
    return

def deleteOperatingOperator(id):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [OperatingOperator] WHERE OperatingOperatorID = " + str(id)
    cursor.execute(sql)
    conn.commit()
    return

def deleteAllOperatingData(order_no, operation_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [OperatingOperator] WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    conn.commit()
    cursor = conn.cursor()
    sql = "DELETE FROM [OperatingWorkCenter] WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def deleteOperationControl(order_no, operation_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [OperationControl] WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def deleteSFR2SAP_Modifier_Change(order_no, operation_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [SFR2SAP_Modifier] WHERE OrderNo = '" + order_no + "' AND OperationNo = '" + operation_no + "' AND (Mode = 22 OR Mode = 21) AND CAST(DateTimeStamp AS DATE) = CAST(CURRENT_TIMESTAMP AS DATE) AND DATEPART(HOUR, DateTimeStamp) = DATEPART(HOUR, CURRENT_TIMESTAMP)"
    cursor.execute(sql)
    conn.commit()
    return

def deleteUser(user_id):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [User] WHERE UserID = '"+user_id+"'"
    cursor.execute(sql)
    conn.commit()
    return

def deleteOrderControl(order_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [OrderControl] WHERE OrderNo = '" + order_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def deleteEmpAtComputer(emp_id):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [EmpAtComputer] WHERE EmpID = '" + str(emp_id) + "'"
    cursor.execute(sql)
    conn.commit()
    return

def deleteAllSFRAndSAPOrder(order_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [OperationControl] WHERE OrderNo = '" + order_no + "'"
    cursor.execute(sql)
    conn.commit()
    sql = "DELETE FROM [OrderControl] WHERE OrderNo = '" + order_no + "'"
    cursor.execute(sql)
    conn.commit()
    sql = "DELETE FROM [SAP_Routing] WHERE ProductionOrderNo = '" + order_no + "'"
    cursor.execute(sql)
    conn.commit()
    sql = "DELETE FROM [SAP_Order] WHERE ProductionOrderNo = '" + order_no + "'"
    cursor.execute(sql)
    conn.commit()
    return

def clearPLNFAI(order_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [PendingPLNFAI] WHERE OrderNo = '"+order_no+"'"
    cursor.execute(sql)
    conn.commit()
    return

def deleteABGraphManualData(mc, date):
    conn = get_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM [ABGraphManualData] WHERE WorkCenterNo = '"+str(mc)+"' and Date = '"+str(date)+"'"
    cursor.execute(sql)
    conn.commit()
    return

def deleteTool(toolh_id):
    conn = get_connection()
    cursor = conn.cursor()
    sql = f"DELETE FROM [ToolItem] WHERE ToolHeaderID = {toolh_id}"
    cursor.execute(sql)
    conn.commit()
    sql = f"DELETE FROM [ToolStep] WHERE ToolHeaderID = {toolh_id}"
    cursor.execute(sql)
    conn.commit()
    sql = f"DELETE FROM [ToolHeader] WHERE ID = {toolh_id}"
    cursor.execute(sql)
    conn.commit()
    return

def deleteToolItem(tooli_id):
    conn = get_connection()
    cursor = conn.cursor()
    sql = f"DELETE FROM [ToolItem] WHERE ID = {tooli_id}"
    cursor.execute(sql)
    conn.commit()
    return

def deleteDoubleRecord(order_no, operation_no):
    conn = get_connection()
    cursor = conn.cursor()
    sql = f"""
            WITH CTE AS 
            ( SELECT *, ROW_NUMBER() OVER (PARTITION BY OrderNo, OperationNo, EmpID, WorkCenterNo, StartDateTime 
            ORDER BY StartDateTime) AS RowNum 
            FROM HistoryOperate WHERE OrderNo = '{order_no}' AND OperationNo = '{operation_no}' AND Type <> 'MANUAL' 
            ) 
            DELETE FROM CTE WHERE RowNum > 1;

            WITH CTE AS 
            ( SELECT *, ROW_NUMBER() OVER (PARTITION BY ProductionOrderNo, OperationNumber, EmployeeID, WorkCenter, StartDate, StartTime
            ORDER BY StartDate, StartTime) AS RowNum 
            FROM SFR2SAP_Report WHERE ProductionOrderNo = '{order_no}' AND OperationNumber = '{operation_no}' AND Yiled = 0 AND Scrap = 0
            ) 
            DELETE FROM CTE WHERE RowNum > 1;
        """
    cursor.execute(sql)
    conn.commit()
    return

################################################################################
################################################################################
################################################################################

def getClientIP(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[-1].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def printString(str):
    print("################################################################################")
    print(str)
    print("################################################################################")
    return

def get_list_for_sql(list):
    if not list:
        return '()'
    else:
        str = '('
        for idx, item in enumerate(list):
            if idx == 0:
                str += '\'' + item + '\''
            else:
                str += ',\'' + item + '\''
        str += ')'
    return str

def generate_random_token(length):
    # Define the characters used for the token
    characters = string.ascii_letters + string.digits
    # Generate the random token
    token = ''.join(secrets.choice(characters) for _ in range(length))
    return token

def frontZero(str, length):
    result = str
    for i in range(length - len(str)):
        result = "0" + result
    return result

def get_day_count(month, year):
    if int(year)%4 == 0 and month == "02":
        return 29
    elif  month == "02":
        return 28
    elif month == "01" or month == "03" or month == "05" or month == "07" or month == "08" or month == "10" or month == "12":
        return 31
    return 30

def get_isoweekdays(size, month, year):
    isoweekdays = [False] * size
    for i in range(size):
        day = frontZero(str(i + 1), 2)
        month = frontZero(month, 2)
        tmp = str(day) + '-' + str(month) + '-' + year
        date = datetime.strptime(tmp, '%d-%m-%Y')
        if date.isoweekday() == 7: # Sunday
            isoweekdays[i] = True
    return isoweekdays

def get_week_rej_per_info(isoweekdays, pro_qtys, rej_qtys):
    week_titles = []
    week_pros = []
    week_rejs = []
    week_pers = []
    week_pro = 0
    week_rej = 0
    week_per = 0
    start_idx = 0
    for idx, is_week in enumerate(isoweekdays):
        week_pro = week_pro + (pro_qtys[idx] if pro_qtys[idx] else 0)
        week_rej = week_rej + (rej_qtys[idx] if rej_qtys[idx] else 0)
        if not is_week:
            week_titles.append(None)
            week_pros.append(0)
            week_rejs.append(0)
            week_pers.append(0)
        if is_week:
            print(week_rej)
            if week_pro > 0:
                week_per = round(((week_rej / week_pro) * 100), 2)
            if(start_idx != idx): 
                week_titles.append(str(start_idx + 1) + " - " + str(idx + 1))
            else:
                week_titles.append(str(idx + 1))
            week_pros.append(week_pro)
            week_rejs.append(week_rej)
            week_pers.append(week_per)
            week_pro = 0
            week_rej = 0
            week_per = 0
            start_idx = idx + 1
    return week_titles, week_pros, week_rejs, week_pers

def get_date_between_week(fweek):
    year = int(fweek[0:4])
    week = int(fweek[6:len(fweek)])
    start_date = datetime(year, 1, 1)
    while start_date.weekday() != 0:
        start_date = start_date + relativedelta(days=+1)
    start_date = start_date + relativedelta(weeks=+week) + relativedelta(days=-7)
    end_date = start_date + relativedelta(days=+6)
    return start_date, end_date

def is_previous_month(month, year):
    # Convert the month and year strings to integers with base 10
    month = int(month)
    year = int(year)
    
    # Get the current month and year
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Calculate the previous month and year from the current month and year
    previous_month, previous_year = (current_month - 1, current_year) if current_month > 1 else (12, current_year - 1)
    
    # Compare the input month and year with the previous month and year, and all months and years before that
    while year < current_year or (year == current_year and month < current_month):
        if (month, year) == (previous_month, previous_year):
            return True
        previous_month, previous_year = (previous_month - 1, previous_year) if previous_month > 1 else (12, previous_year - 1)
        if previous_month == 12:
            year -= 1
    return False

def run_collect_ab_data(fmonth):
    year = fmonth[0:4]
    month = fmonth[5:7]
    # WorkCenterGroup
    wcgs = getMachineWorkCenterGroupList()
    for wcg in wcgs:
        fwcg = wcg.WorkCenterGroup
        wcs = getWorkCenterInGroupActiveList(fwcg, fmonth)
        is_collected = isCollectedABGraphData('WCG', fwcg, month, year)
        if not is_collected:
            print('Collecting AB Graph Data of WorkCenterGroup ', fwcg, ' on ', fmonth)
            #-- Work
            for rs in getMonthlyWorkCenterOperForABGraph('WCG', fwcg, fmonth):
                saveABGraphData('WCG',fwcg,'W', rs.Foper, rs.Fday, month, year)
            #-- Setup
            for rs in getMonthlyWorkCenterSetupForABGraph('WCG', fwcg, fmonth):
                saveABGraphData('WCG',fwcg,'S', rs.Fsetup, rs.Fday, month, year)
        # Machine
        for wc in wcs:
            fwc = wc.WorkCenterNo
            is_collected = isCollectedABGraphData('WC', fwc, month, year)
            if not is_collected:
                print('Collecting AB Graph Data of Machine ', fwc, ' on ', fmonth)
                #-- Work
                for rs in getMonthlyWorkCenterOperForABGraph('WC', fwc, fmonth):
                    saveABGraphData('WC',fwc,'W', rs.Foper, rs.Fday, month, year)
                #-- Setup
                for rs in getMonthlyWorkCenterSetupForABGraph('WC', fwc, fmonth):
                    saveABGraphData('WC',fwc,'S', rs.Fsetup, rs.Fday, month, year)
    # Routing
    rts = getOnRoutingList()
    for rt in rts:
        frt = rt.WorkCenterNo
        wcs = getWorkCenterOnRoutingActiveList(frt, fmonth)
        is_collected = isCollectedABGraphData('RT', frt, month, year)
        if not is_collected:
            print('Collecting AB Graph Data of Routing ', frt, ' on ', fmonth)
            #-- Work
            for rs in getMonthlyWorkCenterOperForABGraph('RT', frt, fmonth):
                saveABGraphData('RT',frt,'W', rs.Foper, rs.Fday, month, year)
            #-- Setup
            for rs in getMonthlyWorkCenterSetupForABGraph('RT', frt, fmonth):
                saveABGraphData('RT',frt,'S', rs.Fsetup, rs.Fday, month, year)
        for wc in wcs:
            fwc = wc.WorkCenterNo
            is_collected = isCollectedABGraphData('WC', fwc, month, year)
            print('Collecting AB Graph Data of Machine ', fwc, ' on ', fmonth)
            if not is_collected:
                #-- Work
                for rs in getMonthlyWorkCenterOperForABGraph('WC', fwc, fmonth):
                    saveABGraphData('WC',fwc,'W', rs.Foper, rs.Fday, month, year)
                #-- Setup
                for rs in getMonthlyWorkCenterSetupForABGraph('WC', fwc, fmonth):
                    saveABGraphData('WC',fwc,'S', rs.Fsetup, rs.Fday, month, year)
    return

def update_employee_master():
    wb = load_workbook(filename = 'media/Employee.xlsx')
    ws = wb.active
    skip_count = 2
    row_count = 0
    new_emp_count = 0
    update_emp_count = 0
    error_emp_count = 0
    print("#########################################")
    print("SFR Employee Update Result")
    print("#########################################")
    for i in range(ws.max_row + 1):
        if i < skip_count:
            continue
        emp_id = "" if ws['A' + str(i)].value == None else int(str(ws['A' + str(i)].value).strip())
        emp_name = "" if ws['B' + str(i)].value == None else ws['B' + str(i)].value
        employment_type = "" if ws['J' + str(i)].value == None else ws['J' + str(i)].value
        profitcenter = "" if ws['K' + str(i)].value == None else ws['K' + str(i)].value
        section = "" if ws['L' + str(i)].value == None else ws['L' + str(i)].value
        section = section.replace('S-','')
        costcenter = "" if ws['M' + str(i)].value == None else ws['M' + str(i)].value
        position = "" if ws['N' + str(i)].value == None else ws['N' + str(i)].value
        jobfunc = "" if ws['O' + str(i)].value == None else ws['O' + str(i)].value
        jobfunc = jobfunc[0:50]
        is_active = 1 if ws['P' + str(i)].value == 1 or ws['P' + str(i)].value == '1' or ws['P' + str(i)].value == 'Active' else 0
        is_active_txt = 'Active' if is_active == 1 else 'In-Active'
        if emp_id and emp_id != "":
            isExist = isExistOperator(str(emp_id))
            if isExist:
                emp = getEmpIDByUserID(emp_id)
                if emp.EmpName != emp_name or emp.Section != section or int(emp.CostCenter) != costcenter or emp.IsActive != is_active or emp.EmploymentType != employment_type or position != emp.Position or emp.ProfitCenter != profitcenter or emp.JobFunction != jobfunc:
                    conn = get_connection()
                    cursor = conn.cursor()
                    sql = f"""UPDATE Employee SET EmpName = '{emp_name}', Section = '{section}', CostCenter = '{costcenter}', IsActive = {str(is_active)}, EmploymentType = '{employment_type}', Position = '{position}', ProfitCenter = '{profitcenter}', JobFunction = '{jobfunc}' WHERE EmpID = '{str(emp_id)}'"""
                    cursor.execute(sql)
                    conn.commit()
                    update_emp_count = update_emp_count + 1
                    print(f'# UPDATE : {emp_id}')
                    if(emp.EmpName != emp_name):
                        print(f'--- Name : {emp.EmpName} => {emp_name}')
                    if(emp.Section != section):
                        print(f'--- Section : {emp.Section} => {section}')
                    if(int(emp.CostCenter) != costcenter):
                        print(f'--- Cost Center : {emp.CostCenter} => {costcenter}')
                    if(emp.ProfitCenter != profitcenter):
                        print(f'--- Profit Center : {emp.ProfitCenter} => {profitcenter}')
                    if(emp.EmploymentType != employment_type):
                        print(f'--- Employement Type : {emp.EmploymentType} => {employment_type}')
                    if(emp.Position != position):
                        print(f'--- Position : {emp.Position} => {position}')
                    if(emp.JobFunction != jobfunc):
                        print(f'--- Job Function : {emp.JobFunction} => {jobfunc}')
                    if(emp.IsActive != is_active):
                        is_active_tmp = 'Active' if emp.IsActive else 'In-Active'
                        print(f'--- Is Active : {is_active_tmp} => {is_active_txt}')
            if not isExist:
                conn = get_connection()
                cursor = conn.cursor()
                sql = f"""INSERT INTO Employee (EmpID,EmpName,Section,CostCenter,IsActive,EmploymentType,Position,ProfitCenter,JobFunction) VALUES ('{str(emp_id)}','{emp_name}','{section}','{costcenter}',{str(is_active)},'{employment_type}','{position}','{profitcenter}','{jobfunc}')"""
                cursor.execute(sql)
                conn.commit()
                new_emp_count = new_emp_count + 1
                print(f'# NEW : {emp_id}')
                print(f'--- Name : {emp_name}')
                print(f'--- Section : {section}')
                print(f'--- Cost Center : {costcenter}')
                print(f'--- Profit Center : {profitcenter}')
                print(f'--- Employement Type : {employment_type}')
                print(f'--- Position : {position}')
                print(f'--- Job Function : {jobfunc}')
                is_active_tmp = 'Active' if is_active == 1 else 'In-Active'
                print(f'--- Is Active : {is_active_tmp}')
        else:
            error_emp_count = error_emp_count + 1
        row_count = row_count + 1
    print("#########################################")
    print("Data Records :", str(row_count))
    print("New Records :", str(new_emp_count))
    print("Update Records :", str(update_emp_count))
    print("Error Records :", str(error_emp_count))
    print("#########################################")
    return

def update_estimate():
    wb = load_workbook(filename = 'media/RTUP.xlsx')
    ws = wb.active
    skip_count = 2 + 69990
    for i in range(ws.max_row + 1):
        if i < skip_count:
            continue
        order_no = ws['A' + str(i)].value
        operation_no = ws['B' + str(i)].value
        setup = ws['C' + str(i)].value
        oper = ws['D' + str(i)].value
        labor = ws['E' + str(i)].value
        operation_no = str(operation_no)
        while len(operation_no) < 4:
            operation_no = '0' + operation_no
        print(i, order_no, operation_no)
        if not order_no or not operation_no:
            continue
        conn = get_connection()
        cursor = conn.cursor()
        sql = f"UPDATE OperationControl SET EstSetupTime = {setup}, EstOperationTime = {oper}, EstLaborTime = {labor} WHERE OrderNo = '{order_no}' AND OperationNo = '{operation_no}'"
        cursor.execute(sql)
        conn.commit()
        conn = get_connection()
        cursor = conn.cursor()
        sql = f"UPDATE SAP_Routing SET EstimateSetTime = {setup}, EstimateOperationTime = {oper}, EstimateLaborTime = {labor} WHERE ProductionOrderNo = '{order_no}' AND OperationNumber = '{operation_no}'"
        cursor.execute(sql)
        conn.commit()
    return

def update_target_cycle_time():
    wb = load_workbook(filename = 'media/target_cycle_time.xlsx')
    ws = wb.active
    skip_count = 3
    stop_count = 999
    for i in range(ws.max_row + 1):
        if i < skip_count:
            continue
        if i > stop_count:
            break
        fg_materialcode = ws['E' + str(i)].value
        mca_3_axis = round(ws['G' + str(i)].value) if ws['G' + str(i)].value else None
        hmc = round(ws['H' + str(i)].value) if ws['H' + str(i)].value else None
        tma = round(ws['I' + str(i)].value) if ws['I' + str(i)].value else None
        mca_5_axis = round(ws['J' + str(i)].value) if ws['J' + str(i)].value else None
        cla = round(ws['K' + str(i)].value) if ws['K' + str(i)].value else None
        bta = round(ws['L' + str(i)].value) if ws['L' + str(i)].value else None
        ig = round(ws['M' + str(i)].value) if ws['M' + str(i)].value else None
        og = round(ws['N' + str(i)].value) if ws['N' + str(i)].value else None
        eda = round(ws['O' + str(i)].value) if ws['O' + str(i)].value else None
        wca = round(ws['P' + str(i)].value) if ws['P' + str(i)].value else None
        cln_as = round(ws['Q' + str(i)].value) if ws['Q' + str(i)].value else None
        tbl_as = round(ws['R' + str(i)].value) if ws['R' + str(i)].value else None
        iqc_as = round(ws['S' + str(i)].value) if ws['S' + str(i)].value else None
        fpi = round(ws['T' + str(i)].value) if ws['T' + str(i)].value else None
        mpi = round(ws['U' + str(i)].value) if ws['U' + str(i)].value else None
        stc = round(ws['V' + str(i)].value) if ws['V' + str(i)].value else None
        std = round(ws['W' + str(i)].value) if ws['W' + str(i)].value else None
        stm = round(ws['X' + str(i)].value) if ws['X' + str(i)].value else None
        sth = round(ws['Y' + str(i)].value) if ws['Y' + str(i)].value else None
        idhen_oem = round(ws['Z' + str(i)].value) if ws['Z' + str(i)].value else None
        msm_oem = round(ws['AA' + str(i)].value) if ws['AA' + str(i)].value else None
        vib_oem = round(ws['AB' + str(i)].value) if ws['AB' + str(i)].value else None
        str_oem = round(ws['AC' + str(i)].value) if ws['AC' + str(i)].value else None
        str_vib = round(ws['AD' + str(i)].value) if ws['AD' + str(i)].value else None
        assy_oem = round(ws['AE' + str(i)].value) if ws['AE' + str(i)].value else None
        pnt_oem = round(ws['AF' + str(i)].value) if ws['AF' + str(i)].value else None
        ink_oem = round(ws['AG' + str(i)].value) if ws['AG' + str(i)].value else None
        stp = round(ws['AH' + str(i)].value) if ws['AH' + str(i)].value else None
        ls_as = round(ws['AI' + str(i)].value) if ws['AI' + str(i)].value else None
        fqc_as = round(ws['AJ' + str(i)].value) if ws['AJ' + str(i)].value else None
        pk_as = round(ws['AK' + str(i)].value) if ws['AK' + str(i)].value else None
        
        keys = ['MCA 3 Axis','HMC','TMA','5 Axis','CLA','BTA','IG','OG','EDA','WCA','CLN_AS','TBL_AS',
                'IQC_AS','FPI','MPI','STC','STD','STM','STH','IDHEN_OEM','MSM_OEM','VIB_OEM','STR_OEM',
                'STR_VIB','ASSY_OEM','PNT_OEM','INK_OEM','STP','LS_AS','FQC_AS','PK_AS']
        values = [mca_3_axis, hmc, tma, mca_5_axis, cla, bta, ig, og, eda, wca, cln_as, tbl_as,
                  iqc_as, fpi, mpi, stc, std, stm, sth, idhen_oem, msm_oem, vib_oem, str_oem,
                  str_vib, assy_oem, pnt_oem, ink_oem, stp, ls_as, fqc_as, pk_as]
        
        for idx, key in enumerate(keys):
            value = values[idx]
            cmd = 'Skip :'
            cursor = get_connection().cursor()
            sql = f"SELECT * FROM [CycleTimeTarget] WHERE fg_materialcode = '{fg_materialcode}' AND MainProcess = '{key}'"
            cursor.execute(sql)
            if (len(cursor.fetchall()) > 0):
                if value: # update
                    cmd = 'Update :'
                    conn = get_connection()
                    cursor = conn.cursor()
                    sql = f"UPDATE CycleTimeTarget SET TargetValue = {value} WHERE fg_materialcode = '{fg_materialcode}' AND MainProcess = '{key}'"
                    cursor.execute(sql)
                    conn.commit()
                else: # delete
                    cmd = 'Delete :'
                    conn = get_connection()
                    cursor = conn.cursor()
                    sql = f"DELETE FROM CycleTimeTarget WHERE fg_materialcode = '{fg_materialcode}' AND MainProcess = '{key}'"
                    cursor.execute(sql)
                    conn.commit()
            elif value: # insert
                cmd = 'Insert :'
                conn = get_connection()
                cursor = conn.cursor()
                sql = f"INSERT INTO [dbo].[CycleTimeTarget] ([FG_MaterialCode],[MainProcess],[TargetValue]) VALUES ('{fg_materialcode}','{key}',{value})"
                cursor.execute(sql)
                conn.commit()
            if cmd != 'Skip :':
                print(i, cmd, fg_materialcode, key, value)
    return    

def run_upload_tool_list(is_validating, fg_matcode, fg_drawing, order_no, operation_no, wc_no, file, sheets):
    wb = load_workbook(filename = f'media/tool_list/{file}')
    sh = wb.worksheets
    print("#########################################")
    print(f"Start : Upload Tool List , Is Validating : {is_validating}")
    print("#########################################")
    tool_header_id = None
    step_order = 0
    confirmed_qty = 0
    ##### FILE TYPE 1 -> MCA ?
    is_type_1 = run_upload_tool_file_type_1(is_validating, fg_matcode, fg_drawing, order_no, operation_no, wc_no, sheets, sh, tool_header_id, step_order, confirmed_qty)
    if is_type_1:
        return True
    ##### FILE TYPE 2 -> BTA & CLA
    is_type_2 = run_upload_tool_file_type_2(is_validating, fg_matcode, fg_drawing, order_no, operation_no, wc_no, sheets, sh, tool_header_id, confirmed_qty)
    if is_type_2:
        return True
    print("#########################################")
    print("Stop : Upload Tool Life")
    print("#########################################")
    return False

def run_upload_tool_file_type_1(is_validating, fg_matcode, fg_drawing, order_no, operation_no, wc_no, sheets, sh, tool_header_id, step_order, confirmed_qty):
    try:
    # if True:
        print("-------------- Header --------------")
        for idx, ws in enumerate(sh):
            drawing_no = ws['C4'].value[1:] 
            drawing_rev = ws['C5'].value[1:]  
            part_name = ws['C6'].value[1:] 
            programmer = ws['C7'].value[1:]  
            tool_list_no = ws['I4'].value[1:]  
            tool_list_rev = ws['I5'].value[1:]  
            customer = ws['I6'].value[1:] 
            file_location = ws['C9'].value
            if not is_validating:
                tool_header_id = getInsertToolHeader(drawing_no, drawing_rev, part_name, programmer, tool_list_no, tool_list_rev, customer, file_location, fg_matcode, fg_drawing, order_no, operation_no, wc_no) 
            print(drawing_no, drawing_rev, part_name, programmer, tool_list_no, tool_list_rev, customer, file_location, order_no, operation_no)
            break
        print("-------------- Step --------------")
        for idx, ws in enumerate(sh):
            if str(idx) not in sheets:
                continue
            skip_count = 13
            old_operation = ''
            old_remark = ''
            for i in range(ws.max_row + 1):
                if i < skip_count:
                    continue
                step_name = ws['A' + str(i)].value
                operation = ws['E' + str(i)].value if ws['E' + str(i)].value else old_operation
                remark = ws['I' + str(i)].value if ws['I' + str(i)].value else old_remark
                if not step_name: # case of no step
                    break
                old_operation = operation
                old_remark = remark
                if not is_validating:
                    InsertToolStep(tool_header_id, step_order, step_name, operation, remark)
                    step_order += 1
                print(step_name, operation, remark)
        print("--------------- Item ---------------")
        used_no = []
        for idx, ws in enumerate(sh):
            if str(idx) not in sheets:
                continue
            skip_count = 19
            for i in range(ws.max_row + 1):
                if i < skip_count:
                    continue
                no = ws['A' + str(i)].value
                tool_no = ''
                type = ws['B' + str(i)].value
                type.replace("'", "")
                ct_code = ws['C' + str(i)].value
                if ct_code and '/' in ct_code:
                    ct_code = ct_code.split('/')[0]
                material = ws['D' + str(i)].value
                dia = ws['E' + str(i)].value 
                flute = ws['F' + str(i)].value 
                ma_code = ws['G' + str(i)].value 
                proj_len = ws['H' + str(i)].value 
                flute_len = ws['I' + str(i)].value 
                tool_life_qty = ws['J' + str(i)].value 
                tool_life_min = ws['K' + str(i)].value
                remark = ws['L' + str(i)].value if ws['L' + str(i)].value else ''
                if not no: # case of no item
                    break
                if no in used_no: # skip same number
                    continue
                used_no.append(no)
                if not is_validating:
                    if not isExistToolMaster(ct_code):
                        InsertToolMaster(ct_code, type, material, dia, flute, ma_code, proj_len, flute_len)
                    InsertToolItem(tool_header_id, no, ct_code, tool_no, tool_life_qty, tool_life_min, remark, confirmed_qty)
                print(no, ct_code, type, material, dia, flute, ma_code, proj_len, flute_len, tool_life_qty, tool_life_min, remark)
    except Exception:
        print('Its not file type #1')
        return False
    return True

def run_upload_tool_file_type_2(is_validating, fg_matcode, fg_drawing, order_no, operation_no, wc_no, sheets, sh, tool_header_id, confirmed_qty):
    try:
    # if True:
        print("-------------- Header --------------")
        for idx, ws in enumerate(sh):
            drawing_no = ws['C4'].value
            drawing_rev = ''
            part_name = ''
            programmer = '' 
            tool_list_no = ws['C2'].value
            tool_list_rev = ''
            customer = ''
            file_location = ''
            if not is_validating:
                tool_header_id = getInsertToolHeader(drawing_no, drawing_rev, part_name, programmer, tool_list_no, tool_list_rev, customer, file_location, fg_matcode, fg_drawing, order_no, operation_no, wc_no) 
            print(drawing_no, drawing_rev, part_name, programmer, tool_list_no, tool_list_rev, customer, file_location, order_no, operation_no)
            break
        print("--------------- Item ---------------")
        used_no = []
        for idx, ws in enumerate(sh):
            if str(idx) not in sheets:
                continue
            skip_count = 9
            for i in range(ws.max_row + 1):
                if i < skip_count:
                    continue
                no = ws['A' + str(i)].value
                tool_no = ws['B' + str(i)].value
                type = ws['D' + str(i)].value # mat_desc
                type = type.replace("'", "")
                ct_code = ws['C' + str(i)].value
                if ct_code and '/' in ct_code:
                    ct_code = ct_code.split('/')[0]
                material = ''
                dia = 0 
                flute = 0
                ma_code = ''
                proj_len = 0
                flute_len = 0
                tool_life_qty = ws['E' + str(i)].value 
                tool_life_min = ws['F' + str(i)].value if ws['F' + str(i)].value != '-' else 0
                remark = ws['L' + str(i)].value if ws['L' + str(i)].value else ''
                if not no: # case of no item
                    break
                if no in used_no: # skip same number
                    continue
                used_no.append(no)
                if not is_validating:
                    if not isExistToolMaster(ct_code):
                        InsertToolMaster(ct_code, type, material, dia, flute, ma_code, proj_len, flute_len)
                    InsertToolItem(tool_header_id, no, ct_code, tool_no, tool_life_qty, tool_life_min, remark, confirmed_qty)
                print(no, ct_code, type, material, dia, flute, ma_code, proj_len, flute_len, tool_life_qty, tool_life_min, remark)
    except Exception:
        print('Its not file type #2 :')
        return False
    return True